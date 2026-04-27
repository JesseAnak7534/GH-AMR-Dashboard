"""
AMR Surveillance Dashboard for Multi-source Surveillance (Environment, Food, Human, Animal, Aquaculture)
Main Streamlit application with multi-page support.
"""
import streamlit as st
import pandas as pd
import os
import uuid
import bcrypt
import json
import logging
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
import plotly.express as px
import plotly.io as pio
import plotly.graph_objects as go

# ── Global Plotly theme ───────────────────────────────────────────────
# Every chart (Home page already transparent; this catches the other pages
# that previously rendered on Plotly's default white background) inherits a
# transparent paper/plot surface so the chart blends with the warm paper
# shell. Also sets the typography + colourway to match the editorial palette.
pio.templates["amrss"] = go.layout.Template(layout=dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter, -apple-system, BlinkMacSystemFont, sans-serif",
              color="#1a170e", size=12),
    colorway=["#194238", "#a27117", "#964517", "#2f5430", "#164161",
              "#741a10", "#b18a3a", "#5a8c6d"],
    xaxis=dict(gridcolor="#e2d7bb", linecolor="#b4a788",
               zerolinecolor="#e2d7bb", tickcolor="#b4a788",
               tickfont=dict(color="#58523e"), title=dict(font=dict(color="#1a170e"))),
    yaxis=dict(gridcolor="#e2d7bb", linecolor="#b4a788",
               zerolinecolor="#e2d7bb", tickcolor="#b4a788",
               tickfont=dict(color="#58523e"), title=dict(font=dict(color="#1a170e"))),
    hoverlabel=dict(bgcolor="#fbf7ec", bordercolor="#b4a788",
                    font=dict(family="Inter, sans-serif", color="#1a170e", size=12)),
    legend=dict(bgcolor="rgba(0,0,0,0)", bordercolor="rgba(0,0,0,0)",
                font=dict(color="#1a170e", size=11)),
    colorscale=dict(sequential=[[0, "#f3ead6"], [0.5, "#a27117"], [1, "#194238"]],
                    diverging=[[0, "#741a10"], [0.5, "#f3ead6"], [1, "#194238"]]),
    title=dict(font=dict(family="Fraunces, Georgia, serif",
                         color="#1a170e", size=16)),
))
pio.templates.default = "amrss"

logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

# Import modules
from src import db, validate, plots, report, analytics
from src import email_utils
from src.page_dashboard import render_dashboard_page
from src.lab_management import (
    get_lab_email_map,
    is_lab_user,
    KoboToolboxManager,
    kobo_submissions_to_frames,
    save_kobo_form_id,
    load_kobo_form_id
)
from src.page_pps import render_pps_page
from src.page_amu import render_amu_page
from src.page_amc import render_amc_page
from src.page_heatmap import render_heatmap_page
from src.page_pathogen_profile import render_pathogen_profile_page
from src.page_hai import render_hai_page

# Page configuration
st.set_page_config(
    page_title="ICBB-AMRSS",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize database (cached: runs once per process, not on every rerun --
# critical for hosted Postgres where ~50 DDL round-trips would otherwise
# happen on every interaction and lock the UI for many seconds)
@st.cache_resource(show_spinner="Connecting to surveillance database…")
def _init_db_once():
    try:
        db.init_database()
        return True
    except Exception:
        logger.exception("db.init_database() failed")
        return False

_init_db_once()


# ── Cloud-data seed: load the committed snapshot if the DB is empty ─────
# We can't reach Supabase from the maintainer's network to run a normal
# data migration, so we ship a small gzip-compressed snapshot of the
# production tables in the repo and replay it the first time the cloud
# app sees an empty `samples` table.  Idempotent: does nothing if
# samples already exist.
@st.cache_resource(show_spinner="Seeding initial surveillance data…")
def _seed_cloud_data_once():
    snapshot_path = Path("db") / "cloud_snapshot.json.gz"
    if not snapshot_path.exists():
        return False
    try:
        conn = db.get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM samples")
        existing = cur.fetchone()[0]
        if existing:
            conn.close()
            return False  # DB already has data; do nothing
        import gzip as _gz
        with _gz.open(snapshot_path, "rt", encoding="utf-8") as f:
            snapshot = json.load(f)
        for table, payload in snapshot.get("tables", {}).items():
            cols = payload.get("columns") or []
            rows = payload.get("rows") or []
            if not rows:
                continue
            placeholders = ", ".join(["%s"] * len(cols))
            collist = ", ".join(cols)
            sql = (
                f"INSERT INTO {table} ({collist}) VALUES ({placeholders}) "
                f"ON CONFLICT DO NOTHING"
            )
            data = [tuple(r.get(c) for c in cols) for r in rows]
            try:
                cur.executemany(sql, data)
                conn.commit()
                logger.info("seeded %s with %d rows", table, len(rows))
            except Exception:
                conn.rollback()
                logger.exception("seed insert failed for table %s", table)
        conn.close()
        return True
    except Exception:
        logger.exception("cloud-data seed failed")
        return False

_seed_cloud_data_once()

# ── Keep-alive: prevent idle WebSocket disconnection ────────────────────
st.markdown(
    """
    <script>
    // Ping the Streamlit server every 30 s to keep the WebSocket alive
    // and ping again as soon as the tab regains focus, since browsers
    // throttle setInterval when the tab is hidden which would otherwise
    // let the connection drop during a demo.
    (function keepAlive() {
        function ping() {
            try {
                fetch(window.location.href, {method: 'HEAD', cache: 'no-store'})
                    .catch(function(){});
            } catch (e) { /* swallow */ }
        }
        setInterval(ping, 30000);
        document.addEventListener('visibilitychange', function() {
            if (document.visibilityState === 'visible') { ping(); }
        });
    })();
    </script>
    """,
    unsafe_allow_html=True,
)

# Email verification via magic link is disabled

def _get_session_timeout_minutes():
    timeout_value = None
    try:
        if hasattr(st, "secrets") and "SESSION_TIMEOUT_MINUTES" in st.secrets:
            timeout_value = st.secrets.get("SESSION_TIMEOUT_MINUTES")
    except (FileNotFoundError, KeyError, AttributeError):
        pass
    if timeout_value is None:
        timeout_value = os.getenv("SESSION_TIMEOUT_MINUTES")
    if timeout_value is None:
        return None
    try:
        timeout_minutes = int(timeout_value)
    except (TypeError, ValueError):
        return None
    return timeout_minutes if timeout_minutes > 0 else None


# Session timeout defaults to 2 hours (120 minutes) so live demos and
# exhibitions don't drop the user mid-session. Override by setting
# SESSION_TIMEOUT_MINUTES in env or Streamlit secrets (set to 0 to disable).
SESSION_TIMEOUT_MINUTES = _get_session_timeout_minutes()
if SESSION_TIMEOUT_MINUTES is None:
    SESSION_TIMEOUT_MINUTES = 120

# Authentication check
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.user_email = None
    st.session_state.is_admin = False
    st.session_state.lab_name = None
    st.session_state.last_activity_time = None
    st.session_state.active_dataset_id = None  # Track selected dataset for filtering dashboards

# Check for session timeout
if SESSION_TIMEOUT_MINUTES and st.session_state.authenticated and st.session_state.last_activity_time:
    time_elapsed = (datetime.now() - st.session_state.last_activity_time).total_seconds() / 60
    if time_elapsed > SESSION_TIMEOUT_MINUTES:
        st.session_state.authenticated = False
        st.session_state.user_email = None
        st.session_state.is_admin = False
        st.session_state.last_activity_time = None
        st.session_state.lab_name = None
        st.warning("Session expired due to inactivity. Please log in again.")
        st.stop()
    else:
        # Update last activity time on each interaction
        st.session_state.last_activity_time = datetime.now()
else:
    # Set initial activity time on login
    if st.session_state.authenticated:
        st.session_state.last_activity_time = datetime.now()

def _get_admin_config():
    admin_email = None
    admin_password = None
    try:
        if hasattr(st, "secrets"):
            if "ADMIN_EMAIL" in st.secrets and "ADMIN_PASSWORD" in st.secrets:
                admin_email = st.secrets["ADMIN_EMAIL"]
                admin_password = st.secrets["ADMIN_PASSWORD"]
    except (FileNotFoundError, KeyError, AttributeError):
        pass
    load_dotenv()
    admin_email = admin_email or os.getenv("ADMIN_EMAIL")
    admin_password = admin_password or os.getenv("ADMIN_PASSWORD")
    return admin_email, admin_password


def _get_lab_email_mapping() -> Dict[str, str]:
    return get_lab_email_map()


def _apply_lab_filter(samples_df: pd.DataFrame, ast_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    lab_name = st.session_state.get("lab_name")
    if st.session_state.get("is_admin") or not lab_name:
        return samples_df, ast_df
    if samples_df.empty:
        return samples_df, ast_df
    lab_mask = samples_df['lab_name'].astype(str).str.strip().str.lower() == lab_name.strip().lower()
    filtered_samples = samples_df[lab_mask]
    if ast_df.empty:
        return filtered_samples, ast_df
    filtered_ast = ast_df[ast_df['sample_id'].astype(str).isin(filtered_samples['sample_id'].astype(str))]
    return filtered_samples, filtered_ast


# ─────────────────────────────────────────────────────────────────────────────
# Shared UI / data helpers
# ─────────────────────────────────────────────────────────────────────────────

def sidebar_multiselect_all(label: str, options, key: str) -> list:
    """Render an 'All'-aware sidebar multiselect.

    Returns the effective selection: all options when "All" is picked or
    nothing is picked, otherwise just the user-chosen values.
    `key` must be unique per call site to isolate widget state across pages.
    """
    options = list(options)
    if not options:
        return []
    opts = ["All"] + options
    picked = st.sidebar.multiselect(label, opts, default=["All"], key=key)
    if not picked or "All" in picked:
        return options
    return [p for p in picked if p != "All"]


@st.cache_data(ttl=300, show_spinner="Loading dataset…")
def _load_dataset_bundle(dataset_id: str, lab_name: Optional[str]):
    """Cached loader for the (samples, ast) dataframes of a dataset.

    Scoped to a single lab when `lab_name` is non-empty; admins pass ``None``.
    Cache entries are keyed on (dataset_id, lab_name), so switching datasets
    or lab users doesn't leak cached rows.
    """
    samples = db.get_dataset_samples(dataset_id)
    ast = db.get_dataset_ast(dataset_id)
    if lab_name and not samples.empty and 'lab_name' in samples.columns:
        mask = samples['lab_name'].astype(str).str.strip().str.lower() == lab_name.strip().lower()
        samples = samples[mask]
        if not ast.empty:
            ast = ast[ast['sample_id'].astype(str).isin(samples['sample_id'].astype(str))]
    return samples, ast


def _load_active_dataset():
    """Convenience wrapper for the common 'load the page's active dataset' pattern.

    Returns (samples_df, ast_df). Callers that have already guarded on
    ``st.session_state.active_dataset_id`` can rely on a non-empty id.
    """
    ds_id = st.session_state.active_dataset_id
    lab = None if st.session_state.get("is_admin") else st.session_state.get("lab_name")
    return _load_dataset_bundle(ds_id, lab)


def _render_dataset_banner(dataset_id):
    """Uniform banner shown above each dataset-scoped page."""
    st.info(f"Viewing dataset: {dataset_id}")


def _csv_download(df: pd.DataFrame, filename: str, key: str, label: str = "⬇ Download CSV"):
    """Render a CSV download button below a chart or table."""
    if df is None or df.empty:
        return
    try:
        csv = df.to_csv(index=False).encode("utf-8")
    except Exception:
        logger.exception("CSV encoding failed for %s", filename)
        return
    st.download_button(label, data=csv, file_name=filename, mime="text/csv", key=key)


def _empty_state(msg: str, icon: str = "📊"):
    """Consistent empty-state helper used in place of ad-hoc st.warning()."""
    st.info(f"{icon} {msg}")


ADMIN_EMAIL, ADMIN_PASSWORD = _get_admin_config()


@st.cache_resource(show_spinner=False)
def _bootstrap_admin_once():
    """Provision the admin account exactly once per process."""
    if not (ADMIN_EMAIL and ADMIN_PASSWORD):
        return False
    try:
        admin_user = db.get_user_by_email(ADMIN_EMAIL)
        password_hash = bcrypt.hashpw(ADMIN_PASSWORD.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        if not admin_user:
            db.create_user(ADMIN_EMAIL, password_hash, is_admin=True)
        else:
            if not admin_user.get("is_admin"):
                db.set_user_admin(ADMIN_EMAIL, True)
            if not admin_user.get("is_active"):
                db.update_user_status(admin_user["user_id"], True)
            db.update_user_password(ADMIN_EMAIL, password_hash)
        try:
            db.set_user_verified(ADMIN_EMAIL, True)
        except Exception:
            logger.exception("admin verified flag write failed")
        return True
    except Exception:
        logger.exception("admin account bootstrap failed")
        return False

_bootstrap_admin_once()


def _bootstrap_lab_accounts() -> None:
    """Provision sentinel-lab logins from the committed hash manifest.

    The deployed Streamlit Cloud app uses its own database (separate from the
    administrator's local Postgres), so each new deployment has to recreate
    the 19 sentinel-lab accounts before any lab can sign in.  We persist the
    bcrypt hashes (never plaintext) to ``db/lab_accounts.json`` and replay
    them here on startup -- this is idempotent: existing accounts get the
    same hash overwritten, so the credentials shared with the labs keep
    working across redeploys.
    """
    manifest = Path("db") / "lab_accounts.json"
    if not manifest.exists():
        return
    try:
        accounts = json.loads(manifest.read_text(encoding="utf-8"))
    except Exception:
        logger.exception("could not read %s", manifest)
        return
    for email, info in accounts.items():
        pw_hash = info.get("password_hash")
        if not pw_hash:
            continue
        try:
            existing = db.get_user_by_email(email)
            if existing is None:
                db.create_user(email, pw_hash, is_admin=False)
            else:
                db.update_user_password(email, pw_hash)
                if not existing.get("is_active"):
                    db.update_user_status(existing["user_id"], True)
            try:
                db.set_user_verified(email, True)
            except Exception:
                logger.exception("lab verified flag write failed for %s", email)
        except Exception:
            logger.exception("lab account bootstrap failed for %s", email)


@st.cache_resource(show_spinner=False)
def _bootstrap_lab_accounts_once():
    try:
        _bootstrap_lab_accounts()
        return True
    except Exception:
        logger.exception("lab account bootstrap top-level failure")
        return False

_bootstrap_lab_accounts_once()

def _get_flag(name: str) -> bool:
    val = None
    try:
        if hasattr(st, "secrets") and name in st.secrets:
            val = st.secrets.get(name)
    except (FileNotFoundError, KeyError, AttributeError):
        pass
    if val is None:
        val = os.getenv(name)
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in ("1", "true", "yes", "on")
    return False

try:
    if _get_flag("PURGE_NON_ADMIN_ON_DEPLOY"):
        flag_path = os.path.join("db", "purge_non_admin.flag")
        if not os.path.exists(flag_path):
            deleted_count, msg = db.delete_non_admin_users(ADMIN_EMAIL)
            os.makedirs("db", exist_ok=True)
            with open(flag_path, "w", encoding="utf-8") as f:
                f.write(f"{datetime.now().isoformat()} - {msg}")
            st.info(f"Startup maintenance: {msg}")
except Exception:
    logger.exception("startup maintenance PURGE_NON_ADMIN_ON_DEPLOY failed")

# If not authenticated, show login page
if not st.session_state.authenticated:
    st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&display=swap');
        /* Palette — deep editorial, scientific institutional */
        :root {
            --paper:         #ebe2cd;
            --paper-alt:     #ddd1b4;
            --surface:       #f5ecd7;
            --surface-alt:   #f9f2e0;
            --border:        #c9bc98;
            --border-strong: #a89777;
            --ink:           #12100a;
            --ink-muted:     #55503e;
            --ink-soft:      #7a735f;
            --forest:        #194238;
            --forest-dark:   #0c2620;
            --forest-deep:   #061712;
            --forest-soft:   #d8e2dc;
            --gold:          #b18a3a;
            --gold-soft:     #e6d39c;
            --terracotta:    #964517;
            --ochre:         #a27117;
            --oxide:         #741a10;
            --moss:          #2f5430;
            --steel:         #164161;
        }

        #MainMenu, footer { visibility: hidden; }
        [data-testid="stSidebar"],
        [data-testid="collapsedControl"],
        header[data-testid="stHeader"] { display: none !important; }

        .stApp {
            background: var(--paper);
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
            color: var(--ink);
        }

        /* Card sits in the middle of the viewport */
        .main {
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            min-height: 100vh !important;
            padding: 2.5rem 1.5rem !important;
        }
        .main .block-container,
        .main .stMainBlockContainer,
        .main [data-testid="stAppViewBlockContainer"],
        [data-testid="stAppViewBlockContainer"],
        .stApp .stMainBlockContainer {
            max-width: 960px !important;
            width: 100% !important;
            margin: 0 auto !important;
            padding: 0 !important;
            background: var(--surface) !important;
            border: 1px solid var(--border) !important;
            border-radius: 18px !important;
            overflow: hidden !important;
        }

        /* Two-column split — brand | form */
        .main [data-testid="stHorizontalBlock"] {
            gap: 0 !important;
            min-height: 560px;
        }
        .main [data-testid="stHorizontalBlock"] > [data-testid="column"],
        .main [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
            padding: 0 !important;
            align-self: stretch;
        }

        /* ── Left brand panel ───────────────────────────────────── */
        .login-brand-pane {
            background:
              radial-gradient(circle at 18% 22%, rgba(230, 211, 156, 0.09) 1px, transparent 1.4px) 0 0/34px 34px,
              radial-gradient(circle at 78% 68%, rgba(230, 211, 156, 0.06) 1px, transparent 1.4px) 17px 17px/44px 44px,
              linear-gradient(165deg, #194238 0%, #0c2620 55%, #061712 100%);
            color: #f4eedd;
            padding: 3rem 2.75rem;
            height: 100%;
            min-height: 580px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            position: relative;
        }
        .login-brand-pane::after {
            content: "";
            position: absolute;
            top: 0; right: -1px;
            width: 3px; height: 100%;
            background: linear-gradient(180deg, transparent 0%, rgba(177, 138, 58, 0.35) 25%, rgba(177, 138, 58, 0.35) 75%, transparent 100%);
        }
        .login-brand-pane .brand-mark {
            margin-bottom: 2rem;
            color: #e6d39c;
            display: inline-block;
            line-height: 0;
        }
        .login-brand-pane .brand-mark svg { display: block; }
        .login-brand-pane .brand-eyebrow {
            font-size: 0.72rem;
            font-weight: 600;
            letter-spacing: 0.18em;
            text-transform: uppercase;
            color: #e6d39c;
            margin-bottom: 0.6rem;
        }
        .login-brand-pane .brand-title {
            font-family: 'Fraunces', 'Georgia', serif;
            font-size: 2.4rem;
            font-weight: 600;
            letter-spacing: -0.02em;
            line-height: 1.1;
            color: #f4eedd;
            margin: 0 0 0.75rem 0;
        }
        .login-brand-pane .brand-sub {
            font-size: 0.98rem;
            color: rgba(244, 238, 221, 0.82);
            line-height: 1.55;
            max-width: 360px;
            margin: 0 0 2rem 0;
        }
        .login-brand-pane .brand-features {
            list-style: none;
            padding: 0;
            margin: 0 0 2rem 0;
            border-top: 1px solid rgba(244, 238, 221, 0.14);
        }
        .login-brand-pane .brand-features li {
            padding: 0.75rem 0;
            border-bottom: 1px solid rgba(244, 238, 221, 0.14);
            font-size: 0.88rem;
            color: rgba(244, 238, 221, 0.88);
            display: flex;
            align-items: center;
            gap: 0.7rem;
        }
        .login-brand-pane .brand-features li::before {
            content: "";
            width: 6px; height: 6px;
            background: #c9a85a;
            border-radius: 50%;
            flex-shrink: 0;
        }
        .login-brand-pane .brand-footer {
            font-size: 0.78rem;
            color: rgba(244, 238, 221, 0.55);
            letter-spacing: 0.04em;
        }

        /* ── Right form panel ───────────────────────────────────── */
        .login-form-pane { padding: 3rem 2.75rem 2rem 2.75rem; }
        .login-form-pane .form-eyebrow {
            font-size: 0.72rem;
            font-weight: 600;
            letter-spacing: 0.16em;
            text-transform: uppercase;
            color: var(--ink-muted);
            margin-bottom: 0.5rem;
        }
        .login-form-pane .form-title {
            font-family: 'Fraunces', 'Georgia', serif;
            font-size: 1.9rem;
            font-weight: 600;
            color: var(--ink);
            letter-spacing: -0.015em;
            margin: 0 0 0.4rem 0;
        }
        .login-form-pane .form-sub {
            font-size: 0.95rem;
            color: var(--ink-muted);
            margin: 0 0 1.75rem 0;
        }

        /* Tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 2rem;
            background: transparent;
            border-bottom: 1px solid var(--border);
            padding: 0;
            border-radius: 0;
        }
        .stTabs [data-baseweb="tab"] {
            background: transparent !important;
            border: none !important;
            border-radius: 0 !important;
            padding: 0.75rem 0 !important;
            font-family: 'Inter', sans-serif;
            font-weight: 500;
            font-size: 0.92rem;
            color: var(--ink-muted);
            position: relative;
        }
        .stTabs [aria-selected="true"] {
            background: transparent !important;
            color: var(--forest) !important;
            font-weight: 600;
        }
        .stTabs [aria-selected="true"]::after {
            content: "";
            position: absolute;
            left: 0; right: 0; bottom: -1px;
            height: 2px;
            background: var(--forest);
        }
        .stTabs [data-baseweb="tab-panel"] { padding-top: 1.5rem; }

        /* Form inputs */
        .stTextInput > div { max-width: none; margin: 0; }
        .stTextInput > label {
            font-size: 0.82rem !important;
            font-weight: 500 !important;
            color: var(--ink) !important;
            letter-spacing: 0.01em;
            margin-bottom: 0.35rem !important;
        }
        .stTextInput > div > div > input {
            background: var(--surface) !important;
            border: 1px solid var(--border) !important;
            border-radius: 10px !important;
            padding: 0.75rem 1rem !important;
            font-size: 0.95rem !important;
            color: var(--ink) !important;
            transition: border-color 0.15s ease;
        }
        .stTextInput > div > div > input::placeholder { color: var(--ink-soft) !important; }
        .stTextInput > div > div > input:focus {
            border-color: var(--forest) !important;
            outline: none !important;
            box-shadow: none !important;
        }

        /* Primary button */
        .stButton > button[kind="primary"] {
            background: var(--forest) !important;
            color: #f4eedd !important;
            border: 1px solid var(--forest) !important;
            border-radius: 10px !important;
            padding: 0.8rem 1.5rem !important;
            font-weight: 600 !important;
            font-size: 0.95rem !important;
            width: 100% !important;
            margin: 1rem 0 0 0 !important;
            display: block !important;
            transition: background 0.15s ease, border-color 0.15s ease;
        }
        .stButton > button[kind="primary"]:hover {
            background: var(--forest-dark) !important;
            border-color: var(--forest-dark) !important;
            transform: none !important;
            box-shadow: none !important;
        }

        /* Alerts inside login */
        .stAlert,
        .stSuccess, .stError, .stInfo, .stWarning {
            border-radius: 10px;
            background: var(--paper) !important;
            border: 1px solid var(--border);
        }
        .stError   { border-left: 3px solid var(--oxide); }
        .stSuccess { border-left: 3px solid var(--moss); }
        .stInfo    { border-left: 3px solid var(--forest); }
        .stWarning { border-left: 3px solid var(--ochre); }

        /* Information tab list prose */
        .login-form-pane .stSubheader,
        .login-form-pane h3 {
            font-family: 'Fraunces', 'Georgia', serif !important;
            color: var(--ink) !important;
            font-weight: 600 !important;
            font-size: 1.15rem !important;
            letter-spacing: -0.01em !important;
        }

        /* Footer under card */
        .login-footer {
            text-align: center;
            color: var(--ink-muted);
            font-size: 0.8rem;
            margin-top: 1.25rem;
        }
        .login-footer p { margin: 0.2rem 0; }
        </style>
    """, unsafe_allow_html=True)

    # Single placeholder wraps the whole login screen.  After a successful
    # submit we call login_placeholder.empty() *before* st.rerun() so the
    # browser DOM is wiped instantly -- otherwise Streamlit would briefly
    # re-render the login form during the rerun (the "login flashback").
    login_placeholder = st.empty()
    with login_placeholder.container():
        # Two-column split — left brand panel, right form panel
        left, right = st.columns([5, 6], gap="medium")

    with left:
        st.markdown("""
            <div class="login-brand-pane">
                <div>
                    <div class="brand-mark">
                        <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 56 56" width="56" height="56" aria-label="ICBB">
                            <path d="M28 3 L50 15.5 L50 40.5 L28 53 L6 40.5 L6 15.5 Z"
                                  fill="none" stroke="#e6d39c" stroke-width="1.4" stroke-linejoin="round"/>
                            <path d="M28 3 L50 15.5 L50 40.5 L28 53 L6 40.5 L6 15.5 Z"
                                  fill="none" stroke="#e6d39c" stroke-width="0.5" stroke-linejoin="round"
                                  transform="translate(0 0) scale(0.86 0.86) translate(2.3 3.9)"/>
                            <circle cx="28" cy="17" r="1.8" fill="#e6d39c"/>
                            <text x="28" y="36" text-anchor="middle"
                                  font-family="'Fraunces', Georgia, serif"
                                  font-size="11" font-weight="700"
                                  fill="#f4eedd" letter-spacing="1.2">ICBB</text>
                        </svg>
                    </div>
                    <div class="brand-eyebrow">National Surveillance</div>
                    <h1 class="brand-title">ICBB AMR Surveillance System</h1>
                    <p class="brand-sub">
                        A unified national platform for antimicrobial resistance reporting
                        across the environment, food, human, animal, and aquaculture sectors.
                    </p>
                    <ul class="brand-features">
                        <li>Sentinel-site laboratory network</li>
                        <li>WHONET &amp; GLASS aligned reporting</li>
                        <li>Geospatial hotspot &amp; trend analytics</li>
                    </ul>
                </div>
                <div class="brand-footer">© ICBB · AMR Surveillance Programme</div>
            </div>
        """, unsafe_allow_html=True)

    with right:
        st.markdown('<div class="login-form-pane">', unsafe_allow_html=True)
        st.markdown('<div class="form-eyebrow">Secure Access</div>', unsafe_allow_html=True)
        st.markdown('<div class="form-title">Welcome back</div>', unsafe_allow_html=True)
        st.markdown('<div class="form-sub">Sign in to continue to your surveillance workspace.</div>', unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["Sign in", "Information"])

        with tab1:
            # Optional one-shot diagnostic: append ?debug=1 to the URL to see
            # whether the lab/admin bootstrap actually populated the cloud DB.
            try:
                if st.query_params.get("debug") == "1":
                    try:
                        all_users = db.get_all_users()
                        st.info(
                            f"DB diagnostic — backend in use, total user rows: {len(all_users)}. "
                            f"Sample emails: {[u['email'] for u in all_users[:5]]}"
                        )
                    except Exception as diag_err:
                        st.warning(f"DB diagnostic failed: {diag_err}")
            except Exception:
                pass

            # Wrap inputs + button in a form so a single Enter / click submits
            # everything in ONE rerun.  Without this, typing in a password and
            # immediately clicking the button can require two clicks because
            # Streamlit processes the input change in one rerun and the button
            # click in the next.
            with st.form("login_form", clear_on_submit=False):
                login_email = st.text_input("Email address", placeholder="name@institution.gh", key="login_email")
                login_password = st.text_input("Password", type="password", placeholder="••••••••", key="login_password")
                submitted = st.form_submit_button("Sign in", use_container_width=True, type="primary")

            if submitted:
                if not login_email or not login_password:
                    st.error("Please fill in all fields")
                else:
                    # Normalize email: lowercase + strip whitespace (the lab
                    # manifest is keyed lowercase, so a user typing "KBTH@..."
                    # or with a stray space would otherwise hit "user not found")
                    login_email = login_email.strip().lower()
                    login_password = login_password.strip()
                    user = db.get_user_by_email(login_email)
                    if user and user['is_active']:
                        try:
                            if bcrypt.checkpw(login_password.encode("utf-8"), user['password_hash'].encode("utf-8")):
                                config_admin_email, _ = _get_admin_config()
                                target_admin_email = (config_admin_email or "").strip().lower()

                                is_admin_flag = user.get('is_admin')
                                if target_admin_email and login_email == target_admin_email:
                                    # Trust the cached bootstrap; do NOT issue
                                    # set_user_admin/update_user_status/
                                    # set_user_verified writes here.  Each one
                                    # is a Postgres round-trip (~200-400ms via
                                    # Supabase pooler) and they made the user
                                    # wait 1-2s after clicking Sign in.  The
                                    # admin bootstrap on startup already keeps
                                    # those flags in sync.
                                    is_admin_flag = 1

                                lab_mapping = _get_lab_email_mapping()
                                is_lab, lab_name = is_lab_user(login_email, lab_mapping)
                                if not is_admin_flag and not is_lab:
                                    st.error("Access denied. This system is restricted to approved labs.")
                                    st.session_state.authenticated = False
                                    st.session_state.user_email = None
                                    st.session_state.is_admin = False
                                    st.session_state.lab_name = None
                                    st.stop()

                                st.session_state.authenticated = True
                                st.session_state.user_email = login_email
                                st.session_state.last_activity_time = datetime.now()
                                st.session_state.is_admin = bool(is_admin_flag)
                                st.session_state.lab_name = lab_name if not is_admin_flag else None

                                # Defer last_login write to the next rerun via
                                # session_state so the UI can render the
                                # dashboard immediately.  We do the actual
                                # UPDATE on the *post-login* page where the
                                # extra ~300ms round-trip is hidden by the
                                # dashboard's data load.
                                st.session_state["_pending_last_login"] = login_email
                                # Flag a one-shot loading overlay for the
                                # *next* run so the user never sees the
                                # login form during the dashboard's first
                                # render after st.rerun().
                                st.session_state["_show_login_overlay"] = True
                                # Wipe the login DOM *before* triggering the
                                # rerun, then show a tiny full-screen loading
                                # overlay so the user never sees the login
                                # form flash again during the script restart.
                                login_placeholder.empty()
                                st.markdown(
                                    """
                                    <div style="position:fixed;inset:0;background:#ebe2cd;
                                                display:flex;align-items:center;justify-content:center;
                                                z-index:99999;font-family:'Fraunces',Georgia,serif;
                                                color:#194238;font-size:1.1rem;letter-spacing:0.02em;">
                                        Loading your workspace…
                                    </div>
                                    """,
                                    unsafe_allow_html=True,
                                )
                                st.rerun()
                            else:
                                st.error("Invalid email or password")
                        except Exception as e:
                            st.error(f"Login error: {str(e)}")
                    else:
                        st.error("Invalid email or password, or account is inactive")

        with tab2:
            st.markdown("#### Approved laboratories")
            st.caption("Lab accounts are pre-configured by the programme administrator. Only personnel from the sentinel sites below have access.")
            st.markdown(
                """
                - Eastern Regional Hospital
                - St. Martin De Porres Hospital Eikwe
                - Sekondi Public Health Reference Laboratory
                - Ho Teaching Hospital
                - Tamale Teaching Hospital
                - Komfo Anokye Teaching Hospital
                - Korle-Bu Teaching Hospital
                - Lekma Hospital
                - Sunyani Teaching Hospital
                - Cape Coast Teaching Hospital
                - National Food Safety Laboratory
                - CSIR Water Research Institute
                - Accra Veterinary Laboratory
                - Kumasi Veterinary Laboratory
                - Quadushah Medical Diagnostic Limited
                - Central Veterinary Laboratory
                - Pong Tamale School
                - Metropolis Health Care Limited
                - Alma Medical Laboratory Ltd
                """
            )
            st.caption("Contact the AMR Surveillance Programme administrator for access.")

        st.markdown('</div>', unsafe_allow_html=True)  # /login-form-pane

    st.markdown(
        '<div class="login-footer">Environment · Food · Human · Animal · Aquaculture</div>',
        unsafe_allow_html=True,
    )

    st.stop()

# ============================================================================
# MAIN APP STYLING (After Authentication)
# ============================================================================

# Render a full-screen overlay on the very first run after login so the
# slow dashboard build doesn't leave the old login DOM visible (the
# "login flashback").  We emit it BEFORE any heavy markdown so Streamlit
# delivers it as one of the earliest deltas; the dashboard renders
# underneath, then a tiny JS timer fades it out once the workspace is
# painted.  One-shot: the flag is consumed immediately.
if st.session_state.pop("_show_login_overlay", False):
    st.markdown(
        """
        <div id="_post_login_overlay" style="position:fixed;inset:0;
                    background:#ebe2cd;display:flex;align-items:center;
                    justify-content:center;z-index:99999;
                    font-family:'Fraunces',Georgia,serif;color:#194238;
                    font-size:1.1rem;letter-spacing:0.02em;
                    transition:opacity 250ms ease;">
            Loading your workspace…
        </div>
        <script>
          // Fade out once the dashboard has finished painting.
          setTimeout(function(){
            var el = window.parent.document.getElementById('_post_login_overlay')
                  || document.getElementById('_post_login_overlay');
            if (el) { el.style.opacity = '0';
                      setTimeout(function(){ el.remove(); }, 300); }
          }, 1500);
        </script>
        """,
        unsafe_allow_html=True,
    )

# Flush any deferred bookkeeping writes that we skipped during the login
# click handler so the dashboard could appear instantly.  Failures here are
# never user-visible.
_pending_login_email = st.session_state.pop("_pending_last_login", None)
if _pending_login_email:
    try:
        db.update_last_login(_pending_login_email)
    except Exception:
        logger.exception("deferred update_last_login failed for %s", _pending_login_email)

# Editorial, warm, human-designed theme for the authenticated app shell
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&display=swap');
    :root {
        /* Clean white surfaces for the main app body so charts and tables
           sit on neutral backgrounds.  The sidebar keeps its own dark
           forest palette via --forest-deep below. */
        --paper:         #ffffff;
        --paper-alt:     #f5f5f5;
        --surface:       #ffffff;
        --surface-alt:   #fafafa;
        --surface-soft:  #ffffff;
        --border:        #e2e8f0;
        --border-strong: #cbd5e1;
        --ink:           #0f172a;
        --ink-muted:     #334155;
        --ink-soft:      #475569;
        --forest:        #1d4d43;
        --forest-dark:   #123830;
        --forest-deep:   #0d2a23;
        --forest-soft:   #d8e2dc;
        --gold:          #b18a3a;
        --gold-soft:     #e6d39c;
        --terracotta:    #964517;
        --ochre:         #a27117;
        --oxide:         #741a10;
        --moss:          #2f5430;
        --steel:         #164161;
        --radius-sm: 8px;
        --radius-md: 10px;
        --radius-lg: 14px;
    }

    /* Global font scale.  Streamlit's default root is 16px; bumping it
       to 17px scales every rem-based size in this stylesheet uniformly
       so text reads better without re-flowing layouts.  em-based sizes
       inside the sidebar inherit naturally. */
    html, body, .stApp { font-size: 17px; }
    .main .block-container,
    .main .stMainBlockContainer,
    [data-testid="stAppViewBlockContainer"] { font-size: 1rem; }
    p, li, label, span, div { font-size: inherit; }
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li { font-size: 1.02rem; line-height: 1.6; }
    .stCaption, [data-testid="stCaptionContainer"], small {
        font-size: 0.92rem !important;
        color: var(--ink-muted) !important;
    }
    /* Inputs / selects: keep them readable but not larger than the
       container so dropdown chips don't wrap. */
    .stSelectbox, .stMultiSelect, .stTextInput, .stNumberInput,
    .stDateInput, .stTextArea, .stRadio, .stCheckbox { font-size: 0.98rem; }
    .stSelectbox label, .stMultiSelect label, .stTextInput label,
    .stNumberInput label, .stDateInput label, .stTextArea label,
    .stRadio label, .stCheckbox label, .stSlider label, .stFileUploader label {
        font-size: 0.95rem !important;
        color: var(--ink) !important;
        font-weight: 500 !important;
    }
    /* Tables / dataframes */
    [data-testid="stDataFrame"] { font-size: 0.95rem; }
    /* Tabs */
    .stTabs [data-baseweb="tab"] { font-size: 1rem !important; }

    .stApp {
        background: var(--paper) !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        color: var(--ink);
        -webkit-font-smoothing: antialiased;
    }

    /* ── Sidebar — deep forest with hex-dot texture ────────────── */
    [data-testid="stSidebar"] {
        background-color: var(--forest-deep) !important;
        background-image:
            radial-gradient(circle at 20% 14%, rgba(230, 211, 156, 0.10) 0, transparent 45%),
            radial-gradient(circle at 82% 78%, rgba(177, 138, 58, 0.08) 0, transparent 50%),
            url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='64' height='64' viewBox='0 0 64 64'><g fill='none' stroke='%23e6d39c' stroke-width='0.8' stroke-opacity='0.08'><polygon points='32,8 52,20 52,44 32,56 12,44 12,20'/><circle cx='32' cy='32' r='1.6' fill='%23e6d39c' fill-opacity='0.12' stroke='none'/></g></svg>") !important;
        background-repeat: no-repeat, no-repeat, repeat !important;
        background-size: auto, auto, 64px 64px !important;
        border-right: 1px solid rgba(230, 211, 156, 0.10);
    }
    [data-testid="stSidebar"] > div:first-child {
        background: transparent;
        padding-top: 0;
    }
    [data-testid="stSidebar"] .stMarkdown,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] label { color: #ede4cb !important; }
    [data-testid="stSidebar"] .stRadio > label { display: none; }

    [data-testid="stSidebar"] .stButton > button {
        background: rgba(138, 36, 25, 0.15) !important;
        color: #f2c8bc !important;
        border: 1px solid rgba(138, 36, 25, 0.35) !important;
        border-radius: var(--radius-sm);
        font-weight: 500;
        font-size: 0.85em;
        padding: 0.45rem 1rem;
        transition: background 0.15s ease, color 0.15s ease;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(138, 36, 25, 0.28) !important;
        color: #fadcd1 !important;
        transform: none;
    }

    [data-testid="stSidebar"] [data-testid="stExpander"],
    [data-testid="stSidebar"] [data-testid="stExpander"] details {
        background: transparent !important;
        border: none !important;
        margin-bottom: 0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary {
        background: transparent !important;
        border: none !important;
        padding: 0.5rem 0.6rem !important;
        color: #a59d85 !important;
        font-size: 0.7em !important;
        font-weight: 600 !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary:hover { color: #e8e0cc !important; }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary p,
    [data-testid="stSidebar"] [data-testid="stExpander"] summary span {
        color: inherit !important;
        font-size: inherit !important;
        font-weight: inherit !important;
        letter-spacing: inherit !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] summary svg {
        width: 12px !important; height: 12px !important;
        color: #7a7260 !important;
    }
    [data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stExpanderDetails"] {
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
    }

    /* Nav items — left accent, flat, higher-contrast text */
    [data-testid="stSidebar"] .nav-item-btn { margin: 0 !important; padding: 0 !important; }
    [data-testid="stSidebar"] .nav-item-btn button {
        background: transparent !important;
        color: #ede4cb !important;
        border: none !important;
        border-left: 3px solid transparent !important;
        text-align: left !important;
        padding: 0.55rem 0.75rem 0.55rem 1rem !important;
        border-radius: 0 var(--radius-sm) var(--radius-sm) 0 !important;
        font-size: 0.9em !important;
        font-weight: 500 !important;
        width: 100% !important;
        margin: 1px 0 !important;
        line-height: 1.4 !important;
        letter-spacing: 0.005em !important;
        transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease;
    }
    [data-testid="stSidebar"] .nav-item-btn button:hover {
        background: rgba(230, 211, 156, 0.10) !important;
        color: #fff8e1 !important;
    }
    [data-testid="stSidebar"] .nav-item-btn.nav-active button {
        background: rgba(230, 211, 156, 0.16) !important;
        color: #f6db9c !important;
        font-weight: 600 !important;
        border-left-color: #e6d39c !important;
    }

    /* Last-updated badge */
    .last-updated-badge {
        display: flex; align-items: center; gap: 0.5rem;
        background: rgba(244, 238, 221, 0.04);
        border: 1px solid rgba(244, 238, 221, 0.08);
        border-radius: var(--radius-sm);
        padding: 0.55rem 0.65rem;
        margin: 0.25rem 0;
        color: #a59d85;
        font-size: 0.75em;
    }
    .last-updated-badge strong { color: #e8e0cc; }
    .last-updated-badge .pulse-dot {
        width: 6px; height: 6px;
        border-radius: 50%;
        background: #c9a85a;
        display: inline-block; flex-shrink: 0;
    }

    .sidebar-user-card {
        background: rgba(244, 238, 221, 0.04);
        border: 1px solid rgba(244, 238, 221, 0.08);
        border-radius: var(--radius-md);
        padding: 0.75rem 0.85rem;
        margin-bottom: 0.5rem;
    }
    .sidebar-user-card .user-label { font-size: 0.68em; color: #a59d85; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 3px; }
    .sidebar-user-card .user-email { font-weight: 600; color: #f4eedd; font-size: 0.9em; margin: 2px 0; }
    .sidebar-user-card .user-role  { font-size: 0.78em; margin: 2px 0; }
    .sidebar-user-card .user-lab   { font-size: 0.78em; color: #a59d85; margin: 2px 0; }

    .sidebar-brand { text-align: center; padding: 1.4rem 0 0.9rem 0; }
    .sidebar-brand .brand-icon {
        display: inline-flex; align-items: center; justify-content: center;
        width: 46px; height: 46px;
        margin-bottom: 0.55rem;
    }
    .sidebar-brand .brand-icon svg { width: 46px; height: 46px; display: block; }
    .sidebar-brand .brand-name {
        font-family: 'Fraunces', 'Georgia', serif;
        font-size: 1.12rem; font-weight: 600; color: #f4eedd;
        letter-spacing: 0.005em;
    }
    .sidebar-brand .brand-sub {
        font-size: 0.7em; color: #c7b88a; margin-top: 4px;
        letter-spacing: 0.14em; text-transform: uppercase; font-weight: 500;
    }

    .sidebar-divider {
        height: 1px;
        background: rgba(244, 238, 221, 0.08);
        margin: 0.7rem 0;
        border: none;
    }

    /* Main content area */
    .main {
        display: block !important;
        align-items: unset !important;
        justify-content: unset !important;
        min-height: unset !important;
        padding: 0 !important;
    }
    .main .block-container,
    .main .stMainBlockContainer,
    .main [data-testid="stAppViewBlockContainer"],
    [data-testid="stAppViewBlockContainer"] {
        max-width: none !important;
        width: auto !important;
        margin: 0 !important;
        padding: 2rem 2.5rem !important;
        background: transparent !important;
        border: none !important;
        border-radius: 0 !important;
    }

    /* Typography — serif display, warm ink body */
    h1 {
        font-family: 'Fraunces', 'Georgia', serif !important;
        color: var(--ink) !important;
        font-weight: 600 !important;
        letter-spacing: -0.02em !important;
        background: none !important;
        -webkit-text-fill-color: initial !important;
    }
    h2 {
        font-family: 'Fraunces', 'Georgia', serif !important;
        color: var(--forest) !important;
        font-weight: 600 !important;
        letter-spacing: -0.015em !important;
    }
    h3, h4 {
        font-family: 'Inter', sans-serif !important;
        color: var(--ink) !important;
        font-weight: 600 !important;
        letter-spacing: -0.005em !important;
    }
    p, li, label, span, div { color: inherit; }
    .stCaption, [data-testid="stCaptionContainer"], small { color: var(--ink-muted) !important; }

    /* Metric cards */
    [data-testid="stMetric"], .stMetric {
        background: var(--surface-soft);
        padding: 1.1rem 1.2rem;
        border-radius: var(--radius-lg);
        border: 1px solid var(--border);
    }
    [data-testid="stMetricLabel"] {
        color: var(--ink-muted) !important;
        font-size: 0.88rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.06em !important;
        text-transform: uppercase !important;
    }
    [data-testid="stMetricValue"] {
        color: var(--ink) !important;
        font-family: 'Fraunces', 'Georgia', serif !important;
        font-weight: 600 !important;
        font-size: 2rem !important;
    }
    [data-testid="stMetricDelta"] { color: var(--ink-muted) !important; font-size: 0.9rem !important; }

    /* Expanders in main content */
    [data-testid="stExpander"] summary {
        background: var(--surface);
        border-radius: var(--radius-md);
        font-weight: 500;
        color: var(--ink) !important;
    }

    /* Bordered containers */
    [data-testid="stVerticalBlockBorderWrapper"] {
        background: var(--surface);
        border: 1px solid var(--border) !important;
        border-radius: var(--radius-lg) !important;
    }

    /* Buttons */
    .stButton > button {
        background: var(--forest);
        color: #f4eedd;
        border: 1px solid var(--forest);
        border-radius: var(--radius-md);
        padding: 0.55rem 1.3rem;
        font-weight: 500;
        letter-spacing: 0.005em;
        transition: background 0.15s ease, border-color 0.15s ease;
    }
    .stButton > button:hover {
        background: var(--forest-dark);
        border-color: var(--forest-dark);
        transform: none;
        color: #f4eedd;
    }
    .stButton > button[kind="secondary"] {
        background: var(--surface);
        color: var(--forest);
        border: 1px solid var(--border-strong);
    }
    .stButton > button[kind="secondary"]:hover {
        background: var(--forest-soft);
        color: var(--forest-dark);
        border-color: var(--forest);
    }

    .stDownloadButton > button {
        background: var(--surface);
        color: var(--forest);
        border: 1px solid var(--border-strong);
        border-radius: var(--radius-md);
        font-weight: 500;
    }
    .stDownloadButton > button:hover {
        background: var(--forest-soft);
        color: var(--forest-dark);
        border-color: var(--forest);
    }

    /* Inputs */
    .stSelectbox > div > div,
    .stMultiSelect > div > div,
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stDateInput > div > div > input,
    .stTextArea textarea {
        border-radius: var(--radius-md);
        border: 1px solid var(--border) !important;
        background: var(--surface) !important;
        color: var(--ink) !important;
        transition: border-color 0.15s ease;
    }
    .stSelectbox > div > div:focus-within,
    .stMultiSelect > div > div:focus-within,
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stDateInput > div > div > input:focus,
    .stTextArea textarea:focus {
        border-color: var(--forest) !important;
        outline: none;
    }

    /* Main tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
        background: transparent;
        border-bottom: 1px solid var(--border);
        padding: 0;
        border-radius: 0;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        border: none !important;
        border-radius: 0 !important;
        padding: 0.75rem 0 !important;
        font-weight: 500;
        color: var(--ink-muted);
        position: relative;
    }
    .stTabs [aria-selected="true"] {
        background: transparent !important;
        color: var(--forest) !important;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"]::after {
        content: "";
        position: absolute;
        left: 0; right: 0; bottom: -1px;
        height: 2px;
        background: var(--forest);
    }

    /* Alerts — warm-paper bg + coloured left border */
    .stSuccess { background: var(--paper-alt); border: 1px solid var(--border); border-left: 3px solid var(--moss);       border-radius: var(--radius-md); color: var(--ink); }
    .stError   { background: var(--paper-alt); border: 1px solid var(--border); border-left: 3px solid var(--oxide);      border-radius: var(--radius-md); color: var(--ink); }
    .stInfo    { background: var(--paper-alt); border: 1px solid var(--border); border-left: 3px solid var(--forest);     border-radius: var(--radius-md); color: var(--ink); }
    .stWarning { background: var(--paper-alt); border: 1px solid var(--border); border-left: 3px solid var(--ochre);      border-radius: var(--radius-md); color: var(--ink); }

    hr {
        border: none;
        height: 1px;
        background: var(--border);
        margin: 1.5rem 0;
    }

    header[data-testid="stHeader"] {
        background: transparent;
        border-bottom: 1px solid var(--border);
    }

    [data-testid="stTable"], .stDataFrame {
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        background: var(--surface);
    }

    /* Plotly charts sit on the paper — no background colour bleed */
    .stPlotlyChart { background: transparent; }

    /* Page title accent — subtle serif system */
    .page-eyebrow {
        font-size: 0.72rem;
        font-weight: 600;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: var(--terracotta);
        margin-bottom: 0.35rem;
    }

    /* ── Sticky page header (title + ticker) ──────────────────── */
    /* Pin the page title and the animated tagline strip to the top of the
       main content area so they stay visible while scrolling. The ticker
       animation inside continues to run normally.

       Streamlit wraps every st.markdown in nested <div> layers, so
       `position: sticky` only works if it is applied to the OUTER
       `stElementContainer` (a direct child of the scrolling main column).
       We use :has() to target that container and lift the wrapper up. */
    div[data-testid="stElementContainer"]:has(> div .amr-sticky-header),
    div[data-testid="stElementContainer"]:has(.amr-sticky-header) {
        position: sticky !important;
        top: 0 !important;
        z-index: 1000 !important;
        /* Fully opaque so scrolled content cannot bleed through. */
        background: var(--paper) !important;
        padding: 0.4rem 0 0.5rem 0 !important;
        margin-bottom: 0 !important;
        border-bottom: 1px solid var(--border);
        box-shadow: 0 8px 16px -10px rgba(0,0,0,0.25);
    }
    .amr-sticky-header {
        background: var(--paper);
    }
    /* Streamlit's own top header (data-testid="stHeader") is positioned
       absolutely at top:0 with z-index ~999990 (above our sticky title)
       and carries a 1px bottom border that visually crosses through the
       "ICBB-AMRSS" text when content scrolls beneath it. The toolbar
       inside it is transparent already, so we just strip the border and
       any background/shadow so the line disappears entirely. */
    header[data-testid="stHeader"] {
        background: transparent !important;
        border-bottom: 0 !important;
        box-shadow: none !important;
    }
    header[data-testid="stHeader"]::before,
    header[data-testid="stHeader"]::after {
        display: none !important;
        border: 0 !important;
        background: transparent !important;
    }
    .amr-sticky-header h1 {
        margin: 0 0 0.35rem 0 !important;
    }
    .amr-sticky-header .amr-ticker {
        margin: 0 !important;
    }
    /* Streamlit auto-injects a heading anchor link icon and (sometimes) a
       thin underline on the heading element itself — strip both inside the
       sticky strip so nothing visually bisects the title text. */
    .amr-sticky-header h1 a,
    .amr-sticky-header h1 [data-testid="stHeaderActionElements"] {
        display: none !important;
    }
    .amr-sticky-header h1::after,
    .amr-sticky-header h1::before {
        display: none !important;
    }

    /* ── Sticky per-page heading ──────────────────────────────────
       The page's heading container is found at runtime by a tiny
       parent-document script (rendered just below). The script applies
       the `amr-stick-heading` / `amr-stick-caption` classes so we can
       style them here. This is more reliable than CSS sibling selectors
       because Streamlit injects empty wrapper elements that vary per
       page. */
    [data-testid="stElementContainer"].amr-stick-heading {
        position: sticky !important;
        top: 140px !important;
        z-index: 998 !important;
        background: var(--paper) !important;
        padding-top: 0.4rem !important;
        padding-bottom: 0.35rem !important;
        border-bottom: 1px solid var(--border);
        box-shadow: 0 6px 14px -10px rgba(0,0,0,0.18);
    }
    [data-testid="stElementContainer"].amr-stick-caption {
        position: sticky !important;
        top: 215px !important;
        z-index: 997 !important;
        background: var(--paper) !important;
        padding-top: 0.2rem !important;
        padding-bottom: 0.45rem !important;
        border-bottom: 1px solid var(--border);
        box-shadow: 0 6px 14px -10px rgba(0,0,0,0.18);
    }
    /* Hide the auto-anchor link icon on any sticky heading too. */
    [data-testid="stElementContainer"].amr-stick-heading [data-testid="stHeaderActionElements"] {
        display: none !important;
    }

    /* ── Animated tagline ticker ──────────────────────────────── */
    .amr-ticker {
        position: relative;
        overflow: hidden;
        width: 100%;
        background: linear-gradient(90deg, var(--paper-alt) 0%, var(--surface-alt) 50%, var(--paper-alt) 100%);
        border-top: 1px solid var(--border);
        border-bottom: 1px solid var(--border);
        padding: 10px 0;
        margin: 0.2rem 0 1.3rem 0;
    }
    .amr-ticker::before, .amr-ticker::after {
        content: "";
        position: absolute; top: 0; bottom: 0; width: 90px; z-index: 2;
        pointer-events: none;
    }
    .amr-ticker::before { left: 0;  background: linear-gradient(90deg, var(--paper-alt) 0%, rgba(235,226,205,0) 100%); }
    .amr-ticker::after  { right: 0; background: linear-gradient(270deg, var(--paper-alt) 0%, rgba(235,226,205,0) 100%); }
    .amr-ticker__track {
        display: inline-flex;
        white-space: nowrap;
        animation: amrTickerScroll 48s linear infinite;
        will-change: transform;
    }
    .amr-ticker:hover .amr-ticker__track { animation-play-state: paused; }
    .amr-ticker__item {
        display: inline-flex; align-items: center;
        font-family: 'Inter', sans-serif;
        font-size: 0.82rem; font-weight: 500;
        color: var(--ink-muted); letter-spacing: 0.04em;
        padding: 0 2.4rem;
    }
    .amr-ticker__item strong {
        font-family: 'Fraunces', 'Georgia', serif;
        font-weight: 600; color: var(--forest-dark);
        margin-right: 0.45rem;
    }
    .amr-ticker__dot {
        display: inline-block; width: 5px; height: 5px; border-radius: 50%;
        background: var(--gold); margin: 0 1.2rem; opacity: 0.85;
    }
    @keyframes amrTickerScroll {
        from { transform: translateX(0); }
        to   { transform: translateX(-50%); }
    }

    /* ── Container blending — subtle cream lift on paper ──────── */
    [data-testid="stVerticalBlockBorderWrapper"],
    div[data-testid="stVerticalBlock"] > div > div[data-testid="stVerticalBlockBorderWrapper"] {
        background: var(--surface-soft) !important;
        border: 1px solid var(--border) !important;
        border-radius: var(--radius-md) !important;
    }
    [data-testid="stExpander"] {
        background: var(--surface-soft) !important;
        border: 1px solid var(--border) !important;
        border-radius: var(--radius-md) !important;
    }
    [data-testid="stMetric"] {
        background: var(--surface-soft) !important;
    }

    /* Plotly charts blend with their parent container — no white slab */
    .stPlotlyChart,
    [data-testid="stPlotlyChart"],
    [data-testid="stPlotlyChart"] > div,
    [data-testid="stPlotlyChart"] .js-plotly-plot,
    [data-testid="stPlotlyChart"] .plot-container,
    [data-testid="stPlotlyChart"] .main-svg {
        background: transparent !important;
    }

    /* Data tables + dataframes — warm surface instead of raw white */
    [data-testid="stDataFrame"],
    [data-testid="stDataFrame"] .glideDataEditor,
    [data-testid="stTable"] {
        background: var(--surface-soft) !important;
    }

    /* Streamlit default markdown rule → warm, subtle */
    hr { border: none; border-top: 1px solid var(--border); margin: 1.2rem 0; }
    </style>
""", unsafe_allow_html=True)

# ── Auto-select most recent accessible dataset on (re)login ────────
# Ensures uploaded data remains visible after logout/login — the DB keeps
# the rows, but `active_dataset_id` in session_state is None on a fresh
# login, which previously left every dashboard page saying "no data".
#
# Cached so we don't re-query the dataset list on every page navigation —
# that single SELECT was opening a new Postgres connection per rerun and
# was the main reason page navigation "got stuck".
@st.cache_data(ttl=30, show_spinner=False)
def _cached_all_datasets():
    return db.get_all_datasets() or []

try:
    _visible = _cached_all_datasets()
    _admin_email_cfg, _ = _get_admin_config()
    _admin_email_norm = (_admin_email_cfg or "").strip().lower()
    if not st.session_state.is_admin and _admin_email_norm:
        _visible = [d for d in _visible
                    if (d.get("uploaded_by") or "").strip().lower() != _admin_email_norm]
    if st.session_state.get("lab_name"):
        _user_norm = (st.session_state.get("user_email") or "").strip().lower()
        _visible = [d for d in _visible
                    if (d.get("uploaded_by") or "").strip().lower() == _user_norm]
    _visible_ids = {d["dataset_id"] for d in _visible}
    _current_id = st.session_state.get("active_dataset_id")
    if _visible and (not _current_id or _current_id not in _visible_ids):
        st.session_state.active_dataset_id = _visible[0]["dataset_id"]
except Exception:
    logger.exception("auto-select of active dataset failed")

# App title and animated tagline ticker — wrapped in a sticky header so they
# stay pinned to the top of the main view when the user scrolls. The ticker
# animation continues to run as normal inside the sticky strip.
_ticker_item = (
    "<span class='amr-ticker__item'>"
    "<strong>ICBB AMR Surveillance System</strong>"
    "Multi-source Surveillance"
    "<span class='amr-ticker__dot'></span>"
    "Environment"
    "<span class='amr-ticker__dot'></span>"
    "Food"
    "<span class='amr-ticker__dot'></span>"
    "Human"
    "<span class='amr-ticker__dot'></span>"
    "Animal"
    "<span class='amr-ticker__dot'></span>"
    "Aquaculture"
    "<span class='amr-ticker__dot'></span>"
    "Ghana"
    "</span>"
)
st.markdown(
    "<div class='amr-sticky-header'>"
    "<h1>ICBB-AMRSS</h1>"
    "<div class='amr-ticker'><div class='amr-ticker__track'>"
    f"{_ticker_item * 4}"
    "</div></div>"
    "</div>",
    unsafe_allow_html=True,
)

# ── Sticky page-heading helper ──────────────────────────────────────
# Streamlit puts a few invisible wrapper containers between the title
# strip and the page's actual heading, which makes pure-CSS sibling
# selectors unreliable. This tiny script (loaded inside an iframe but
# reaching parent.document) finds the FIRST visible heading that comes
# after the title strip on every rerun and tags its container so the
# CSS rules above can pin it. Same for the optional caption/subtitle
# that sits directly under the heading.
import streamlit.components.v1 as _components
_components.html(
    """
<script>
(function() {
  const apply = () => {
    try {
      const doc = window.parent.document;
      const titleEl = doc.querySelector('.amr-sticky-header');
      if (!titleEl) return;
      const titleEC = titleEl.closest('[data-testid="stElementContainer"]');
      if (!titleEC) return;

      // Clear previous tags so we don't accumulate stickies after reruns.
      doc.querySelectorAll('.amr-stick-heading, .amr-stick-caption')
         .forEach(n => n.classList.remove('amr-stick-heading', 'amr-stick-caption'));

      // Walk forward siblings; tag the first heading-bearing container,
      // then tag the immediately following caption/paragraph container
      // (if any). Stop as soon as we hit anything heavier so we don't
      // pin charts/tables.
      let n = titleEC.nextElementSibling;
      let headingDone = false;
      let steps = 0;
      while (n && steps < 12) {
        steps++;
        const isEC = n.matches && n.matches('[data-testid="stElementContainer"]');
        if (!isEC) { n = n.nextElementSibling; continue; }

        // Skip our own zero-height components.html iframe wrapper and any
        // other invisible/empty containers (Streamlit injects a few of
        // these between the title strip and the page's real heading).
        const rect = n.getBoundingClientRect();
        const heading = n.querySelector('h1, h2, h3, [data-testid="stHeading"]');
        const caption = n.querySelector('[data-testid="stCaptionContainer"]');
        const para    = n.querySelector('.stMarkdown > div > p, .stMarkdown p');
        const heavy   = n.querySelector(
          '[data-testid="stHorizontalBlock"], [data-testid="stMetric"],'
          + ' [data-testid="stDataFrame"], [data-testid="stPlotlyChart"],'
          + ' [data-testid="stTable"], [data-testid="stForm"], canvas, svg.main-svg'
        );
        const isInvisible = rect.height < 4 || (!heading && !caption && !para && !heavy);

        if (!headingDone) {
          if (heading) {
            n.classList.add('amr-stick-heading');
            headingDone = true;
          } else if (heavy) {
            // Heading missing on this page; nothing to pin.
            break;
          } else if (isInvisible) {
            // Skip empty wrapper / our own iframe and keep looking.
            n = n.nextElementSibling; continue;
          }
        } else {
          if (caption || (para && !heavy)) {
            n.classList.add('amr-stick-caption');
          }
          // Stop after one extra container regardless.
          break;
        }
        n = n.nextElementSibling;
      }
    } catch (e) { /* swallow — never break the page */ }
  };

  // Run now and also re-run when Streamlit re-renders the main column.
  apply();
  try {
    const doc = window.parent.document;
    const root = doc.querySelector('section[data-testid="stMain"]') || doc.body;
    const obs = new MutationObserver(() => { apply(); });
    obs.observe(root, { childList: true, subtree: true });
  } catch (e) { /* ignore */ }
})();
</script>
""",
    height=0,
)

# Sidebar navigation with user info and admin panel
with st.sidebar:
    # Logo/Title
    st.markdown("""
        <div class="sidebar-brand">
            <div class="brand-icon">
                <svg viewBox="0 0 56 56" xmlns="http://www.w3.org/2000/svg" aria-label="ICBB logo">
                    <path d="M28 4 L50 17 L50 39 L28 52 L6 39 L6 17 Z"
                          fill="none" stroke="#e6d39c" stroke-width="1.6" stroke-linejoin="round"/>
                    <circle cx="28" cy="12.5" r="1.6" fill="#e6d39c"/>
                    <text x="28" y="33" text-anchor="middle"
                          font-family="Fraunces, Georgia, serif" font-size="12" font-weight="600"
                          letter-spacing="0.5" fill="#f4eedd">ICBB</text>
                </svg>
            </div>
            <div class="brand-name">ICBB-AMRSS</div>
            <div class="brand-sub">Surveillance System</div>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

    # ── Last Updated indicator ────────────────────────────────────
    # Cached so we don't open a fresh DB connection on every nav click.
    @st.cache_data(ttl=60, show_spinner=False)
    def _sidebar_last_uploaded_at():
        _conn = db.get_connection()
        try:
            row = _conn.execute(
                "SELECT uploaded_at FROM datasets ORDER BY uploaded_at DESC LIMIT 1"
            ).fetchone()
            return row[0] if row and row[0] else None
        finally:
            try:
                _conn.close()
            except Exception:
                pass

    try:
        _last_at = _sidebar_last_uploaded_at()
        if _last_at:
            _last_ts = datetime.fromisoformat(_last_at)
            _delta = datetime.now() - _last_ts
            if _delta.days > 0:
                _ago = f"{_delta.days}d ago"
            elif _delta.seconds >= 3600:
                _ago = f"{_delta.seconds // 3600}h ago"
            else:
                _ago = f"{max(1, _delta.seconds // 60)}m ago"
            st.markdown(f"""
                <div class="last-updated-badge">
                    <span class="pulse-dot"></span>
                    <span>Updated: <strong>{_last_ts.strftime('%d %b %Y, %H:%M')}</strong> · {_ago}</span>
                </div>
            """, unsafe_allow_html=True)
    except Exception:
        logger.exception("last-updated sidebar badge render failed")

    st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

    # ── Grouped Navigation ─────────────────────────────────────────
    # Initialise active page in session state
    if "active_page" not in st.session_state:
        st.session_state.active_page = "Home"

    _NAV_GROUPS = [
        ("📂", "Data Management", [
            ("📤", "Upload & Data Quality"),
            ("🗄️", "Data Management"),
        ]),
        ("🔬", "Surveillance", [
            ("📊", "Resistance Overview"),
            ("🟥", "Resistance Heat Map"),
            ("🦠", "Pathogen Profile"),
            ("🏥", "HAI Profile"),
        ]),
        ("📈", "Analytics", [
            ("📉", "Trends"),
            ("🗺️", "Map Hotspots"),
            ("🧠", "Advanced Analytics"),
            ("⚠️", "Risk Assessment"),
            ("🔄", "Comparative Analysis"),
        ]),
        ("🌍", "One Health", [
            ("💊", "PPS Dashboard"),
            ("💉", "AMU Dashboard"),
            ("🧪", "AMC Dashboard"),
        ]),
        ("📋", "Reports & Tools", [
            ("🔔", "Alerts Dashboard"),
            ("🧬", "Antibiogram"),
            ("📁", "WHONET Export"),
            ("📄", "Report Export"),
        ]),
    ]
    if st.session_state.is_admin:
        _NAV_GROUPS.append(("⚙️", "Administration", [
            ("👥", "Admin - Users"),
            ("📂", "Admin - Datasets"),
        ]))

    # Determine which group the current page belongs to
    _active = st.session_state.active_page
    _active_group = None
    for _gicon, _gname, _gitems in _NAV_GROUPS:
        if any(pname == _active for _, pname in _gitems):
            _active_group = _gname
            break

    # ── Home button (standalone, always visible) ──
    _home_cls = "nav-item-btn nav-active" if _active == "Home" else "nav-item-btn"
    st.markdown(f'<div class="{_home_cls}">', unsafe_allow_html=True)
    if st.button("🏠  Home", key="nav_Home", use_container_width=True):
        st.session_state.active_page = "Home"
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    for _gicon, _gname, _gitems in _NAV_GROUPS:
        _is_open = (_gname == _active_group)
        with st.expander(f"{_gicon}  {_gname}", expanded=_is_open):
            for _picon, _pname in _gitems:
                _css_class = "nav-item-btn nav-active" if _pname == _active else "nav-item-btn"
                st.markdown(f'<div class="{_css_class}">', unsafe_allow_html=True)
                if st.button(f"{_picon}  {_pname}", key=f"nav_{_pname}", use_container_width=True):
                    st.session_state.active_page = _pname
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

    # ── Sign out at bottom ─────────────────────────────────────────
    st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
    if st.button("Sign out", use_container_width=True):
        st.session_state.authenticated = False
        st.session_state.user_email = None
        st.session_state.is_admin = False
        st.session_state.last_activity_time = None
        st.session_state.lab_name = None
        st.success("Logged out successfully")
        st.rerun()

    page = st.session_state.active_page

# ── Priority Pathogen Quick Filter (sidebar) ──────────────────────────
_pp_filter_path = os.path.join("data", "lookups", "priority_pathogens.json")
if os.path.exists(_pp_filter_path):
    with open(_pp_filter_path, "r") as _pf:
        _pp_lookup = json.load(_pf)
    _who_pp = _pp_lookup.get("WHO_2024_PRIORITY_PATHOGENS", {})
    _ghana_pp = _pp_lookup.get("GHANA_PRIORITY_PATHOGENS", {})
    _all_pp_names = sorted(set(
        p for pp_dict in [_who_pp, _ghana_pp]
        for tier_list in pp_dict.values()
        for p in tier_list
    ))
    if _all_pp_names:
        with st.sidebar.expander("Priority Pathogen Filter", expanded=False):
            pp_source = st.selectbox("List", ["WHO 2024", "Ghana"], key="pp_source")
            pp_dict = _who_pp if pp_source == "WHO 2024" else _ghana_pp
            pp_tier = st.selectbox("Tier", ["All"] + list(pp_dict.keys()), key="pp_tier")
            if pp_tier == "All":
                pp_names = sorted(set(p for lst in pp_dict.values() for p in lst))
            else:
                pp_names = pp_dict.get(pp_tier, [])
            st.session_state['priority_pathogen_filter'] = pp_names
            st.caption(f"{len(pp_names)} pathogens selected")

# ============================================================================
# DASHBOARD LANDING PAGE
# ============================================================================
if page == "Home":
    render_dashboard_page()

# ============================================================================
# PAGE 1: UPLOAD & DATA QUALITY
# ============================================================================
elif page == "Upload & Data Quality":
    st.header("Upload & Data Quality")
    
    col1, col2 = st.columns([2, 1])
    
    with col2:
        st.subheader("📋 Template Download")
        os.makedirs("templates", exist_ok=True)
        try:
            from src.validate_onehealth import create_unified_template
            unified_bytes = create_unified_template()
            st.download_button(
                label="⬇ Download Unified Template",
                data=unified_bytes,
                file_name="AMR_OneHealth_Unified_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Single workbook with all sheets: samples, ast_results, pps_survey, prescriptions, amu_data, amc_data",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error creating unified template: {e}")
        st.caption("One Excel file · 6 sheets · fill only what you need")
    
    with col1:
        st.subheader("Upload Data")
        uploaded_file = st.file_uploader(
            "Upload the unified Excel template (samples + AST + PPS + AMU + AMC sheets)",
            type=["xlsx", "xls"]
        )
        
        if uploaded_file:
            if st.button("Validate & Upload", type="primary"):
                with st.spinner("Validating all sheets..."):
                    # ── Core AMR (samples + ast_results) ────────────────
                    is_valid, errors, samples_df, ast_df = validate.validate_upload(uploaded_file)
                    
                    amr_saved = False
                    if is_valid:
                        if st.session_state.lab_name:
                            lab_values = samples_df['lab_name'].astype(str).str.strip().unique().tolist()
                            if len(lab_values) != 1 or lab_values[0].strip().lower() != st.session_state.lab_name.strip().lower():
                                st.error("Uploaded data must contain only your laboratory in the lab_name column.")
                                st.stop()

                        auto_interpreted_count = ast_df['auto_interpreted'].sum() if 'auto_interpreted' in ast_df.columns else 0
                        if auto_interpreted_count > 0:
                            st.info(f"🔬 Automated interpretation: {int(auto_interpreted_count)} AST results (CLSI/EUCAST)")

                        dataset_id = str(uuid.uuid4())[:8]
                        success, msg = db.save_dataset(
                            dataset_id,
                            uploaded_file.name.replace('.xlsx', ''),
                            samples_df,
                            ast_df,
                            uploaded_by=(st.session_state.user_email or "Anonymous")
                        )
                        if success:
                            st.success(f"✅ AMR data saved (ID: {dataset_id}) — {len(samples_df)} samples, {len(ast_df)} tests")
                            amr_saved = True
                        else:
                            st.error(f"Database error: {msg}")
                    elif errors:
                        # Only show AMR errors if sheets existed
                        uploaded_file.seek(0)
                        from openpyxl import load_workbook as _lwb
                        _wb = _lwb(uploaded_file, read_only=True)
                        _has_amr = 'samples' in _wb.sheetnames or 'ast_results' in _wb.sheetnames
                        _wb.close()
                        if _has_amr:
                            st.warning("⚠ AMR sheets (samples/ast_results) had errors:")
                            for i, error in enumerate(errors, 1):
                                st.markdown(f"  {i}. {error}")

                    # ── One Health sheets (PPS / AMU / AMC) ─────────────
                    uploaded_file.seek(0)
                    from src.validate_onehealth import validate_unified_upload
                    oh_ok, oh_errors, oh_result = validate_unified_upload(uploaded_file)

                    oh_any = bool(oh_result)
                    if oh_errors:
                        for e in oh_errors:
                            st.warning(f"⚠ {e}")

                    user_email = st.session_state.get('user_email', 'unknown')

                    # PPS
                    if 'pps_survey' in oh_result and 'pps_prescriptions' in oh_result:
                        import uuid as _uuid
                        survey_raw = oh_result['pps_survey']
                        # Handle both DataFrame and legacy dict
                        survey_df = survey_raw if isinstance(survey_raw, pd.DataFrame) else pd.DataFrame([survey_raw])
                        rx_df = oh_result['pps_prescriptions']
                        n_surveys = len(survey_df)
                        n_rx = len(rx_df)
                        rx_per = max(1, n_rx // n_surveys) if n_surveys else 0
                        rx_idx = 0
                        pps_ok_count = 0
                        for _, srow in survey_df.iterrows():
                            sid = f"PPS-{_uuid.uuid4().hex[:8]}"
                            end_idx = min(rx_idx + rx_per, n_rx)
                            chunk = rx_df.iloc[rx_idx:end_idx] if rx_idx < n_rx else pd.DataFrame()
                            rx_idx = end_idx
                            ok, msg = db.save_pps_survey(
                                sid,
                                str(srow['facility_name']),
                                str(srow['survey_date']),
                                str(srow.get('region', '')),
                                str(srow.get('district', '')),
                                int(srow.get('total_patients', 0)),
                                int(srow.get('patients_on_antibiotics', 0)),
                                chunk,
                                uploaded_by=user_email,
                            )
                            if ok:
                                pps_ok_count += 1
                        if pps_ok_count:
                            st.success(f"✅ PPS saved — {pps_ok_count} surveys, {n_rx} prescriptions")
                        else:
                            st.error("PPS save error: no surveys could be saved")

                    # AMU
                    if 'amu_data' in oh_result:
                        amu_df = oh_result['amu_data']
                        ok_a, msg_a = db.save_amu_records(amu_df, user_email)
                        if ok_a:
                            st.success(f"✅ AMU saved — {len(amu_df)} records")
                        else:
                            st.error(f"AMU save error: {msg_a}")

                    # AMC
                    if 'amc_data' in oh_result:
                        amc_df = oh_result['amc_data']
                        ok_c, msg_c = db.save_amc_records(amc_df, user_email)
                        if ok_c:
                            st.success(f"✅ AMC saved — {len(amc_df)} records")
                        else:
                            st.error(f"AMC save error: {msg_c}")

                    if amr_saved or oh_any:
                        st.balloons()
    
    st.markdown("---")
    
    # Show existing datasets
    st.subheader("Existing Datasets")
    datasets = _cached_all_datasets()
    # Hide admin-owned datasets from non-admin users
    config_admin_email, _ = _get_admin_config()
    admin_email = (config_admin_email or "").strip().lower()
    if not st.session_state.is_admin and admin_email:
        datasets = [ds for ds in datasets if (ds.get('uploaded_by') or '').strip().lower() != admin_email]
    if st.session_state.lab_name:
        datasets = [
            ds for ds in datasets
            if (ds.get('uploaded_by') or '').strip().lower() == (st.session_state.user_email or '').strip().lower()
        ]
    
    if datasets:
        for ds in datasets:
            col1, col2, col3 = st.columns([2, 1, 1])
            
            with col1:
                st.write(f"**{ds['dataset_name']}**")
                st.caption(f"ID: {ds['dataset_id']} | Uploaded: {ds['uploaded_at'][:10]}")
            
            with col2:
                st.metric("Samples", ds['rows_samples'])
                st.metric("Tests", ds['rows_tests'])
            
            with col3:
                if st.button("Delete", key=f"del_{ds['dataset_id']}"):
                    success, msg = db.delete_dataset(ds['dataset_id'])
                    if success:
                        st.success("Deleted!")
                        st.rerun()
                    else:
                        st.error(msg)
    else:
        st.info("No datasets uploaded yet. Upload one above.")

# ============================================================================
# PAGE 2: DATA MANAGEMENT
# ============================================================================
elif page == "Data Management":
    st.header("Data Management")
    st.markdown("Manage, review, and maintain your AMR surveillance datasets")

    # Get all datasets
    datasets = _cached_all_datasets()
    # Hide admin-owned datasets from non-admin users
    config_admin_email, _ = _get_admin_config()
    admin_email = (config_admin_email or "").strip().lower()
    if not st.session_state.is_admin and admin_email:
        datasets = [ds for ds in datasets if (ds.get('uploaded_by') or '').strip().lower() != admin_email]
    if st.session_state.lab_name:
        datasets = [
            ds for ds in datasets
            if (ds.get('uploaded_by') or '').strip().lower() == (st.session_state.user_email or '').strip().lower()
        ]

    if not datasets:
        st.info("No datasets available. Please upload data first on the 'Upload & Data Quality' page.")
    else:
        # Dataset selection — option values are raw dataset_ids; format_func shows the friendly label
        ds_label_by_id = {ds['dataset_id']: f"{ds['dataset_name']} (ID: {ds['dataset_id']})" for ds in datasets}
        selected_dataset_id = st.selectbox(
            "Select Dataset to Manage",
            list(ds_label_by_id.keys()),
            format_func=lambda ds_id: ds_label_by_id.get(ds_id, ds_id),
            key="data_mgmt_dataset"
        )
        if selected_dataset_id:
            st.session_state.active_dataset_id = selected_dataset_id
            st.success(f"Active dataset: {selected_dataset_id}")
elif page == "Admin - Datasets":
    st.header("Admin - Datasets")
    config_admin_email, _ = _get_admin_config()
    admin_email = (config_admin_email or "").strip().lower()

    all_datasets = _cached_all_datasets()

    main_datasets = db.get_main_datasets(country="Ghana")
    main_label_by_id = {d['dataset_id']: f"{d['dataset_name']} ({d['dataset_id']})" for d in main_datasets}
    st.selectbox(
        "National Main Dataset (Ghana)",
        list(main_label_by_id.keys()) or [""],
        format_func=lambda k: main_label_by_id.get(k, "None"),
        key="main_ds_select",
    )

    st.markdown("---")
    st.subheader("Mark a dataset as National Main")
    all_label_by_id = {d['dataset_id']: f"{d['dataset_name']} ({d['dataset_id']})" for d in all_datasets}
    target_id = st.selectbox(
        "Select dataset",
        list(all_label_by_id.keys()),
        format_func=lambda k: all_label_by_id.get(k, k),
        key="mark_main_select",
    )
    if st.button("Set as National Main (Ghana)", type="primary"):
        try:
            ok, msg = db.set_dataset_main(target_id, True, country="Ghana")
            if ok:
                st.success("Main dataset updated")
                st.rerun()
            else:
                st.error(msg)
        except Exception as e:
            st.error(f"Error: {e}")

    st.markdown("---")
    st.subheader("Merge User Dataset into National Main")
    user_datasets = [d for d in all_datasets if (d.get('uploaded_by') or '').strip().lower() != admin_email]
    user_label_by_id = {
        d['dataset_id']: f"{d['uploaded_by'] or 'Unknown'}: {d['dataset_name']} ({d['dataset_id']})"
        for d in user_datasets
    }
    src_id = st.selectbox(
        "Select user dataset",
        list(user_label_by_id.keys()) or [""],
        format_func=lambda k: user_label_by_id.get(k, "No user datasets"),
        key="merge_src_select",
    )
    # Refresh main choices
    main_datasets = db.get_main_datasets(country="Ghana")
    merge_main_label_by_id = {d['dataset_id']: f"{d['dataset_name']} ({d['dataset_id']})" for d in main_datasets}
    merge_target_id = st.selectbox(
        "Target main dataset",
        list(merge_main_label_by_id.keys()) or [""],
        format_func=lambda k: merge_main_label_by_id.get(k, "None"),
        key="merge_target_select",
    )

    if st.button("Merge into National Main", type="primary"):
        try:
            if not main_datasets:
                st.error("Please mark a dataset as National Main first")
            elif not src_id or not merge_target_id:
                st.error("Please select a source and a target dataset")
            else:
                ok, msg = db.merge_dataset_into_main(src_id, merge_target_id)
                if ok:
                    st.success(msg)
                else:
                    st.error(msg)
        except Exception as e:
            st.error(f"Error: {e}")

    st.markdown("---")
    st.subheader("KoboToolbox Sync")
    with st.expander("Import submissions from KoboToolbox", expanded=False):
        saved_form_id = load_kobo_form_id() or ""
        form_id = st.text_input("KoboToolbox Form ID", value=saved_form_id, key="kobo_form_id")

        col_a, col_b = st.columns([1, 1])
        with col_a:
            if st.button("Create KoboToolbox Form", key="kobo_create_form"):
                with st.spinner("Creating KoboToolbox form..."):
                    kobo = KoboToolboxManager()
                    ok, msg, form_data = kobo.create_amr_form()
                    if not ok:
                        st.error(msg)
                    else:
                        new_form_id = str(form_data.get("uid") or form_data.get("id") or "").strip()
                        if new_form_id:
                            save_ok, save_msg = save_kobo_form_id(new_form_id)
                            st.success(f"Form created. Form ID: {new_form_id}")
                            if not save_ok:
                                st.warning(save_msg)
                        else:
                            st.warning("Form created but Form ID was not returned. Please copy it manually from KoboToolbox.")
        with col_b:
            if st.button("Save Form ID", key="kobo_save_form_id"):
                if not form_id:
                    st.error("Please enter a Form ID to save.")
                else:
                    ok, msg = save_kobo_form_id(form_id)
                    if ok:
                        st.success(msg)
                    else:
                        st.error(msg)

        if st.button("Sync KoboToolbox Submissions", key="kobo_sync"):
            if not form_id:
                st.error("Please provide a KoboToolbox Form ID.")
            else:
                with st.spinner("Syncing data from KoboToolbox..."):
                    kobo = KoboToolboxManager()
                    ok, msg, submissions_df = kobo.fetch_submitted_data(form_id)
                    if not ok:
                        st.error(msg)
                    elif submissions_df is None or submissions_df.empty:
                        st.info("No submissions available for import.")
                    else:
                        samples_df, ast_df = kobo_submissions_to_frames(submissions_df)
                        sample_valid, sample_errors = validate.validate_samples(samples_df)
                        ast_valid, ast_errors = validate.validate_ast_results(ast_df, set(samples_df['sample_id'].dropna().astype(str)))
                        errors = sample_errors + ast_errors

                        if errors:
                            st.error("KoboToolbox data validation failed.")
                            for err in errors[:10]:
                                st.write(f"- {err}")
                        else:
                            # Deduplicate by existing sample_id and (isolate_id + antibiotic) combination
                            existing_samples_df = db.get_all_samples()
                            existing_ast_df = db.get_all_ast_results()

                            existing_sample_ids = set(existing_samples_df['sample_id'].dropna().astype(str)) if not existing_samples_df.empty else set()
                            
                            # Create set of (isolate_id + antibiotic) combinations to allow same isolate tested against different antibiotics
                            existing_ast_combos = set()
                            if not existing_ast_df.empty:
                                existing_ast_combos = set(
                                    (str(row['isolate_id']), str(row['antibiotic'])) 
                                    for idx, row in existing_ast_df.iterrows()
                                    if pd.notna(row['isolate_id']) and pd.notna(row['antibiotic'])
                                )

                            before_samples = len(samples_df)
                            before_tests = len(ast_df)

                            samples_df = samples_df[~samples_df['sample_id'].astype(str).isin(existing_sample_ids)]
                            
                            # Filter AST data by (isolate_id + antibiotic) combination
                            ast_df['_combo'] = list(zip(
                                ast_df['isolate_id'].astype(str),
                                ast_df['antibiotic'].astype(str)
                            ))
                            ast_df = ast_df[~ast_df['_combo'].isin(existing_ast_combos)]
                            ast_df = ast_df.drop(columns=['_combo'])

                            # Ensure AST rows correspond to remaining samples
                            ast_df = ast_df[ast_df['sample_id'].astype(str).isin(samples_df['sample_id'].astype(str))]

                            dropped_samples = before_samples - len(samples_df)
                            dropped_tests = before_tests - len(ast_df)

                            if samples_df.empty or ast_df.empty:
                                st.info("No new unique records found after deduplication.")
                                st.stop()

                            dataset_id = str(uuid.uuid4())[:8]
                            dataset_name = f"Kobo Sync {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                            success, save_msg = db.save_dataset(
                                dataset_id,
                                dataset_name,
                                samples_df,
                                ast_df,
                                uploaded_by=(st.session_state.user_email or "System")
                            )
                            if success:
                                st.success(f"KoboToolbox data imported as dataset {dataset_id}")
                                if dropped_samples or dropped_tests:
                                    st.info(f"Skipped duplicates: {dropped_samples} samples, {dropped_tests} tests")

                                # Option C — per-lab visibility:
                                # In addition to the consolidated national
                                # dataset above (visible to the admin), we
                                # also create one per-lab dataset slice so
                                # that when the lab itself logs in, the
                                # existing visibility filter on Data
                                # Management (which keys on `uploaded_by`
                                # == lab email) shows them only their own
                                # rows.  The per-lab slices are de-duped
                                # by the same logic as the national insert.
                                try:
                                    lab_email_map = _get_lab_email_mapping() or {}
                                    if 'lab_name' in samples_df.columns and lab_email_map:
                                        per_lab_created = 0
                                        for lab_name, lab_email in lab_email_map.items():
                                            lab_samples = samples_df[
                                                samples_df['lab_name'].astype(str) == str(lab_name)
                                            ]
                                            if lab_samples.empty:
                                                continue
                                            lab_ast = ast_df[
                                                ast_df['sample_id'].astype(str).isin(
                                                    lab_samples['sample_id'].astype(str)
                                                )
                                            ]
                                            if lab_ast.empty:
                                                continue
                                            lab_dataset_id = str(uuid.uuid4())[:8]
                                            lab_dataset_name = (
                                                f"{lab_name} — Kobo "
                                                f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
                                            )
                                            ok_lab, _ = db.save_dataset(
                                                lab_dataset_id,
                                                lab_dataset_name,
                                                lab_samples,
                                                lab_ast,
                                                uploaded_by=lab_email,
                                            )
                                            if ok_lab:
                                                per_lab_created += 1
                                        if per_lab_created:
                                            st.success(
                                                f"Created {per_lab_created} per-lab dataset "
                                                f"slice(s) — each lab now sees only its own rows."
                                            )
                                        # Bust the dataset list cache so the
                                        # new slices appear immediately on
                                        # the Data Management page.
                                        try:
                                            _cached_all_datasets.clear()
                                        except Exception:
                                            pass
                                except Exception:
                                    logger.exception("per-lab slice creation failed")
                            else:
                                st.error(save_msg)

        # Admin page continues without dataset preview block to avoid undefined variables

# ============================================================================
# PAGE 3: RESISTANCE OVERVIEW
# ============================================================================
elif page == "Resistance Overview":
    st.header("Resistance Overview")
    
    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    # Get data for active dataset only (cached loader applies lab-filter)
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # Filters
        st.sidebar.markdown("### Filters")

        def _uniq(df, col):
            return sorted(df[col].dropna().astype(str).unique().tolist()) if col in df.columns else []

        selected_labs = sidebar_multiselect_all(
            "Sentinel Site / Lab", _uniq(all_samples, "lab_name"), key="ro_lab"
        )
        selected_organisms = sidebar_multiselect_all(
            "Organism", _uniq(all_ast, "organism"), key="ro_org"
        )
        selected_antibiotics = sidebar_multiselect_all(
            "Antibiotic", _uniq(all_ast, "antibiotic"), key="ro_abx"
        )
        selected_categories = sidebar_multiselect_all(
            "Source Category", _uniq(all_samples, "source_category"), key="ro_cat"
        )
        selected_source_types = sidebar_multiselect_all(
            "Source Type", _uniq(all_samples, "source_type"), key="ro_src"
        )
        selected_site_types = sidebar_multiselect_all(
            "Site Type", _uniq(all_samples, "site_type"), key="ro_site"
        )
        selected_regions = sidebar_multiselect_all(
            "Region", _uniq(all_samples, "region"), key="ro_region"
        )
        selected_districts = sidebar_multiselect_all(
            "District", _uniq(all_samples, "district"), key="ro_district"
        )

        # Apply filters with validation
        if selected_categories and selected_regions and selected_districts:
            _mask = (
                (all_samples['source_category'].astype(str).isin(selected_categories)) &
                (all_samples['source_type'].astype(str).isin(selected_source_types)) &
                (all_samples['site_type'].astype(str).isin(selected_site_types)) &
                (all_samples['region'].astype(str).isin(selected_regions)) &
                (all_samples['district'].astype(str).isin(selected_districts))
            )
            if selected_labs and 'lab_name' in all_samples.columns:
                _mask = _mask & (all_samples['lab_name'].astype(str).isin(selected_labs))
            filtered_samples = all_samples[_mask]
        else:
            filtered_samples = all_samples
        
        if selected_organisms and selected_antibiotics:
            filtered_ast = all_ast[
                (all_ast['organism'].astype(str).isin(selected_organisms)) &
                (all_ast['antibiotic'].astype(str).isin(selected_antibiotics)) &
                (all_ast['sample_id'].astype(str).isin(filtered_samples['sample_id'].astype(str)))
            ]
        else:
            filtered_ast = all_ast[
                all_ast['sample_id'].astype(str).isin(filtered_samples['sample_id'].astype(str))
            ]
        
        if filtered_ast.empty:
            st.warning("No data matches the selected filters.")
        else:
            # Metrics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                resistant_count = (filtered_ast['result'] == 'R').sum()
                total_tests = len(filtered_ast)
                pct = resistant_count / total_tests * 100 if total_tests > 0 else 0
                st.metric("Resistance %", f"{pct:.1f}%", delta=f"{resistant_count}/{total_tests}")
            
            with col2:
                st.metric("Total Tests", total_tests)
            
            with col3:
                st.metric("Unique Samples", filtered_samples['sample_id'].nunique())
            
            with col4:
                st.metric("Unique Organisms", filtered_ast['organism'].nunique())
            
            # ── Sentinel Phenotype & MDRO KPI Row ──────────────────────
            from src.analytics import detect_sentinel_phenotypes, calculate_mdro_incidence
            _sentinel = detect_sentinel_phenotypes(filtered_ast)
            _mdro = calculate_mdro_incidence(filtered_ast)

            if _sentinel or _mdro.get("mdr_isolates", 0) > 0:
                st.markdown("""
                <style>
                .sentinel-row { display:flex; gap:0.6rem; flex-wrap:wrap; margin-bottom:1rem; }
                .sentinel-card {
                    flex:1; min-width:110px; padding:0.7rem 0.6rem; border-radius:10px;
                    text-align:center; border:1px solid rgba(0,0,0,0.08);
                }
                .sentinel-card.crit { background:linear-gradient(135deg,#5c1e1e,#3d1010); }
                .sentinel-card.high { background:linear-gradient(135deg,#5c4a1e,#3d3010); }
                .sentinel-card.info { background:linear-gradient(135deg,#1e3a5f,#0d253f); }
                .sentinel-card .sv { font-size:1.5rem; font-weight:700; color:#fff; }
                .sentinel-card .sl { font-size:0.72rem; color:rgba(255,255,255,0.65); margin-top:0.15rem; }
                </style>
                """, unsafe_allow_html=True)

                cards_html = '<div class="sentinel-row">'
                # MDRO card
                _mdr_n = _mdro.get("mdr_isolates", 0)
                _mdr_pct = _mdro.get("mdr_rate_pct", 0)
                cards_html += (
                    f'<div class="sentinel-card {"crit" if _mdr_pct >= 30 else "info"}">'
                    f'<div class="sv">{_mdr_n}</div>'
                    f'<div class="sl">MDR Isolates ({_mdr_pct:.0f}%)</div></div>'
                )
                # Sentinel phenotype cards (top 5)
                for sp in _sentinel[:5]:
                    tier_class = "crit" if sp["who_tier"] == "Critical" else "high"
                    cards_html += (
                        f'<div class="sentinel-card {tier_class}">'
                        f'<div class="sv">{sp["isolate_count"]}</div>'
                        f'<div class="sl">{sp["code"]} ({sp["resistance_rate"]:.0f}%)</div></div>'
                    )
                cards_html += '</div>'
                st.markdown(cards_html, unsafe_allow_html=True)

                with st.expander("Sentinel Phenotype Details", expanded=False):
                    if _sentinel:
                        _sp_df = pd.DataFrame([{
                            "Phenotype": s["code"],
                            "Description": s["label"],
                            "WHO Tier": s["who_tier"],
                            "Positive Isolates": s["isolate_count"],
                            "Tested": s["total_tested"],
                            "Rate %": s["resistance_rate"],
                        } for s in _sentinel])
                        st.dataframe(_sp_df, use_container_width=True, hide_index=True)
                    else:
                        st.info("No WHO sentinel phenotypes detected in the current data.")

            st.markdown("---")
            
            # Charts
            col1, col2 = st.columns(2)
            
            with col1:
                st.plotly_chart(
                    plots.plot_top_antibiotics(filtered_ast),
                    use_container_width=True
                )
            
            with col2:
                st.plotly_chart(
                    plots.plot_resistance_distribution(filtered_ast),
                    use_container_width=True
                )
            
            st.plotly_chart(
                plots.plot_resistance_by_category(filtered_ast, filtered_samples),
                use_container_width=True
            )
            
            st.plotly_chart(
                plots.plot_resistance_by_source_type(filtered_ast, filtered_samples),
                use_container_width=True
            )
            
            st.info("📊 For a detailed organism × antibiotic resistance matrix, visit the **Resistance Heat Map** page.")
            
            st.markdown("---")
            
            # Co-resistance patterns
            st.subheader("🔗 Co-Resistance Patterns")
            
            co_resistance = plots.get_co_resistance_patterns(filtered_ast)
            if not co_resistance.empty:
                st.dataframe(co_resistance, use_container_width=True)
            else:
                st.info("No co-resistance patterns detected")
            
            st.markdown("---")
            
            # Resistance Mechanisms
            st.subheader("🧬 Resistance Mechanisms")
            
            col1, col2 = st.columns([2, 1])
            with col1:
                mech_fig = plots.plot_resistance_mechanisms(filtered_ast)
                st.plotly_chart(mech_fig, use_container_width=True)
            
            with col2:
                from src.analytics import detect_resistance_mechanisms
                mechanisms_df = detect_resistance_mechanisms(filtered_ast)
                if not mechanisms_df.empty:
                    st.dataframe(mechanisms_df[['isolate_id', 'organism', 'resistance_mechanism', 'confidence']].head(50), use_container_width=True)
                else:
                    st.info("No resistance mechanisms detected")
            
            st.markdown("---")
            
            # Cross-resistance patterns
            st.subheader("🔄 Cross-Resistance Patterns")
            
            col1, col2 = st.columns([2, 1])
            with col1:
                cross_fig = plots.plot_cross_resistance_patterns(filtered_ast)
                st.plotly_chart(cross_fig, use_container_width=True)
            
            with col2:
                from src.analytics import detect_cross_resistance
                cross_df = detect_cross_resistance(filtered_ast)
                if not cross_df.empty:
                    st.dataframe(cross_df[['isolate_id', 'organism', 'antibiotic_class', 'cross_resistance_level']].head(50), use_container_width=True)
                else:
                    st.info("No cross-resistance patterns detected")
            
            st.markdown("---")
            
            # Data preview
            st.subheader("Data Preview")
            display_df = filtered_ast[['sample_id', 'organism', 'antibiotic', 'result', 'method', 'test_date']].head(100)
            st.dataframe(display_df, use_container_width=True)

# ============================================================================
# PAGE: RESISTANCE HEAT MAP
# ============================================================================
elif page == "Resistance Heat Map":
    render_heatmap_page()

# ============================================================================
# PAGE: PATHOGEN PROFILE
# ============================================================================
elif page == "Pathogen Profile":
    render_pathogen_profile_page()

# ============================================================================
# PAGE: HAI PROFILE
# ============================================================================
elif page == "HAI Profile":
    render_hai_page()

# ============================================================================
# PAGE 4: TRENDS
# ============================================================================
elif page == "Trends":
    st.header("Resistance Trends")
    
    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # Filters
        st.sidebar.markdown("### Trend Filters")

        def _uniq(df, col):
            return sorted(df[col].dropna().astype(str).unique().tolist()) if col in df.columns else []

        _sel_labs_tr = sidebar_multiselect_all(
            "Sentinel Site / Lab", _uniq(all_samples, "lab_name"), key="tr_lab"
        )
        if 'lab_name' in all_samples.columns and _sel_labs_tr != _uniq(all_samples, "lab_name"):
            all_samples = all_samples[all_samples['lab_name'].astype(str).isin(_sel_labs_tr)]
            all_ast = all_ast[all_ast['sample_id'].astype(str).isin(all_samples['sample_id'].astype(str))]

        selected_organisms = sidebar_multiselect_all(
            "Organism", _uniq(all_ast, "organism"), key="tr_org"
        )
        selected_antibiotics = sidebar_multiselect_all(
            "Antibiotic", _uniq(all_ast, "antibiotic"), key="tr_abx"
        )

        # Time aggregation
        time_agg = st.sidebar.selectbox("Time Aggregation", ["Monthly", "Quarterly", "Yearly"], key="tr_time")
        
        # Apply filters
        if selected_organisms and selected_antibiotics:
            filtered_ast = all_ast[
                (all_ast['organism'].astype(str).isin(selected_organisms)) &
                (all_ast['antibiotic'].astype(str).isin(selected_antibiotics))
            ]
        else:
            filtered_ast = all_ast
        
        if filtered_ast.empty:
            st.warning("No data matches the selected filters. Try selecting different filters.")
        else:
            # Overall trend
            st.plotly_chart(
                plots.plot_resistance_trends(filtered_ast, time_agg),
                use_container_width=True
            )
            
            st.markdown("---")
            
            # Show summary statistics
            st.subheader("Trend Summary")
            
            col1, col2, col3 = st.columns(3)
            
            # Calculate oldest and newest dates
            filtered_ast['test_date_parsed'] = pd.to_datetime(filtered_ast['test_date'], errors='coerce')
            valid_dates = filtered_ast[filtered_ast['test_date_parsed'].notna()]['test_date_parsed']
            
            if not valid_dates.empty:
                earliest = valid_dates.min().strftime('%Y-%m-%d')
                latest = valid_dates.max().strftime('%Y-%m-%d')
                
                with col1:
                    st.metric("Earliest Test", earliest)
                with col2:
                    st.metric("Latest Test", latest)
                with col3:
                    st.metric("Date Range", f"{len(valid_dates)} tests")
            
            st.markdown("---")
            
            # Data preview
            st.subheader("Recent Test Data")
            display_df = filtered_ast[['test_date', 'organism', 'antibiotic', 'result', 'sample_id']].sort_values('test_date', ascending=False).head(100)
            st.dataframe(display_df, use_container_width=True)

# ============================================================================
# PAGE 5: MAP HOTSPOTS
# ============================================================================
elif page == "Map Hotspots":
    st.header("Geographic Hotspots & Regional Analysis")
    
    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    # Sentinel Site / Lab sidebar filter
    _map_labs = sorted(all_samples['lab_name'].dropna().astype(str).unique().tolist()) if 'lab_name' in all_samples.columns else []
    if _map_labs:
        st.sidebar.markdown("### Map Filters")
        _sel_map_labs = sidebar_multiselect_all("Sentinel Site / Lab", _map_labs, key="map_lab")
        if _sel_map_labs != _map_labs:
            all_samples = all_samples[all_samples['lab_name'].astype(str).isin(_sel_map_labs)]
            all_ast = all_ast[all_ast['sample_id'].astype(str).isin(all_samples['sample_id'].astype(str))]

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # Check if geographic data exists
        samples_with_coords = all_samples[
            all_samples['latitude'].notna() & 
            all_samples['longitude'].notna()
        ]
        has_coords = len(samples_with_coords) > 0
        
        if has_coords:
            # Enhanced Interactive Folium Map
            st.subheader("📍 Interactive Ghana Map - Resistance Hotspots")
            st.markdown(f"**{len(samples_with_coords)}** samples with geographic coordinates | Interactive map below")
            
            try:
                # Import the enhanced mapping module
                from src import ghana_map
                from streamlit_folium import folium_static
                
                # Create and display interactive Folium map
                m = ghana_map.create_interactive_ghana_map(samples_with_coords, all_ast)
                
                # Display using folium_static with full width
                folium_static(m, width=1200, height=600)
                
                # Map instructions
                with st.expander("📚 How to Use the Interactive Map", expanded=False):
                    st.markdown("""
                    **Data Point Display:**
                    - Each colored circle represents a sample location
                    - Circle size = number of tests from that location
                    - Circle color = resistance rate:
                      - **Red**: High resistance (>50%)
                      - **Orange**: Medium resistance (30-50%)
                      - **Green**: Low resistance (<30%)
                    
                    **How to Interact:**
                    - **Hover** over circles to see detailed information
                    - **Click** on circles for popup with detailed data
                    - **Drag** to pan around the map
                    - **Scroll** to zoom in/out
                    - **Double-click** to zoom to location
                    """)
                
            except Exception as e:
                st.warning(f"Map rendering issue: {str(e)}")
                st.info("Displaying data in tabular format...")
                
                # Fallback display: Show data as table
                display_cols = ['sample_id', 'district', 'region', 'latitude', 'longitude']
                available_cols = [col for col in display_cols if col in samples_with_coords.columns]
                
                if available_cols:
                    st.dataframe(
                        samples_with_coords[available_cols].head(50),
                        use_container_width=True
                    )
                    st.caption(f"Showing first 50 of {len(samples_with_coords)} samples")
            
            st.markdown("---")
        else:
            st.info("📍 No geographic coordinates in uploaded data. Add latitude/longitude to samples sheet to enable location mapping.")
        
        # Regional Analysis
        st.subheader("Resistance by Region")
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.plotly_chart(
                plots.plot_resistance_by_region(all_ast, all_samples),
                use_container_width=True
            )
        
        with col2:
            st.plotly_chart(
                plots.plot_resistance_percentage_by_region(all_ast, all_samples),
                use_container_width=True
            )
        
        st.markdown("---")
        
        # District-level Analysis
        st.subheader("District-Level Resistance Hotspots")
        
        # Detailed district analysis
        st.plotly_chart(
            plots.plot_resistance_by_district_detailed(all_ast, all_samples, top_n=15),
            use_container_width=True
        )
        
        st.markdown("---")
        
        # Top districts table
        st.subheader("Top Districts Summary Table")
        
        top_districts = plots.get_resistance_by_district_detailed(all_ast, all_samples)
        
        if not top_districts.empty:
            st.dataframe(
                top_districts[['district', 'region', 'total_tests', 'susceptible', 'intermediate', 'resistant', 'percent_resistant']],
                use_container_width=True,
                height=500
            )
        else:
            st.info("No district data available.")
        
        st.markdown("---")
        
        # Surveillance alerts
        st.subheader("Surveillance Alerts & Warnings")
        
        alerts = plots.get_surveillance_alerts(all_ast, all_samples)
        
        if alerts:
            for alert in alerts:
                if alert['severity'] == 'HIGH':
                    st.error(f"**{alert['severity']}**: {alert['message']}")
                elif alert['severity'] == 'MEDIUM':
                    st.warning(f"**{alert['severity']}**: {alert['message']}")
                else:
                    st.info(f"**{alert['severity']}**: {alert['message']}")
        else:
            st.success("No critical alerts detected")



# ============================================================================
# PAGE 6: ADVANCED ANALYTICS
# ============================================================================
elif page == "Advanced Analytics":
    st.header("Advanced Analytics & Insights")
    
    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    # Sentinel Site / Lab sidebar filter
    _aa_labs = sorted(all_samples['lab_name'].dropna().astype(str).unique().tolist()) if 'lab_name' in all_samples.columns else []
    if _aa_labs:
        st.sidebar.markdown("### Analytics Filters")
        _sel_lab_opts_aa = sidebar_multiselect_all("Sentinel Site / Lab", _aa_labs, key="aa_lab")
        if _sel_lab_opts_aa != _aa_labs:
            all_samples = all_samples[all_samples['lab_name'].astype(str).isin(_sel_lab_opts_aa)]
            all_ast = all_ast[all_ast['sample_id'].astype(str).isin(all_samples['sample_id'].astype(str))]

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # Tab selection
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "Statistics",
            "Trends & Forecasts",
            "Emerging Patterns",
            "Antibiotic Insights",
            "Data Quality"
        ])
        
        # TAB 1: STATISTICS
        with tab1:
            st.subheader("Comprehensive Resistance Statistics")
            
            col1, col2, col3 = st.columns(3)
            
            # Overall stats
            stats = analytics.calculate_resistance_statistics(all_ast)
            
            with col1:
                st.metric("Resistance Rate", f"{stats.get('resistance_rate', 0):.1f}%")
            with col2:
                st.metric("Tests Analyzed", stats.get('total_tests', 0))
            with col3:
                st.metric("Organisms", all_ast['organism'].nunique())
            
            # Detailed breakdown
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.info(f"**Susceptible**: {stats.get('susceptible_count', 0)} ({stats.get('susceptible_rate', 0):.1f}%)")
            with col2:
                st.warning(f"**Intermediate**: {stats.get('intermediate_count', 0)} ({stats.get('intermediate_rate', 0):.1f}%)")
            with col3:
                st.error(f"**Resistant**: {stats.get('resistant_count', 0)} ({stats.get('resistance_rate', 0):.1f}%)")
            
            st.markdown("---")
            
            # Trend direction
            st.subheader("Trend Analysis")
            trend_info = analytics.calculate_trend_direction(all_ast)
            
            if trend_info:
                col1, col2 = st.columns(2)
                
                with col1:
                    trend = trend_info.get('trend', 'stable').upper()
                    risk = trend_info.get('risk_level', 'LOW')
                    
                    if trend == 'INCREASING':
                        st.error(f"**{trend}** - Risk: {risk}")
                    elif trend == 'DECREASING':
                        st.success(f"**{trend}** - Risk: {risk}")
                    else:
                        st.info(f"**{trend}** - Risk: {risk}")
                
                with col2:
                    st.metric(
                        "Change in Resistance",
                        f"{trend_info.get('change_percentage', 0):.2f}%",
                        delta=f"{trend_info.get('change_percentage', 0):.2f}%"
                    )
            
            st.markdown("---")
            
            # Organism comparison
            st.subheader("Organism Resistance Comparison")
            org_comparison = analytics.compare_organisms(all_ast)
            if not org_comparison.empty:
                st.dataframe(org_comparison, use_container_width=True)
            
            st.markdown("---")
            
            # Antibiotic comparison
            st.subheader("Antibiotic Efficacy Comparison")
            abx_comparison = analytics.compare_antibiotics(all_ast)
            if not abx_comparison.empty:
                st.dataframe(abx_comparison, use_container_width=True)
        
        # TAB 2: TRENDS & FORECASTS
        with tab2:
            st.subheader("Resistance Trends & Forecasting")
            
            col1, col2 = st.columns(2)
            
            with col1:
                forecast_periods = st.slider("Forecast Periods (months)", 1, 12, 3)
            
            with col2:
                st.empty()
            
            # Forecast
            forecast = analytics.forecast_resistance_trend(all_ast, forecast_periods)
            
            if 'forecasts' in forecast:
                st.info(f"Trend: {forecast['forecasts'][0]['trend'].upper()}")
                
                forecast_df = pd.DataFrame(forecast['forecasts'])
                st.dataframe(forecast_df, use_container_width=True)
                
                # Visualization
                fig = px.line(
                    forecast_df,
                    x='months_ahead',
                    y='predicted_resistance_rate',
                    markers=True,
                    title='Forecasted Resistance Rate',
                    labels={'months_ahead': 'Months Ahead', 'predicted_resistance_rate': 'Predicted Resistance %'}
                )
                st.plotly_chart(fig, use_container_width=True)
            elif 'error' in forecast:
                st.warning(f"{forecast['error']}")
        
        # TAB 3: EMERGING PATTERNS
        with tab3:
            st.subheader("Emerging Resistance Patterns")
            
            emerging = analytics.identify_emerging_resistance(all_ast, all_samples)
            
            if emerging:
                emerging_df = pd.DataFrame(emerging)
                st.dataframe(emerging_df, use_container_width=True)
                
                st.warning(f"🚨 {len(emerging)} emerging resistance patterns detected in the last 3 months")
            else:
                st.success("No concerning emerging patterns detected")
        
        # TAB 4: ANTIBIOTIC INSIGHTS
        with tab4:
            st.subheader("Antibiotic Recommendations")
            
            recommendations = analytics.generate_antibiotic_recommendations(all_ast)
            
            if recommendations:
                # Priority breakdown
                col1, col2, col3, col4 = st.columns(4)
                
                preferred = len([r for r in recommendations if r['priority'] == 1])
                good = len([r for r in recommendations if r['priority'] == 2])
                caution = len([r for r in recommendations if r['priority'] == 3])
                avoid = len([r for r in recommendations if r['priority'] == 4])
                
                with col1:
                    st.success(f"**Preferred**: {preferred}")
                with col2:
                    st.info(f"**Good**: {good}")
                with col3:
                    st.warning(f"**Caution**: {caution}")
                with col4:
                    st.error(f"**Avoid**: {avoid}")
                
                st.markdown("---")
                
                # Detailed recommendations
                rec_df = pd.DataFrame(recommendations).sort_values('priority')
                st.dataframe(rec_df, use_container_width=True)
        
        # TAB 5: DATA QUALITY
        with tab5:
            st.subheader("Surveillance System Quality Metrics")
            
            quality = analytics.assess_data_quality(all_samples, all_ast)
            kpis = analytics.calculate_kpis(all_samples, all_ast)
            
            if quality:
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Samples", quality.get('total_samples', 0))
                with col2:
                    st.metric("Total Tests", quality.get('total_tests', 0))
                with col3:
                    st.metric("Completeness", f"{quality.get('completeness_score', 0):.1f}%")
                with col4:
                    st.metric("Geographic Coverage", f"{quality.get('samples_with_coordinates', 0)} samples")
                
                st.markdown("---")
                
                if quality.get('data_quality_issues'):
                    st.warning("**Data Quality Issues Detected:**")
                    for issue in quality['data_quality_issues']:
                        st.warning(f"• {issue}")
                else:
                    st.success("No data quality issues detected")
                
                st.markdown("---")
                
                # KPIs
                st.subheader("Key Performance Indicators")
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Tests/Sample", kpis.get('tests_per_sample', 0))
                with col2:
                    st.metric("Organisms ID'd", kpis.get('organisms_identified', 0))
                with col3:
                    st.metric("Antibiotics Tested", kpis.get('antibiotics_tested', 0))


# ============================================================================
# PAGE 7: RISK ASSESSMENT
# ============================================================================
elif page == "Risk Assessment":
    st.header("Risk Assessment & Alerts")
    
    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    # Sentinel Site / Lab sidebar filter
    _risk_labs = sorted(all_samples['lab_name'].dropna().astype(str).unique().tolist()) if 'lab_name' in all_samples.columns else []
    if _risk_labs:
        st.sidebar.markdown("### Risk Filters")
        _sel_lab_opts_risk = sidebar_multiselect_all("Sentinel Site / Lab", _risk_labs, key="risk_lab")
        if _sel_lab_opts_risk != _risk_labs:
            all_samples = all_samples[all_samples['lab_name'].astype(str).isin(_sel_lab_opts_risk)]
            all_ast = all_ast[all_ast['sample_id'].astype(str).isin(all_samples['sample_id'].astype(str))]

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # Tabs
        tab1, tab2 = st.tabs(["Risk Scores", "Resistance Burden"])
        
        # TAB 1: ORGANISM RISK SCORES
        with tab1:
            st.subheader("Organism Risk Scores")
            
            # Risk threshold slider
            risk_threshold = st.slider("Show organisms with resistance rate ≥", 0, 100, 50, step=1)
            
            high_risk = analytics.get_high_risk_organisms(all_ast, risk_threshold)
            
            if high_risk:
                for risk_item in high_risk:
                    with st.expander(f"{risk_item['organism']} - Risk: {risk_item['risk_level']} ({risk_item['risk_score']}/100)"):
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.metric("Risk Score", risk_item['risk_score'])
                        with col2:
                            st.metric("Resistance Rate", f"{risk_item['resistance_rate']:.1f}%")
                        with col3:
                            st.metric("Tests", risk_item['test_count'])
                        
                        st.markdown("**Risk Factors:**")
                        for factor in risk_item['risk_factors']:
                            st.write(f"• {factor}")
                        
                        # Recommendation
                        if risk_item['risk_level'] == 'CRITICAL':
                            st.error("**Urgent intervention required** - Consider alternative treatment options")
                        elif risk_item['risk_level'] == 'HIGH':
                            st.warning("**Enhanced surveillance** - Monitor trends closely")
                        else:
                            st.info("**Monitor** - Continue standard surveillance")
            else:
                st.success(f"No organisms above risk threshold ({risk_threshold})")

            # Detailed single-organism assessment
            st.markdown("---")
            st.subheader("Detailed Organism Assessment")
            organisms = sorted(all_ast['organism'].dropna().astype(str).unique().tolist())
            if organisms:
                selected_org = st.selectbox("Select Organism for Detail", organisms, key="risk_org_detail")
                if selected_org:
                    org_risk = analytics.calculate_organism_risk_score(all_ast, selected_org)
                    if org_risk:
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Risk Score", org_risk['risk_score'])
                        with col2:
                            st.metric("Risk Level", org_risk['risk_level'])
                        with col3:
                            st.metric("Resistance Rate", f"{org_risk['resistance_rate']:.1f}%")
                        with col4:
                            st.metric("Tests", org_risk['test_count'])
                        st.markdown("**Risk Factors:**")
                        for factor in org_risk['risk_factors']:
                            st.write(f"• {factor}")
                        if org_risk['risk_level'] == 'CRITICAL':
                            st.error("**CRITICAL** — Implement enhanced infection control, review treatment guidelines, consider alternative antimicrobials, report to national health authorities.")
                        elif org_risk['risk_level'] == 'HIGH':
                            st.warning("**HIGH** — Increase surveillance frequency, review empiric treatment protocols, consider antimicrobial stewardship interventions.")
                        else:
                            st.info("**MODERATE/LOW** — Continue routine surveillance, monitor for changes in resistance patterns.")
        
        # TAB 2: RESISTANCE BURDEN
        with tab2:
            st.subheader("Overall Resistance Burden")
            
            burden = analytics.calculate_resistance_burden(all_samples, all_ast)
            
            if burden:
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Resistant Tests", burden.get('total_resistant_tests', 0))
                with col2:
                    st.metric("Overall Resistance Rate", f"{burden.get('overall_resistance_rate', 0):.1f}%")
                with col3:
                    st.metric("Total Tests", burden.get('total_tests', 0))
                
                st.markdown("---")
                
                # Public health impact
                impact = burden.get('public_health_impact', '')
                if 'CRITICAL' in impact:
                    st.error(f"{impact}")
                elif 'HIGH' in impact:
                    st.warning(f"{impact}")
                else:
                    st.info(f"{impact}")
                
                st.markdown("---")
                
                # By category
                if burden.get('resistance_by_category'):
                    st.subheader("Resistance by Source Category")
                    
                    category_data = pd.DataFrame(
                        list(burden['resistance_by_category'].items()),
                        columns=['Category', 'Resistance Rate (%)']
                    )
                    
                    fig = px.bar(
                        category_data,
                        x='Category',
                        y='Resistance Rate (%)',
                        color='Resistance Rate (%)',
                        color_continuous_scale='RdYlGn_r',
                        title='Resistance Burden by Source Category'
                    )
                    st.plotly_chart(fig, use_container_width=True)


# ============================================================================
# PAGE 8: COMPARATIVE ANALYSIS
# ============================================================================
elif page == "Comparative Analysis":
    st.header("Comparative Analysis")
    st.markdown("Compare resistance patterns across different categories, time periods, and sources")

    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    # Sentinel Site / Lab sidebar filter
    _comp_labs = sorted(all_samples['lab_name'].dropna().astype(str).unique().tolist()) if 'lab_name' in all_samples.columns else []
    if _comp_labs:
        st.sidebar.markdown("### Comparison Filters")
        _sel_lab_opts_comp = sidebar_multiselect_all("Sentinel Site / Lab", _comp_labs, key="comp_lab")
        if _sel_lab_opts_comp != _comp_labs:
            all_samples = all_samples[all_samples['lab_name'].astype(str).isin(_sel_lab_opts_comp)]
            all_ast = all_ast[all_ast['sample_id'].astype(str).isin(all_samples['sample_id'].astype(str))]

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # Analysis type selection
        analysis_type = st.selectbox(
            "Select Comparison Type",
            ["Category Comparison", "Time Period Comparison", "Source Type Comparison", "Multi-Parameter Comparison", "Cross-Variable Comparison", "Custom Comparison"],
            key="comparison_type"
        )

        st.markdown("---")

        if analysis_type == "Category Comparison":
            st.subheader("Category Comparison")

            # Get available categories
            available_categories = sorted(all_samples['source_category'].dropna().unique())

            if len(available_categories) >= 2:
                col1, col2 = st.columns(2)

                with col1:
                    category_a = st.selectbox(
                        "Select First Category",
                        available_categories,
                        index=0 if len(available_categories) > 0 else None,
                        key="category_a"
                    )

                with col2:
                    # Filter out the selected category A from options for category B
                    remaining_categories = [cat for cat in available_categories if cat != category_a]
                    category_b = st.selectbox(
                        "Select Second Category",
                        remaining_categories,
                        index=0 if len(remaining_categories) > 0 else None,
                        key="category_b"
                    )

                if st.button("Compare Categories", key="compare_categories"):
                    # Get data for each selected category
                    cat_a_samples = all_samples[all_samples['source_category'] == category_a]
                    cat_b_samples = all_samples[all_samples['source_category'] == category_b]

                    cat_a_ast = all_ast[all_ast['sample_id'].isin(cat_a_samples['sample_id'])]
                    cat_b_ast = all_ast[all_ast['sample_id'].isin(cat_b_samples['sample_id'])]

                    if not cat_a_ast.empty and not cat_b_ast.empty:
                        # Create comparison metrics
                        col1, col2, col3 = st.columns(3)

                        with col1:
                            cat_a_resistance = (cat_a_ast['result'] == 'R').sum() / len(cat_a_ast) * 100
                            st.metric(f"{category_a} Resistance Rate", f"{cat_a_resistance:.1f}%", delta=f"{(cat_a_ast['result'] == 'R').sum()}/{len(cat_a_ast)}")

                        with col2:
                            cat_b_resistance = (cat_b_ast['result'] == 'R').sum() / len(cat_b_ast) * 100
                            st.metric(f"{category_b} Resistance Rate", f"{cat_b_resistance:.1f}%", delta=f"{(cat_b_ast['result'] == 'R').sum()}/{len(cat_b_ast)}")

                        with col3:
                            diff = cat_a_resistance - cat_b_resistance
                            st.metric(f"Difference ({category_a} - {category_b})", f"{diff:+.1f}%")

                        # Side-by-side charts
                        st.markdown("### Resistance Distribution Comparison")

                        col1, col2 = st.columns(2)

                        with col1:
                            st.markdown(f"**{category_a} Sources**")
                            try:
                                cat_a_fig = plots.plot_resistance_distribution(cat_a_ast)
                                st.plotly_chart(cat_a_fig, use_container_width=True)
                            except Exception as e:
                                st.warning(f"Unable to generate {category_a} sources chart: {str(e)}")

                        with col2:
                            st.markdown(f"**{category_b} Sources**")
                            try:
                                cat_b_fig = plots.plot_resistance_distribution(cat_b_ast)
                                st.plotly_chart(cat_b_fig, use_container_width=True)
                            except Exception as e:
                                st.warning(f"Unable to generate {category_b} sources chart: {str(e)}")

                        # Top antibiotics comparison
                        st.markdown("### Top Antibiotics Comparison")

                        try:
                            cat_a_top = plots.get_antibiotic_resistance_rates(cat_a_ast)
                            cat_b_top = plots.get_antibiotic_resistance_rates(cat_b_ast)

                            if not cat_a_top.empty and not cat_b_top.empty:
                                # Create comparison chart
                                comparison_data = []

                                # Get top 10 antibiotics from both
                                all_antibiotics = set(cat_a_top.head(10)['antibiotic']) | set(cat_b_top.head(10)['antibiotic'])

                                for antibiotic in all_antibiotics:
                                    cat_a_rate = cat_a_top.loc[cat_a_top['antibiotic'] == antibiotic, 'resistance_rate'].iloc[0] if antibiotic in cat_a_top['antibiotic'].values else 0
                                    cat_b_rate = cat_b_top.loc[cat_b_top['antibiotic'] == antibiotic, 'resistance_rate'].iloc[0] if antibiotic in cat_b_top['antibiotic'].values else 0

                                    comparison_data.append({
                                        'Antibiotic': antibiotic,
                                        category_a: cat_a_rate,
                                        category_b: cat_b_rate
                                    })

                                comparison_df = pd.DataFrame(comparison_data)

                                fig = px.bar(comparison_df, x='Antibiotic', y=[category_a, category_b],
                                           title=f'Antibiotic Resistance: {category_a} vs {category_b}',
                                           barmode='group', color_discrete_sequence=['#FF6B6B', '#4ECDC4'])
                                fig.update_layout(xaxis_tickangle=-45)
                                st.plotly_chart(fig, use_container_width=True)
                            else:
                                st.info("No antibiotic resistance data available for comparison.")
                        except Exception as e:
                            st.warning(f"Unable to generate antibiotic comparison: {str(e)}")

                    else:
                        st.warning(f"Insufficient data for {category_a} vs {category_b} comparison. Need both categories to have AST results.")
            else:
                st.warning("Need data from at least 2 categories for comparison.")

        elif analysis_type == "Time Period Comparison":
            st.subheader("📅 Time Period Comparison")

            if 'test_date' in all_ast.columns:
                # Get date range
                dates = pd.to_datetime(all_ast['test_date'].dropna())
                min_date = dates.min()
                max_date = dates.max()

                col1, col2 = st.columns(2)

                with col1:
                    period1 = st.date_input("First Period Start-End", value=(min_date, min_date + (max_date - min_date)/2), key="period1")
                    if len(period1) == 2:
                        period1_start, period1_end = period1

                with col2:
                    period2 = st.date_input("Second Period Start-End", value=(min_date + (max_date - min_date)/2, max_date), key="period2")
                    if len(period2) == 2:
                        period2_start, period2_end = period2

                if st.button("Compare Periods", key="compare_periods"):
                    # Filter data for each period
                    period1_data = all_ast[
                        (pd.to_datetime(all_ast['test_date']) >= pd.Timestamp(period1_start)) &
                        (pd.to_datetime(all_ast['test_date']) <= pd.Timestamp(period1_end))
                    ]

                    period2_data = all_ast[
                        (pd.to_datetime(all_ast['test_date']) >= pd.Timestamp(period2_start)) &
                        (pd.to_datetime(all_ast['test_date']) <= pd.Timestamp(period2_end))
                    ]

                    if not period1_data.empty and not period2_data.empty:
                        # Comparison metrics
                        col1, col2, col3 = st.columns(3)

                        with col1:
                            p1_resistance = (period1_data['result'] == 'R').sum() / len(period1_data) * 100
                            st.metric(f"Period 1 Resistance ({period1_start.strftime('%Y-%m')})", f"{p1_resistance:.1f}%")

                        with col2:
                            p2_resistance = (period2_data['result'] == 'R').sum() / len(period2_data) * 100
                            st.metric(f"Period 2 Resistance ({period2_start.strftime('%Y-%m')})", f"{p2_resistance:.1f}%")

                        with col3:
                            diff = p2_resistance - p1_resistance
                            trend = "Increasing" if diff > 0 else "Decreasing" if diff < 0 else "Stable"
                            st.metric("Trend", f"{diff:+.1f}%", trend)

                        # Trend visualization
                        trend_data = pd.DataFrame({
                            'Period': [f"{period1_start.strftime('%Y-%m')}", f"{period2_start.strftime('%Y-%m')}"],
                            'Resistance_Rate': [p1_resistance, p2_resistance]
                        })

                        fig = px.line(trend_data, x='Period', y='Resistance_Rate',
                                    title='Resistance Trend Over Time',
                                    markers=True, color_discrete_sequence=['#FF6B6B'])
                        fig.update_layout(yaxis_title='Resistance Rate (%)')
                        st.plotly_chart(fig, use_container_width=True)

                        # Organism comparison
                        st.markdown("### Organism Resistance Changes")

                        p1_org = period1_data.groupby('organism').agg({
                            'result': lambda x: (x == 'R').sum() / len(x) * 100
                        }).round(1)

                        p2_org = period2_data.groupby('organism').agg({
                            'result': lambda x: (x == 'R').sum() / len(x) * 100
                        }).round(1)

                        # Find organisms present in both periods
                        common_orgs = set(p1_org.index) & set(p2_org.index)

                        if common_orgs:
                            comparison_org = []
                            for org in common_orgs:
                                comparison_org.append({
                                    'Organism': org,
                                    'Period_1': p1_org.loc[org, 'result'],
                                    'Period_2': p2_org.loc[org, 'result'],
                                    'Change': p2_org.loc[org, 'result'] - p1_org.loc[org, 'result']
                                })

                            org_comparison = pd.DataFrame(comparison_org).sort_values('Change', key=abs, ascending=False)

                            fig = px.bar(org_comparison.head(10), x='Organism', y='Change',
                                       title='Organism Resistance Changes (Period 2 - Period 1)',
                                       color='Change',
                                       color_continuous_scale=['green', 'red'])
                            fig.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.warning("One or both periods have no data. Please adjust the date ranges.")
            else:
                st.warning("Date information not available for time period comparison.")

        elif analysis_type == "Source Type Comparison":
            st.subheader("🏭 Source Type Comparison")

            source_types = sorted(all_samples['source_type'].dropna().unique())
            if len(source_types) > 1:
                selected_sources = st.multiselect(
                    "Select Source Types to Compare",
                    source_types,
                    default=source_types[:3] if len(source_types) >= 3 else source_types,
                    key="source_comparison"
                )

                if len(selected_sources) >= 2 and st.button("Compare Sources", key="compare_sources"):
                    # Similar logic to regional comparison but for source types
                    source_data = {}

                    for source in selected_sources:
                        source_samples = all_samples[all_samples['source_type'] == source]
                        source_ast = all_ast[all_ast['sample_id'].isin(source_samples['sample_id'])]

                        if not source_ast.empty:
                            resistance_rate = (source_ast['result'] == 'R').sum() / len(source_ast) * 100
                            source_data[source] = {
                                'resistance_rate': resistance_rate,
                                'total_tests': len(source_ast),
                                'resistant_count': (source_ast['result'] == 'R').sum(),
                                'data': source_ast
                            }

                    if len(source_data) >= 2:
                        # Create comparison table
                        comparison_table = []
                        for source, data in source_data.items():
                            comparison_table.append({
                                'Source Type': source,
                                'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                'Total Tests': data['total_tests'],
                                'Resistant Isolates': data['resistant_count']
                            })

                        st.dataframe(pd.DataFrame(comparison_table))

                        # Resistance rate comparison chart
                        fig = px.bar(
                            pd.DataFrame(comparison_table),
                            x='Source Type',
                            y='Resistance Rate (%)',
                            title='Source Type Resistance Comparison',
                            color='Resistance Rate (%)',
                            color_continuous_scale='Blues'
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.warning("Need data from at least 2 source types for comparison.")
                else:
                    st.info("Select at least 2 source types to compare.")
            else:
                st.warning("Need data from multiple source types for comparison.")

        elif analysis_type == "Multi-Parameter Comparison":
            st.subheader("Multi-Parameter Comparison")
            st.markdown("Compare resistance patterns across multiple values of a single parameter (e.g., multiple regions, organisms, or antibiotics)")

            # Parameter selection
            parameter_type = st.selectbox(
                "Select Parameter to Compare",
                ["Regions", "Organisms", "Antibiotics", "Categories", "Source Types"],
                key="multi_param_type"
            )

            st.markdown("---")

            if parameter_type == "Regions":
                regions = sorted(all_samples['region'].dropna().unique())
                if len(regions) >= 2:
                    selected_items = st.multiselect(
                        "Select Regions to Compare",
                        regions,
                        default=regions[:min(5, len(regions))],
                        key="multi_regions"
                    )

                    if len(selected_items) >= 2 and st.button("Compare Multiple Regions", key="multi_region_compare"):
                        comparison_data = {}
                        
                        for region in selected_items:
                            region_samples = all_samples[all_samples['region'] == region]
                            region_ast = all_ast[all_ast['sample_id'].isin(region_samples['sample_id'])]
                            
                            if not region_ast.empty:
                                resistance_rate = (region_ast['result'] == 'R').sum() / len(region_ast) * 100
                                comparison_data[region] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(region_ast),
                                    'resistant_count': (region_ast['result'] == 'R').sum(),
                                    'susceptible_count': (region_ast['result'] == 'S').sum(),
                                    'intermediate_count': (region_ast['result'] == 'I').sum(),
                                    'data': region_ast
                                }
                        
                        if comparison_data:
                            # Create comparison table
                            comp_table = pd.DataFrame([
                                {
                                    'Parameter': param,
                                    'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                    'Total Tests': data['total_tests'],
                                    'Resistant': data['resistant_count'],
                                    'Susceptible': data['susceptible_count'],
                                    'Intermediate': data['intermediate_count']
                                }
                                for param, data in comparison_data.items()
                            ]).sort_values('Resistance Rate (%)', ascending=False)
                            
                            st.dataframe(comp_table, use_container_width=True)
                            
                            # Bar chart comparison
                            fig = px.bar(
                                comp_table,
                                x='Parameter',
                                y='Resistance Rate (%)',
                                title='Resistance Rate Comparison Across Regions',
                                color='Resistance Rate (%)',
                                color_continuous_scale='Reds',
                                text='Resistance Rate (%)'
                            )
                            fig.update_traces(textposition='outside')
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Result distribution
                            st.markdown("### Result Distribution by Region")
                            
                            result_data = []
                            for region, data in comparison_data.items():
                                total = data['total_tests']
                                result_data.append({
                                    'Region': region,
                                    'Resistant': (data['resistant_count'] / total * 100) if total > 0 else 0,
                                    'Susceptible': (data['susceptible_count'] / total * 100) if total > 0 else 0,
                                    'Intermediate': (data['intermediate_count'] / total * 100) if total > 0 else 0
                                })
                            
                            result_df = pd.DataFrame(result_data)
                            fig = px.bar(result_df, x='Region', y=['Resistant', 'Susceptible', 'Intermediate'],
                                       title='Result Distribution (%)',
                                       barmode='stack',
                                       color_discrete_map={'Resistant': '#FF6B6B', 'Susceptible': '#51CF66', 'Intermediate': '#FFD93D'})
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Top organisms comparison
                            st.markdown("### Top Organisms by Region")
                            
                            org_cols = st.columns(min(3, len(comparison_data)))
                            for idx, (region, data) in enumerate(list(comparison_data.items())[:3]):
                                with org_cols[idx]:
                                    st.markdown(f"**{region}**")
                                    org_data = data['data'].groupby('organism').agg({
                                        'result': lambda x: (x == 'R').sum() / len(x) * 100
                                    }).round(1).sort_values(by='result', ascending=False).head(5)
                                    org_data.columns = ['Resistance %']
                                    st.dataframe(org_data, use_container_width=True)
                else:
                    st.warning("Need data from at least 2 regions for comparison.")

            elif parameter_type == "Organisms":
                organisms = sorted(all_ast['organism'].dropna().unique())
                if len(organisms) >= 2:
                    selected_items = st.multiselect(
                        "Select Organisms to Compare",
                        organisms,
                        default=organisms[:min(5, len(organisms))],
                        key="multi_organisms"
                    )

                    if len(selected_items) >= 2 and st.button("Compare Multiple Organisms", key="multi_org_compare"):
                        comparison_data = {}
                        
                        for organism in selected_items:
                            org_ast = all_ast[all_ast['organism'] == organism]
                            
                            if not org_ast.empty:
                                resistance_rate = (org_ast['result'] == 'R').sum() / len(org_ast) * 100
                                comparison_data[organism] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(org_ast),
                                    'resistant_count': (org_ast['result'] == 'R').sum(),
                                    'susceptible_count': (org_ast['result'] == 'S').sum(),
                                    'intermediate_count': (org_ast['result'] == 'I').sum(),
                                    'data': org_ast
                                }
                        
                        if comparison_data:
                            # Create comparison table
                            comp_table = pd.DataFrame([
                                {
                                    'Organism': param,
                                    'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                    'Total Tests': data['total_tests'],
                                    'Resistant': data['resistant_count'],
                                    'Susceptible': data['susceptible_count'],
                                    'Intermediate': data['intermediate_count']
                                }
                                for param, data in comparison_data.items()
                            ]).sort_values('Resistance Rate (%)', ascending=False)
                            
                            st.dataframe(comp_table, use_container_width=True)
                            
                            # Bar chart comparison
                            fig = px.bar(
                                comp_table,
                                x='Organism',
                                y='Resistance Rate (%)',
                                title='Resistance Rate Comparison Across Organisms',
                                color='Resistance Rate (%)',
                                color_continuous_scale='Purples',
                                text='Resistance Rate (%)'
                            )
                            fig.update_traces(textposition='outside')
                            fig.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Top antibiotics comparison
                            st.markdown("### Top Antibiotics by Organism")
                            
                            org_cols = st.columns(min(3, len(comparison_data)))
                            for idx, (organism, data) in enumerate(list(comparison_data.items())[:3]):
                                with org_cols[idx]:
                                    st.markdown(f"**{organism}**")
                                    try:
                                        antibiotic_data = data['data'].groupby('antibiotic').agg({
                                            'result': lambda x: (x == 'R').sum() / len(x) * 100
                                        }).round(1).sort_values(by='result', ascending=False).head(5)
                                        antibiotic_data.columns = ['Resistance %']
                                        st.dataframe(antibiotic_data, use_container_width=True)
                                    except Exception as e:
                                        st.warning(f"Error processing {organism}: {str(e)}")
                else:
                    st.warning("Need data from at least 2 organisms for comparison.")

            elif parameter_type == "Antibiotics":
                antibiotics = sorted(all_ast['antibiotic'].dropna().unique())
                if len(antibiotics) >= 2:
                    selected_items = st.multiselect(
                        "Select Antibiotics to Compare",
                        antibiotics,
                        default=antibiotics[:min(8, len(antibiotics))],
                        key="multi_antibiotics"
                    )

                    if len(selected_items) >= 2 and st.button("Compare Multiple Antibiotics", key="multi_antibiotic_compare"):
                        comparison_data = {}
                        
                        for antibiotic in selected_items:
                            antibiotic_ast = all_ast[all_ast['antibiotic'] == antibiotic]
                            
                            if not antibiotic_ast.empty:
                                resistance_rate = (antibiotic_ast['result'] == 'R').sum() / len(antibiotic_ast) * 100
                                comparison_data[antibiotic] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(antibiotic_ast),
                                    'resistant_count': (antibiotic_ast['result'] == 'R').sum(),
                                    'susceptible_count': (antibiotic_ast['result'] == 'S').sum(),
                                    'intermediate_count': (antibiotic_ast['result'] == 'I').sum(),
                                    'data': antibiotic_ast
                                }
                        
                        if comparison_data:
                            # Create comparison table
                            comp_table = pd.DataFrame([
                                {
                                    'Antibiotic': param,
                                    'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                    'Total Tests': data['total_tests'],
                                    'Resistant': data['resistant_count'],
                                    'Susceptible': data['susceptible_count'],
                                    'Intermediate': data['intermediate_count']
                                }
                                for param, data in comparison_data.items()
                            ]).sort_values('Resistance Rate (%)', ascending=False)
                            
                            st.dataframe(comp_table, use_container_width=True)
                            
                            # Bar chart comparison
                            fig = px.bar(
                                comp_table,
                                x='Antibiotic',
                                y='Resistance Rate (%)',
                                title='Resistance Rate Comparison Across Antibiotics',
                                color='Resistance Rate (%)',
                                color_continuous_scale='Oranges',
                                text='Resistance Rate (%)'
                            )
                            fig.update_traces(textposition='outside')
                            fig.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Top organisms by antibiotic
                            st.markdown("### Top Organisms by Antibiotic")
                            
                            org_cols = st.columns(min(4, len(comparison_data)))
                            for idx, (antibiotic, data) in enumerate(list(comparison_data.items())[:4]):
                                with org_cols[idx]:
                                    st.markdown(f"**{antibiotic}**")
                                    try:
                                        org_data = data['data'].groupby('organism').agg({
                                            'result': lambda x: (x == 'R').sum() / len(x) * 100
                                        }).round(1).sort_values(by='result', ascending=False).head(5)
                                        org_data.columns = ['Resistance %']
                                        st.dataframe(org_data, use_container_width=True)
                                    except Exception as e:
                                        st.warning(f"Error processing {antibiotic}: {str(e)}")
                else:
                    st.warning("Need data from at least 2 antibiotics for comparison.")

            elif parameter_type == "Categories":
                categories = sorted(all_samples['source_category'].dropna().unique())
                if len(categories) >= 2:
                    selected_items = st.multiselect(
                        "Select Categories to Compare",
                        categories,
                        default=categories[:min(5, len(categories))],
                        key="multi_categories"
                    )

                    if len(selected_items) >= 2 and st.button("Compare Multiple Categories", key="multi_cat_compare"):
                        comparison_data = {}
                        
                        for category in selected_items:
                            cat_samples = all_samples[all_samples['source_category'] == category]
                            cat_ast = all_ast[all_ast['sample_id'].isin(cat_samples['sample_id'])]
                            
                            if not cat_ast.empty:
                                resistance_rate = (cat_ast['result'] == 'R').sum() / len(cat_ast) * 100
                                comparison_data[category] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(cat_ast),
                                    'resistant_count': (cat_ast['result'] == 'R').sum(),
                                    'susceptible_count': (cat_ast['result'] == 'S').sum(),
                                    'intermediate_count': (cat_ast['result'] == 'I').sum(),
                                    'data': cat_ast
                                }
                        
                        if comparison_data:
                            # Create comparison table
                            comp_table = pd.DataFrame([
                                {
                                    'Category': param,
                                    'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                    'Total Tests': data['total_tests'],
                                    'Resistant': data['resistant_count'],
                                    'Susceptible': data['susceptible_count'],
                                    'Intermediate': data['intermediate_count']
                                }
                                for param, data in comparison_data.items()
                            ]).sort_values('Resistance Rate (%)', ascending=False)
                            
                            st.dataframe(comp_table, use_container_width=True)
                            
                            # Bar chart comparison
                            fig = px.bar(
                                comp_table,
                                x='Category',
                                y='Resistance Rate (%)',
                                title='Resistance Rate Comparison Across Categories',
                                color='Resistance Rate (%)',
                                color_continuous_scale='Greens',
                                text='Resistance Rate (%)'
                            )
                            fig.update_traces(textposition='outside')
                            st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("Need data from at least 2 categories for comparison.")

            elif parameter_type == "Source Types":
                source_types = sorted(all_samples['source_type'].dropna().unique())
                if len(source_types) >= 2:
                    selected_items = st.multiselect(
                        "Select Source Types to Compare",
                        source_types,
                        default=source_types[:min(5, len(source_types))],
                        key="multi_sources"
                    )

                    if len(selected_items) >= 2 and st.button("Compare Multiple Source Types", key="multi_source_compare"):
                        comparison_data = {}
                        
                        for source_type in selected_items:
                            source_samples = all_samples[all_samples['source_type'] == source_type]
                            source_ast = all_ast[all_ast['sample_id'].isin(source_samples['sample_id'])]
                            
                            if not source_ast.empty:
                                resistance_rate = (source_ast['result'] == 'R').sum() / len(source_ast) * 100
                                comparison_data[source_type] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(source_ast),
                                    'resistant_count': (source_ast['result'] == 'R').sum(),
                                    'susceptible_count': (source_ast['result'] == 'S').sum(),
                                    'intermediate_count': (source_ast['result'] == 'I').sum(),
                                    'data': source_ast
                                }
                        
                        if comparison_data:
                            # Create comparison table
                            comp_table = pd.DataFrame([
                                {
                                    'Source Type': param,
                                    'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                    'Total Tests': data['total_tests'],
                                    'Resistant': data['resistant_count'],
                                    'Susceptible': data['susceptible_count'],
                                    'Intermediate': data['intermediate_count']
                                }
                                for param, data in comparison_data.items()
                            ]).sort_values('Resistance Rate (%)', ascending=False)
                            
                            st.dataframe(comp_table, use_container_width=True)
                            
                            # Bar chart comparison
                            fig = px.bar(
                                comp_table,
                                x='Source Type',
                                y='Resistance Rate (%)',
                                title='Resistance Rate Comparison Across Source Types',
                                color='Resistance Rate (%)',
                                color_continuous_scale='Blues',
                                text='Resistance Rate (%)'
                            )
                            fig.update_traces(textposition='outside')
                            fig.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(fig, use_container_width=True)
                else:
                    st.warning("Need data from at least 2 source types for comparison.")

        elif analysis_type == "Cross-Variable Comparison":
            st.subheader("Cross-Variable Comparison")
            st.markdown("Compare a specific organism-antibiotic combination across different variables (regions, source types, categories, etc.)")

            st.markdown("---")

            # Step 1: Select organism and antibiotic
            col1, col2 = st.columns(2)

            organisms = sorted(all_ast['organism'].dropna().unique())
            antibiotics = sorted(all_ast['antibiotic'].dropna().unique())

            with col1:
                selected_organism = st.selectbox(
                    "Select Organism",
                    organisms,
                    key="cross_organism"
                )

            with col2:
                selected_antibiotic = st.selectbox(
                    "Select Antibiotic",
                    antibiotics,
                    key="cross_antibiotic"
                )

            st.markdown("---")

            # Step 2: Select variable to compare across
            comparison_variable = st.selectbox(
                "Compare This Combination Across:",
                ["Regions", "Districts", "Source Types", "Categories", "Sources", "Time Periods"],
                key="cross_variable"
            )

            st.markdown("---")

            if st.button("Compare Across Variable", key="cross_compare"):
                # Filter for the selected organism and antibiotic
                filtered_ast = all_ast[
                    (all_ast['organism'] == selected_organism) & 
                    (all_ast['antibiotic'] == selected_antibiotic)
                ]

                if filtered_ast.empty:
                    st.warning(f"No data found for {selected_organism} tested against {selected_antibiotic}")
                else:
                    # Merge with samples data to get location/source information
                    filtered_with_samples = filtered_ast.merge(
                        all_samples[['sample_id', 'region', 'district', 'source_type', 'source_category', 'collection_date']],
                        on='sample_id',
                        how='left'
                    )

                    comparison_data = {}

                    if comparison_variable == "Regions":
                        regions = sorted(filtered_with_samples['region'].dropna().unique())
                        
                        for region in regions:
                            region_data = filtered_with_samples[filtered_with_samples['region'] == region]
                            if not region_data.empty:
                                resistance_rate = (region_data['result'] == 'R').sum() / len(region_data) * 100
                                comparison_data[region] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(region_data),
                                    'resistant_count': (region_data['result'] == 'R').sum(),
                                    'susceptible_count': (region_data['result'] == 'S').sum(),
                                    'intermediate_count': (region_data['result'] == 'I').sum()
                                }

                    elif comparison_variable == "Districts":
                        districts = sorted(filtered_with_samples['district'].dropna().unique())
                        
                        for district in districts:
                            district_data = filtered_with_samples[filtered_with_samples['district'] == district]
                            if not district_data.empty:
                                resistance_rate = (district_data['result'] == 'R').sum() / len(district_data) * 100
                                comparison_data[district] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(district_data),
                                    'resistant_count': (district_data['result'] == 'R').sum(),
                                    'susceptible_count': (district_data['result'] == 'S').sum(),
                                    'intermediate_count': (district_data['result'] == 'I').sum()
                                }

                    elif comparison_variable == "Source Types":
                        source_types = sorted(filtered_with_samples['source_type'].dropna().unique())
                        
                        for source_type in source_types:
                            source_data = filtered_with_samples[filtered_with_samples['source_type'] == source_type]
                            if not source_data.empty:
                                resistance_rate = (source_data['result'] == 'R').sum() / len(source_data) * 100
                                comparison_data[source_type] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(source_data),
                                    'resistant_count': (source_data['result'] == 'R').sum(),
                                    'susceptible_count': (source_data['result'] == 'S').sum(),
                                    'intermediate_count': (source_data['result'] == 'I').sum()
                                }

                    elif comparison_variable == "Categories":
                        categories = sorted(filtered_with_samples['source_category'].dropna().unique())
                        
                        for category in categories:
                            cat_data = filtered_with_samples[filtered_with_samples['source_category'] == category]
                            if not cat_data.empty:
                                resistance_rate = (cat_data['result'] == 'R').sum() / len(cat_data) * 100
                                comparison_data[category] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(cat_data),
                                    'resistant_count': (cat_data['result'] == 'R').sum(),
                                    'susceptible_count': (cat_data['result'] == 'S').sum(),
                                    'intermediate_count': (cat_data['result'] == 'I').sum()
                                }

                    elif comparison_variable == "Sources":
                        sources = sorted(filtered_with_samples.get('source', filtered_with_samples.get('source_type', pd.Series(dtype='object'))).dropna().unique())
                        
                        for source in sources:
                            source_data = filtered_with_samples[filtered_with_samples.get('source', filtered_with_samples['source_type']) == source]
                            if not source_data.empty:
                                resistance_rate = (source_data['result'] == 'R').sum() / len(source_data) * 100
                                comparison_data[source] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(source_data),
                                    'resistant_count': (source_data['result'] == 'R').sum(),
                                    'susceptible_count': (source_data['result'] == 'S').sum(),
                                    'intermediate_count': (source_data['result'] == 'I').sum()
                                }

                    elif comparison_variable == "Time Periods":
                        filtered_with_samples['test_month'] = pd.to_datetime(filtered_with_samples.get('collection_date', filtered_with_samples.get('test_date', pd.Series(dtype='object')))).dt.to_period('M')
                        time_periods = sorted(filtered_with_samples['test_month'].dropna().unique())
                        
                        for period in time_periods:
                            period_data = filtered_with_samples[filtered_with_samples['test_month'] == period]
                            if not period_data.empty:
                                resistance_rate = (period_data['result'] == 'R').sum() / len(period_data) * 100
                                comparison_data[str(period)] = {
                                    'resistance_rate': resistance_rate,
                                    'total_tests': len(period_data),
                                    'resistant_count': (period_data['result'] == 'R').sum(),
                                    'susceptible_count': (period_data['result'] == 'S').sum(),
                                    'intermediate_count': (period_data['result'] == 'I').sum()
                                }

                    if comparison_data:
                        # Create summary header
                        st.markdown(f"### {selected_organism} vs {selected_antibiotic} - Across {comparison_variable}")
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Total Tests (All)", len(filtered_ast))
                        with col2:
                            overall_resistance = (filtered_ast['result'] == 'R').sum() / len(filtered_ast) * 100
                            st.metric("Overall Resistance Rate", f"{overall_resistance:.1f}%")
                        with col3:
                            st.metric("Locations/Variables", len(comparison_data))

                        st.markdown("---")

                        # Create comparison table
                        comp_table = pd.DataFrame([
                            {
                                comparison_variable.rstrip('s'): param,
                                'Resistance Rate (%)': round(data['resistance_rate'], 1),
                                'Total Tests': data['total_tests'],
                                'Resistant': data['resistant_count'],
                                'Susceptible': data['susceptible_count'],
                                'Intermediate': data['intermediate_count']
                            }
                            for param, data in comparison_data.items()
                        ]).sort_values('Resistance Rate (%)', ascending=False)

                        st.dataframe(comp_table, use_container_width=True)

                        st.markdown("---")

                        # Bar chart comparison
                        fig = px.bar(
                            comp_table,
                            x=comparison_variable.rstrip('s'),
                            y='Resistance Rate (%)',
                            title=f'{selected_organism} + {selected_antibiotic} Resistance Rate Across {comparison_variable}',
                            color='Resistance Rate (%)',
                            color_continuous_scale='RdYlGn_r',
                            text='Resistance Rate (%)',
                            height=500
                        )
                        fig.update_traces(textposition='outside')
                        fig.update_layout(xaxis_tickangle=-45)
                        st.plotly_chart(fig, use_container_width=True)

                        # Result distribution across variable
                        st.markdown(f"### Result Distribution by {comparison_variable}")

                        result_dist = []
                        for var, data in comparison_data.items():
                            total = data['total_tests']
                            result_dist.append({
                                comparison_variable.rstrip('s'): var,
                                'Resistant (%)': (data['resistant_count'] / total * 100) if total > 0 else 0,
                                'Susceptible (%)': (data['susceptible_count'] / total * 100) if total > 0 else 0,
                                'Intermediate (%)': (data['intermediate_count'] / total * 100) if total > 0 else 0
                            })

                        result_dist_df = pd.DataFrame(result_dist)
                        fig = px.bar(
                            result_dist_df,
                            x=comparison_variable.rstrip('s'),
                            y=['Resistant (%)', 'Susceptible (%)', 'Intermediate (%)'],
                            title=f'Result Distribution Across {comparison_variable}',
                            barmode='stack',
                            color_discrete_map={'Resistant (%)': '#FF6B6B', 'Susceptible (%)': '#51CF66', 'Intermediate (%)': '#FFD93D'},
                            height=500
                        )
                        fig.update_layout(xaxis_tickangle=-45)
                        st.plotly_chart(fig, use_container_width=True)

                        # Heatmap style visualization
                        st.markdown(f"### Heatmap: {selected_organism} + {selected_antibiotic} Resistance")
                        
                        heatmap_data = comp_table.set_index(comparison_variable.rstrip('s'))
                        heatmap_vals = heatmap_data[['Resistant', 'Susceptible', 'Intermediate']]
                        
                        fig = px.imshow(
                            heatmap_vals.T,
                            labels=dict(x=comparison_variable.rstrip('s'), y='Result', color='Count'),
                            title=f'Test Result Distribution Heatmap',
                            color_continuous_scale='Blues',
                            height=300
                        )
                        st.plotly_chart(fig, use_container_width=True)

                    else:
                        st.warning(f"No data available for {selected_organism} vs {selected_antibiotic} across {comparison_variable}")

        elif analysis_type == "Custom Comparison":
            st.subheader("🎯 Custom Comparison")

            st.markdown("Create custom comparisons by selecting specific filter combinations:")

            # Custom filter setup
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("**Group A Filters:**")
                group_a_categories = st.multiselect(
                    "Categories (Group A)",
                    sorted(all_samples['source_category'].dropna().unique()),
                    key="group_a_cat"
                )
                group_a_regions = st.multiselect(
                    "Regions (Group A)",
                    sorted(all_samples['region'].dropna().unique()),
                    key="group_a_reg"
                )

            with col2:
                st.markdown("**Group B Filters:**")
                group_b_categories = st.multiselect(
                    "Categories (Group B)",
                    sorted(all_samples['source_category'].dropna().unique()),
                    key="group_b_cat"
                )
                group_b_regions = st.multiselect(
                    "Regions (Group B)",
                    sorted(all_samples['region'].dropna().unique()),
                    key="group_b_reg"
                )

            group_a_name = st.text_input("Group A Name", value="Group A", key="group_a_name")
            group_b_name = st.text_input("Group B Name", value="Group B", key="group_b_name")

            if st.button("Run Custom Comparison", key="custom_comparison"):
                # Apply filters for Group A
                group_a_samples = all_samples
                if group_a_categories:
                    group_a_samples = group_a_samples[group_a_samples['source_category'].isin(group_a_categories)]
                if group_a_regions:
                    group_a_samples = group_a_samples[group_a_samples['region'].isin(group_a_regions)]

                # Apply filters for Group B
                group_b_samples = all_samples
                if group_b_categories:
                    group_b_samples = group_b_samples[group_b_samples['source_category'].isin(group_b_categories)]
                if group_b_regions:
                    group_b_samples = group_b_samples[group_b_samples['region'].isin(group_b_regions)]

                # Get AST data
                group_a_ast = all_ast[all_ast['sample_id'].isin(group_a_samples['sample_id'])]
                group_b_ast = all_ast[all_ast['sample_id'].isin(group_b_samples['sample_id'])]

                if not group_a_ast.empty and not group_b_ast.empty:
                    # Comparison metrics
                    a_resistance = (group_a_ast['result'] == 'R').sum() / len(group_a_ast) * 100
                    b_resistance = (group_b_ast['result'] == 'R').sum() / len(group_b_ast) * 100

                    col1, col2, col3 = st.columns(3)

                    with col1:
                        st.metric(f"{group_a_name} Resistance", f"{a_resistance:.1f}%")
                    with col2:
                        st.metric(f"{group_b_name} Resistance", f"{b_resistance:.1f}%")
                    with col3:
                        diff = b_resistance - a_resistance
                        st.metric("Difference", f"{diff:+.1f}%")

                    # Side-by-side charts
                    col1, col2 = st.columns(2)

                    with col1:
                        st.markdown(f"**{group_a_name}**")
                        try:
                            a_fig = plots.plot_resistance_distribution(group_a_ast)
                            st.plotly_chart(a_fig, use_container_width=True)
                        except Exception as e:
                            st.warning(f"Unable to generate chart for {group_a_name}: {str(e)}")

                    with col2:
                        st.markdown(f"**{group_b_name}**")
                        try:
                            b_fig = plots.plot_resistance_distribution(group_b_ast)
                            st.plotly_chart(b_fig, use_container_width=True)
                        except Exception as e:
                            st.warning(f"Unable to generate chart for {group_b_name}: {str(e)}")

                else:
                    st.warning("One or both groups have no data. Please adjust your filters.")

# ============================================================================
# PAGE 9: ALERTS DASHBOARD
# ============================================================================
elif page == "Alerts Dashboard":
    st.header("Alerts Dashboard")
    
    # Require dataset selection
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    # Import alerts module
    from src.alerts import (
        generate_all_alerts, alerts_to_dataframe, get_alert_summary,
        AlertSeverity, AlertType
    )
    
    all_samples, all_ast = _load_active_dataset()
    _render_dataset_banner(st.session_state.active_dataset_id)

    if all_ast.empty:
        _empty_state("No AST data available for alert generation.")
    else:
        # Alert Configuration
        with st.expander("Alert Configuration", expanded=False):
            col1, col2, col3 = st.columns(3)
            with col1:
                critical_threshold = st.slider("Critical Threshold (%)", 50, 95, 80, 5)
            with col2:
                high_threshold = st.slider("High Threshold (%)", 30, 80, 60, 5)
            with col3:
                medium_threshold = st.slider("Medium Threshold (%)", 20, 60, 40, 5)
        
        # Build thresholds dict
        custom_thresholds = {
            'critical': critical_threshold,
            'high': high_threshold,
            'medium': medium_threshold
        }
        
        # Generate alerts
        with st.spinner("Analyzing data for alerts..."):
            alerts = generate_all_alerts(
                all_ast, 
                all_samples,
                thresholds=custom_thresholds
            )
        
        # Alert Summary Cards
        summary = get_alert_summary(alerts)
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #ef4444, #dc2626); color: white; padding: 20px; border-radius: 10px; text-align: center;">
                <div style="font-size: 32px; font-weight: bold;">{summary['critical']}</div>
                <div>Critical Alerts</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #f97316, #ea580c); color: white; padding: 20px; border-radius: 10px; text-align: center;">
                <div style="font-size: 32px; font-weight: bold;">{summary['high']}</div>
                <div>High Priority</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #eab308, #ca8a04); color: white; padding: 20px; border-radius: 10px; text-align: center;">
                <div style="font-size: 32px; font-weight: bold;">{summary['medium']}</div>
                <div>Medium Priority</div>
            </div>
            """, unsafe_allow_html=True)
        with col4:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #22c55e, #16a34a); color: white; padding: 20px; border-radius: 10px; text-align: center;">
                <div style="font-size: 32px; font-weight: bold;">{summary['low']}</div>
                <div>Low Priority</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        if alerts:
            # Filter alerts by type
            st.subheader("Alert Details")
            
            alert_type_filter = st.multiselect(
                "Filter by Alert Type",
                options=[t.value.replace('_', ' ').title() for t in AlertType],
                default=[t.value.replace('_', ' ').title() for t in AlertType]
            )
            
            severity_filter = st.multiselect(
                "Filter by Severity",
                options=[s.value.upper() for s in AlertSeverity],
                default=[s.value.upper() for s in AlertSeverity]
            )
            
            # Convert to dataframe for display
            alerts_df = alerts_to_dataframe(alerts)
            
            # Apply filters using correct column names
            filtered_alerts = alerts_df[
                (alerts_df['Type'].isin(alert_type_filter)) &
                (alerts_df['Severity'].isin(severity_filter))
            ]
            
            if not filtered_alerts.empty:
                # Display alerts
                for _, alert in filtered_alerts.iterrows():
                    severity_color = {
                        'CRITICAL': '#ef4444',
                        'HIGH': '#f97316',
                        'MEDIUM': '#eab308',
                        'LOW': '#22c55e'
                    }.get(alert['Severity'], '#64748b')
                    
                    with st.expander(f"{alert['Title']}", expanded=alert['Severity'] == 'CRITICAL'):
                        st.markdown(f"""
                        <div style="border-left: 4px solid {severity_color}; padding-left: 15px;">
                            <p><strong>Severity:</strong> <span style="color: {severity_color}; font-weight: bold;">{alert['Severity']}</span></p>
                            <p><strong>Type:</strong> {alert['Type']}</p>
                            <p><strong>Description:</strong> {alert['Description']}</p>
                            <p><strong>Detected:</strong> {alert['Created']}</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if alert['Organism'] != '-':
                            st.write(f"**Organism:** {alert['Organism']}")
                        if alert['Antibiotic'] != '-':
                            st.write(f"**Antibiotic:** {alert['Antibiotic']}")
                        if alert['Current Value'] != '-':
                            st.write(f"**Current Value:** {alert['Current Value']}")
            else:
                st.info("No alerts match the selected filters.")
        else:
            st.success("No alerts detected based on current thresholds. Your data looks good!")

# ============================================================================
# PAGE 10: ANTIBIOGRAM
# ============================================================================
elif page == "Antibiogram":
    st.header("Cumulative Antibiogram")
    
    # Require dataset selection
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    # Import antibiogram module
    from src.antibiogram import (
        generate_antibiogram, antibiogram_to_html, antibiogram_to_excel,
        generate_quarterly_antibiograms, compare_antibiograms,
        CLSI_MIN_ISOLATES
    )
    
    all_samples, all_ast = _load_active_dataset()
    _render_dataset_banner(st.session_state.active_dataset_id)

    if all_ast.empty:
        _empty_state("No AST data available for antibiogram generation.")
    else:
        # Configuration
        st.subheader("Antibiogram Configuration")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            min_isolates = st.slider(
                "Minimum Isolates for Reporting",
                min_value=5, max_value=50, value=30, step=5,
                help=f"CLSI recommends minimum {CLSI_MIN_ISOLATES} isolates for cumulative antibiograms"
            )
        with col2:
            include_all = st.checkbox(
                "Include combinations below threshold",
                value=False,
                help="Show all combinations (marked with *) even if below minimum isolates"
            )
        with col3:
            lab_filter = st.selectbox(
                "Filter by Laboratory",
                options=["All Laboratories"] + sorted(all_samples['lab_name'].dropna().unique().tolist())
            )
        
        # Apply lab filter
        if lab_filter != "All Laboratories":
            filtered_samples = all_samples[all_samples['lab_name'] == lab_filter]
            filtered_ast = all_ast[all_ast['sample_id'].isin(filtered_samples['sample_id'])]
            lab_name = lab_filter
        else:
            filtered_ast = all_ast
            lab_name = "All Laboratories"
        
        # Generate antibiogram
        with st.spinner("Generating antibiogram..."):
            antibiogram = generate_antibiogram(
                filtered_ast,
                lab_name=lab_name,
                min_isolates=min_isolates,
                include_all=include_all
            )
        
        if 'error' in antibiogram and antibiogram.get('matrix') is None:
            st.error(antibiogram['error'])
        else:
            # Summary statistics
            st.subheader("Summary")
            summary = antibiogram.get('summary', {})
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Isolates", summary.get('total_isolates', 0))
            with col2:
                st.metric("Organisms", summary.get('total_organisms', 0))
            with col3:
                st.metric("Antibiotics", summary.get('total_antibiotics', 0))
            with col4:
                avg_susc = summary.get('overall_susceptibility', 0)
                st.metric("Avg Susceptibility", f"{avg_susc:.1f}%" if avg_susc else "N/A")
            
            st.markdown("---")
            
            # Display antibiogram
            st.subheader("Antibiogram Matrix")
            st.markdown("*Values show % Susceptible (number tested)*")
            
            # Display as interactive DataFrame with color styling
            matrix = antibiogram.get('matrix', pd.DataFrame())
            numeric_matrix = antibiogram.get('numeric_matrix', pd.DataFrame())
            
            if not matrix.empty:
                # Create styled dataframe
                def color_cells(val):
                    try:
                        # Extract numeric value from string like "85 (20)"
                        if pd.isna(val) or val == '-':
                            return 'background-color: #f0f0f0'
                        num_str = str(val).split('(')[0].strip().replace('*', '')
                        num = float(num_str) if num_str else 0
                        if num >= 90:
                            return 'background-color: #10b981; color: white'
                        elif num >= 70:
                            return 'background-color: #84cc16; color: white'
                        elif num >= 50:
                            return 'background-color: #fbbf24; color: #1f2937'
                        elif num >= 30:
                            return 'background-color: #f97316; color: white'
                        else:
                            return 'background-color: #ef4444; color: white'
                    except:
                        return 'background-color: #f0f0f0'
                
                # pandas >= 2.1 removed Styler.applymap in favour of
                # Styler.map; fall back to the old name for older envs.
                _styler = matrix.style
                _apply = getattr(_styler, "map", None) or _styler.applymap
                styled_df = _apply(color_cells)
                st.dataframe(styled_df, use_container_width=True, height=400)
                
                # Legend
                st.markdown("""
                <div style="margin-top: 15px; padding: 10px; background: #f8fafc; border-radius: 8px;">
                    <p style="font-weight: 600; margin-bottom: 5px;">Legend:</p>
                    <div style="display: flex; gap: 15px; flex-wrap: wrap; font-size: 12px;">
                        <span><span style="display: inline-block; width: 15px; height: 15px; background: #10b981; margin-right: 5px; vertical-align: middle;"></span>≥90% Susceptible</span>
                        <span><span style="display: inline-block; width: 15px; height: 15px; background: #84cc16; margin-right: 5px; vertical-align: middle;"></span>70-89%</span>
                        <span><span style="display: inline-block; width: 15px; height: 15px; background: #fbbf24; margin-right: 5px; vertical-align: middle;"></span>50-69%</span>
                        <span><span style="display: inline-block; width: 15px; height: 15px; background: #f97316; margin-right: 5px; vertical-align: middle;"></span>30-49%</span>
                        <span><span style="display: inline-block; width: 15px; height: 15px; background: #ef4444; margin-right: 5px; vertical-align: middle;"></span>&lt;30%</span>
                    </div>
                    <p style="margin-top: 10px; font-size: 11px; color: #64748b;">* indicates fewer than minimum isolates (interpret with caution)</p>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.warning("No antibiogram matrix data available.")
            
            # High resistance alerts
            if summary.get('lowest_susceptibility_combinations'):
                st.markdown("---")
                st.subheader("High Resistance Alerts")
                st.markdown("*Organism-antibiotic combinations with lowest susceptibility:*")
                
                for combo in summary['lowest_susceptibility_combinations']:
                    resistance = 100 - combo['pct_susceptible']
                    color = '#ef4444' if resistance >= 70 else '#f97316' if resistance >= 50 else '#eab308'
                    st.markdown(f"""
                    <div style="background: #f8fafc; padding: 10px; margin: 5px 0; border-radius: 5px; border-left: 4px solid {color};">
                        <strong>{combo['organism']}</strong> vs <strong>{combo['antibiotic']}</strong>: 
                        <span style="color: {color}; font-weight: bold;">{resistance:.0f}% resistant</span>
                        ({combo['total']} tested)
                    </div>
                    """, unsafe_allow_html=True)
            
            # Export options
            st.markdown("---")
            st.subheader("Export Antibiogram")
            
            col1, col2 = st.columns(2)
            with col1:
                # HTML export
                html_data = antibiogram_to_html(antibiogram)
                st.download_button(
                    label="Download as HTML",
                    data=html_data,
                    file_name=f"antibiogram_{lab_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.html",
                    mime="text/html"
                )
            with col2:
                # Excel export
                try:
                    excel_data = antibiogram_to_excel(antibiogram)
                    st.download_button(
                        label="Download as Excel",
                        data=excel_data,
                        file_name=f"antibiogram_{lab_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.warning(f"Excel export requires openpyxl: {e}")

# ============================================================================
# PAGE 11: WHONET EXPORT
# ============================================================================
elif page == "WHONET Export":
    st.header("WHONET Data Export")
    st.markdown("*Export data in WHONET format for integration with WHO GLASS and global surveillance networks.*")
    
    # Require dataset selection
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    # Import WHONET module
    from src.whonet import (
        convert_to_whonet_format, export_to_whonet_txt, export_to_whonet_excel,
        generate_glass_report, validate_whonet_data, generate_glass_html_report
    )
    
    all_samples, all_ast = _load_active_dataset()
    _render_dataset_banner(st.session_state.active_dataset_id)

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available for WHONET export.")
    else:
        # Lab configuration
        st.subheader("Laboratory Information")
        
        col1, col2 = st.columns(2)
        with col1:
            lab_code = st.text_input("Laboratory Code", value="GH001", help="WHONET laboratory identifier")
        with col2:
            lab_name = st.text_input("Laboratory Name", value="AMR Surveillance Lab Ghana")
        
        lab_info = {'code': lab_code, 'name': lab_name}
        
        # Convert to WHONET format
        with st.spinner("Converting data to WHONET format..."):
            whonet_df = convert_to_whonet_format(all_samples, all_ast, lab_info)
        
        if whonet_df.empty:
            st.error("Unable to convert data to WHONET format.")
        else:
            # Validate data
            validation = validate_whonet_data(whonet_df)
            
            # Summary
            st.subheader("Export Summary")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Records", len(whonet_df))
            with col2:
                st.metric("Unique Organisms", validation['statistics'].get('unique_organisms', 0))
            with col3:
                st.metric("Antibiotics", len(validation['statistics'].get('antibiotics', [])))
            with col4:
                status_icon = "Valid" if validation['is_valid'] else "Issues"
                st.metric("Status", status_icon)
            
            # Validation results
            if not validation['is_valid'] or validation['warnings']:
                with st.expander("Validation Details", expanded=not validation['is_valid']):
                    if validation['errors']:
                        for error in validation['errors']:
                            st.error(error)
                    if validation['warnings']:
                        for warning in validation['warnings']:
                            st.warning(warning)
            
            # Preview
            st.subheader("Data Preview")
            st.dataframe(whonet_df.head(20), use_container_width=True)
            
            st.markdown("---")
            
            # GLASS Report
            st.subheader("WHO GLASS Summary")
            
            glass_report = generate_glass_report(whonet_df)
            
            if 'error' not in glass_report:
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Organism Distribution:**")
                    for org, count in list(glass_report.get('organisms', {}).items())[:10]:
                        st.write(f"- {org}: {count}")
                
                with col2:
                    st.markdown("**Specimen Types:**")
                    for spec, count in glass_report.get('specimen_distribution', {}).items():
                        st.write(f"- {spec}: {count}")
                
                # Priority pathogen resistance rates
                if glass_report.get('resistance_rates'):
                    st.markdown("---")
                    st.markdown("**Priority Pathogen Resistance Rates:**")
                    
                    for org, data in glass_report['resistance_rates'].items():
                        with st.expander(f"{org} (n={data['isolate_count']})"):
                            for ab, ab_data in data.get('antibiotics', {}).items():
                                rate = ab_data.get('resistance_rate', 0)
                                color = 'red' if rate >= 50 else 'orange' if rate >= 30 else 'green'
                                st.markdown(f"- **{ab}**: {rate:.1f}% resistant ({ab_data['tested']} tested)")
            
            st.markdown("---")
            
            # Export options
            st.subheader("Export Options")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Tab-delimited text (standard WHONET format)
                txt_data = export_to_whonet_txt(whonet_df)
                st.download_button(
                    label="Download WHONET Text",
                    data=txt_data,
                    file_name=f"WHONET_export_{datetime.now().strftime('%Y%m%d')}.txt",
                    mime="text/plain",
                    help="Standard WHONET tab-delimited format"
                )
            
            with col2:
                # Excel format
                try:
                    excel_data = export_to_whonet_excel(whonet_df)
                    st.download_button(
                        label="Download WHONET Excel",
                        data=excel_data,
                        file_name=f"WHONET_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.warning(f"Excel export requires openpyxl: {e}")
            
            with col3:
                # GLASS summary HTML report
                glass_html = generate_glass_html_report(glass_report)
                st.download_button(
                    label="Download GLASS Report",
                    data=glass_html,
                    file_name=f"GLASS_Report_{datetime.now().strftime('%Y%m%d')}.html",
                    mime="text/html",
                    help="WHO GLASS formatted HTML report"
                )

# ============================================================================
# PAGE 12: REPORT EXPORT
# ============================================================================
elif page == "Report Export":
    st.header("Report Export")

    # Require dataset selection before showing dashboard
    if not st.session_state.active_dataset_id:
        st.warning("Please select a dataset in the 'Data Management' page first.")
        st.stop()
    
    all_samples, all_ast = _load_active_dataset()

    _render_dataset_banner(st.session_state.active_dataset_id)

    if all_ast.empty or all_samples.empty:
        _empty_state("No data available in the selected dataset.")
    else:
        # ============================================================================
        # FILTERING CONTROLS (Same as Resistance Overview)
        # ============================================================================
        st.subheader("Report Filters")
        st.markdown("Configure filters to generate reports based on specific data subsets:")

        col1, col2, col3 = st.columns(3)

        with col1:
            # Category filter
            categories = sorted(all_samples['source_category'].dropna().astype(str).unique().tolist())
            if categories:
                category_options = ["All"] + categories
                selected_category_options = st.multiselect(
                    "Source Category",
                    category_options,
                    default=["All"],
                    key="report_categories"
                )
                # If "All" is selected, use all categories; otherwise use selected ones
                if "All" in selected_category_options:
                    selected_categories = categories
                else:
                    selected_categories = [opt for opt in selected_category_options if opt != "All"]
            else:
                selected_categories = []

        with col2:
            # Source type filter
            source_types = sorted(all_samples['source_type'].dropna().astype(str).unique().tolist())
            if source_types:
                source_type_options = ["All"] + source_types
                selected_source_type_options = st.multiselect(
                    "Source Type",
                    source_type_options,
                    default=["All"],
                    key="report_source_types"
                )
                # If "All" is selected, use all source types; otherwise use selected ones
                if "All" in selected_source_type_options:
                    selected_source_types = source_types
                else:
                    selected_source_types = [opt for opt in selected_source_type_options if opt != "All"]
            else:
                selected_source_types = []

        with col3:
            # Site type filter
            site_types = sorted(all_samples['site_type'].dropna().astype(str).unique().tolist())
            if site_types:
                site_type_options = ["All"] + site_types
                selected_site_type_options = st.multiselect(
                    "Site Type",
                    site_type_options,
                    default=["All"],
                    key="report_site_types"
                )
                # If "All" is selected, use all site types; otherwise use selected ones
                if "All" in selected_site_type_options:
                    selected_site_types = site_types
                else:
                    selected_site_types = [opt for opt in selected_site_type_options if opt != "All"]
            else:
                selected_site_types = []

        col4, col5, col6 = st.columns(3)

        with col4:
            # Region filter
            regions = sorted(all_samples['region'].dropna().astype(str).unique().tolist())
            if regions:
                region_options = ["All"] + regions
                selected_region_options = st.multiselect(
                    "Region",
                    region_options,
                    default=["All"],
                    key="report_regions"
                )
                # If "All" is selected, use all regions; otherwise use selected ones
                if "All" in selected_region_options:
                    selected_regions = regions
                else:
                    selected_regions = [opt for opt in selected_region_options if opt != "All"]
            else:
                selected_regions = []

        with col5:
            # District filter
            districts = sorted(all_samples['district'].dropna().astype(str).unique().tolist())
            if districts:
                district_options = ["All"] + districts
                selected_district_options = st.multiselect(
                    "District",
                    district_options,
                    default=["All"],
                    key="report_districts"
                )
                # If "All" is selected, use all districts; otherwise use selected ones
                if "All" in selected_district_options:
                    selected_districts = districts
                else:
                    selected_districts = [opt for opt in selected_district_options if opt != "All"]
            else:
                selected_districts = []

        with col6:
            # Date range filter
            if 'test_date' in all_ast.columns:
                min_date = pd.to_datetime(all_ast['test_date'].dropna()).min()
                max_date = pd.to_datetime(all_ast['test_date'].dropna()).max()

                if pd.notna(min_date) and pd.notna(max_date):
                    date_range = st.date_input(
                        "Date Range",
                        value=(min_date.date(), max_date.date()),
                        key="report_date_range"
                    )
                    if len(date_range) == 2:
                        start_date, end_date = date_range
                    else:
                        start_date, end_date = min_date.date(), max_date.date()
                else:
                    start_date, end_date = None, None
            else:
                start_date, end_date = None, None

        col7, col8 = st.columns(2)

        with col7:
            # Organism filter
            organisms = sorted(all_ast['organism'].dropna().astype(str).unique().tolist())
            if organisms:
                organism_options = ["All"] + organisms
                selected_organism_options = st.multiselect(
                    "Organisms",
                    organism_options,
                    default=["All"],
                    key="report_organisms"
                )
                # If "All" is selected, use all organisms; otherwise use selected ones
                if "All" in selected_organism_options:
                    selected_organisms = organisms
                else:
                    selected_organisms = [opt for opt in selected_organism_options if opt != "All"]
            else:
                selected_organisms = []

        with col8:
            # Antibiotic filter
            antibiotics = sorted(all_ast['antibiotic'].dropna().astype(str).unique().tolist())
            if antibiotics:
                antibiotic_options = ["All"] + antibiotics
                selected_antibiotic_options = st.multiselect(
                    "Antibiotics",
                    antibiotic_options,
                    default=["All"],
                    key="report_antibiotics"
                )
                # If "All" is selected, use all antibiotics; otherwise use selected ones
                if "All" in selected_antibiotic_options:
                    selected_antibiotics = antibiotics
                else:
                    selected_antibiotics = [opt for opt in selected_antibiotic_options if opt != "All"]
            else:
                selected_antibiotics = []

        # Apply filters to get filtered data
        st.markdown("---")

        # Apply sample filters
        if selected_categories and selected_regions and selected_districts:
            filtered_samples = all_samples[
                (all_samples['source_category'].astype(str).isin(selected_categories)) &
                (all_samples['source_type'].astype(str).isin(selected_source_types)) &
                (all_samples['site_type'].astype(str).isin(selected_site_types)) &
                (all_samples['region'].astype(str).isin(selected_regions)) &
                (all_samples['district'].astype(str).isin(selected_districts))
            ]
        else:
            filtered_samples = all_samples

        # Apply AST filters
        base_ast_filter = all_ast['sample_id'].astype(str).isin(filtered_samples['sample_id'].astype(str))

        if selected_organisms:
            base_ast_filter &= all_ast['organism'].astype(str).isin(selected_organisms)

        if selected_antibiotics:
            base_ast_filter &= all_ast['antibiotic'].astype(str).isin(selected_antibiotics)

        # Apply date filter if available
        if start_date and end_date and 'test_date' in all_ast.columns:
            date_filter = pd.to_datetime(all_ast['test_date']).dt.date.between(start_date, end_date)
            base_ast_filter &= date_filter

        filtered_ast = all_ast[base_ast_filter]

        # Display filter summary
        st.subheader("Filtered Data Summary")
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Filtered Samples", filtered_samples['sample_id'].nunique())
        with col2:
            st.metric("Filtered Tests", len(filtered_ast))
        with col3:
            resistant_count = (filtered_ast['result'] == 'R').sum()
            resistance_rate = resistant_count / len(filtered_ast) * 100 if len(filtered_ast) > 0 else 0
            st.metric("Resistance Rate", f"{resistance_rate:.1f}%")
        with col4:
            st.metric("Organisms", filtered_ast['organism'].nunique())

        st.markdown("---")

        # ============================================================================
        # REPORT GENERATION
        # ============================================================================
        st.subheader("Generate Technical Report")

        if filtered_ast.empty:
            st.warning("No data matches the selected filters. Please adjust your filters.")
        else:
            # Report configuration
            report_title = st.text_input(
                "Report Title",
                value=f"AMR Technical Surveillance Report - {datetime.now().strftime('%B %Y')}",
                key="report_title"
            )

            # Dataset selection (optional - for metadata)
            datasets = _cached_all_datasets()
            # Hide admin-owned datasets from non-admin users
            config_admin_email, _ = _get_admin_config()
            admin_email = (config_admin_email or "").strip().lower()
            if not st.session_state.is_admin and admin_email:
                datasets = [d for d in datasets if (d.get('uploaded_by') or '').strip().lower() != admin_email]
            dataset_names = [f"{d['dataset_name']} ({d['dataset_id']})" for d in datasets]

            selected_dataset_name = "Filtered Dataset"
            if dataset_names:
                selected_dataset_display = st.selectbox(
                    "Reference Dataset (optional)",
                    ["None"] + dataset_names,
                    key="reference_dataset"
                )
                if selected_dataset_display != "None":
                    selected_dataset_name = selected_dataset_display.split('(')[0].strip()

            if st.button("Generate Technical Report", type="primary", use_container_width=True):
                with st.spinner("Generating comprehensive technical report with filtered data..."):
                    try:
                        # Generate HTML report with filtered data
                        html_content = report.generate_filtered_html_report(
                            report_title,
                            filtered_samples,
                            filtered_ast,
                            selected_categories,
                            selected_regions,
                            selected_organisms,
                            selected_antibiotics,
                            pps_df=db.get_pps_surveys(),
                            pps_rx_df=db.get_pps_prescriptions(),
                            amu_df=db.get_amu_records(),
                            amc_df=db.get_amc_records(),
                        )

                        # Success message
                        st.success("Professional HTML report generated successfully!")
                        st.info("Report includes embedded interactive visualizations and comprehensive filtered data analysis")

                        # Preview section
                        with st.expander("Report Preview", expanded=False):
                            st.markdown("**Report will include:**")
                            st.markdown("- Executive summary with key metrics")
                            st.markdown("- Interactive resistance distribution charts")
                            st.markdown("- Resistance Heat Map — Critical Combinations")
                            st.markdown("- Pathogen Profile — Top 5 Organisms")
                            st.markdown("- HAI Profile with ESKAPE Surveillance")
                            st.markdown("- Geographic and temporal analysis")
                            st.markdown("- Advanced analytics and risk assessment")
                            st.markdown("- One Health: PPS, AMU & AMC summaries")
                            st.markdown("- Professional formatting with no text overlap")

                        # Download buttons
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        filename = f"AMR_Report_Filtered_{timestamp}.html"
                        col_html, col_pdf = st.columns(2)
                        with col_html:
                            st.download_button(
                                label="Download HTML Report",
                                data=html_content,
                                file_name=filename,
                                mime="text/html",
                                use_container_width=True
                            )
                        with col_pdf:
                            try:
                                from src.report import html_to_pdf
                                pdf_bytes = html_to_pdf(html_content)
                                st.download_button(
                                    label="Download PDF Report",
                                    data=pdf_bytes,
                                    file_name=filename.replace('.html', '.pdf'),
                                    mime="application/pdf",
                                    use_container_width=True
                                )
                            except Exception:
                                logger.exception("PDF report generation failed; falling back to HTML")
                                st.info("💡 To save as PDF: download the HTML report, open it in your browser, then press **Ctrl+P → Save as PDF**.")

                    except Exception as e:
                        st.error(f"Error generating report: {str(e)}")
                        st.info("Please check your data and try again. If the error persists, contact support.")

# ============================================================================
# PAGE 14: PPS DASHBOARD
# ============================================================================
elif page == "PPS Dashboard":
    render_pps_page()

# ============================================================================
# PAGE 16: AMU DASHBOARD
# ============================================================================
elif page == "AMU Dashboard":
    render_amu_page()

# ============================================================================
# PAGE 17: AMC DASHBOARD
# ============================================================================
elif page == "AMC Dashboard":
    render_amc_page()

# ============================================================================
# PAGE 10: ADMIN - USER MANAGEMENT
# ============================================================================
elif page == "Admin - Users":
    if not st.session_state.is_admin:
        st.error("🚫 Access denied. Admin privileges required.")
        st.stop()
    
    st.header("👥 User Management")
    st.markdown("Manage user accounts and permissions")
    st.markdown("---")
    
    # Get all users
    all_users = db.get_all_users()
    
    if not all_users:
        st.info("📭 No users registered yet.")
    else:
        # Display users in a table
        st.subheader("Registered Users")
        
        # Create columns for display
        users_df = pd.DataFrame(all_users)
        users_df['created_at'] = pd.to_datetime(users_df['created_at']).dt.strftime('%Y-%m-%d %H:%M')
        users_df['last_login'] = users_df['last_login'].apply(
            lambda x: pd.to_datetime(x).strftime('%Y-%m-%d %H:%M') if x else "Never"
        )
        users_df['Status'] = users_df['is_active'].apply(lambda x: "Active" if x else "Inactive")
        users_df['Role'] = users_df['is_admin'].apply(lambda x: "Admin" if x else "User")
        
        # Display table
        display_df = users_df[['email', 'created_at', 'last_login', 'Status', 'Role']].copy()
        display_df.columns = ['Email', 'Created', 'Last Login', 'Status', 'Role']
        
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        st.markdown("---")
        
        # User management actions
        st.subheader("User Actions")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Deactivate User")
            selected_user = st.selectbox(
                "Select user to deactivate",
                [u for u in all_users if u['is_active']],
                format_func=lambda x: x['email'],
                key="deactivate_user"
            )
            if st.button("Deactivate", use_container_width=True, key="btn_deactivate"):
                success, msg = db.update_user_status(selected_user['user_id'], False)
                if success:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
        
        with col2:
            st.subheader("Reactivate User")
            selected_inactive = st.selectbox(
                "Select user to reactivate",
                [u for u in all_users if not u['is_active']],
                format_func=lambda x: x['email'],
                key="reactivate_user"
            )
            if st.button("Reactivate", use_container_width=True, key="btn_reactivate"):
                success, msg = db.update_user_status(selected_inactive['user_id'], True)
                if success:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
        
        st.markdown("---")
        
        # Reset password section
        st.subheader("Reset Password")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            user_for_reset = st.selectbox(
                "Select user to reset password",
                all_users,
                format_func=lambda x: x['email'],
                key="reset_user"
            )
        
        with col2:
            st.write("")  # Spacing
            if st.button("Generate Temporary Password", use_container_width=True):
                # Generate a temporary password
                temp_password = f"Temp@{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}"
                password_hash = bcrypt.hashpw(temp_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                success, msg = db.update_user_password(user_for_reset['email'], password_hash)
                
                if success:
                    st.success(msg)
                    st.info(f"Temporary Password: `{temp_password}`")
                    st.warning("Please share this password securely with the user. They should change it on first login.")
                else:
                    st.error(msg)
        
        st.markdown("---")
        
        # User statistics
        st.subheader("User Statistics")
        
        col1, col2, col3, col4 = st.columns(4)
        
        total_users = len(all_users)
        active_users = len([u for u in all_users if u['is_active']])
        inactive_users = len([u for u in all_users if not u['is_active']])
        admin_users = len([u for u in all_users if u['is_admin']])
        
        with col1:
            st.metric("Total Users", total_users)
        with col2:
            st.metric("Active Users", active_users)
        with col3:
            st.metric("Inactive Users", inactive_users)
        with col4:
            st.metric("Admins", admin_users)

# ============================================================================
# Footer
# ============================================================================
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #7f8c8d; font-size: 12px; margin-top: 30px;">
    <p>ICBB-AMRSS | ICBB AMR Surveillance System | Ghana</p>
    <p>Data stored locally in SQLite. No internet required.</p>
    <p><em>For academic and policy use. Always consult AMR experts for decision-making.</em></p>
</div>
""", unsafe_allow_html=True)

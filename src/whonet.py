"""
WHONET Export Module for AMR Surveillance Dashboard.
Exports data in WHONET format for integration with WHO GLASS and global surveillance networks.
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from io import BytesIO, StringIO
import re


# WHONET field mappings
WHONET_SPECIMEN_CODES = {
    'blood': 'bld',
    'urine': 'uri',
    'stool': 'sto',
    'wound': 'wou',
    'sputum': 'spu',
    'csf': 'csf',
    'cerebrospinal fluid': 'csf',
    'respiratory': 'res',
    'throat': 'thr',
    'eye': 'eye',
    'ear': 'ear',
    'genital': 'gen',
    'pus': 'pus',
    'tissue': 'tis',
    'catheter': 'cat',
    'other': 'oth',
    'unknown': 'unk'
}

WHONET_PATIENT_TYPE = {
    'inpatient': 'in',
    'outpatient': 'out',
    'emergency': 'er',
    'icu': 'icu',
    'unknown': 'unk'
}

# Common organism name to WHONET code mapping
WHONET_ORGANISM_CODES = {
    'escherichia coli': 'eco',
    'e. coli': 'eco',
    'klebsiella pneumoniae': 'kpn',
    'k. pneumoniae': 'kpn',
    'staphylococcus aureus': 'sau',
    's. aureus': 'sau',
    'pseudomonas aeruginosa': 'pae',
    'p. aeruginosa': 'pae',
    'acinetobacter baumannii': 'aba',
    'a. baumannii': 'aba',
    'enterococcus faecalis': 'efa',
    'e. faecalis': 'efa',
    'enterococcus faecium': 'efm',
    'e. faecium': 'efm',
    'streptococcus pneumoniae': 'spn',
    's. pneumoniae': 'spn',
    'salmonella': 'sal',
    'salmonella typhi': 'sty',
    's. typhi': 'sty',
    'salmonella typhimurium': 'stm',
    'salmonella enterica': 'sen',
    'shigella': 'shi',
    'shigella flexneri': 'sfl',
    'shigella sonnei': 'sso',
    'vibrio cholerae': 'vch',
    'v. cholerae': 'vch',
    'campylobacter': 'cam',
    'campylobacter jejuni': 'cje',
    'campylobacter coli': 'cco',
    'proteus mirabilis': 'pmi',
    'enterobacter cloacae': 'ecl',
    'serratia marcescens': 'sma',
    'citrobacter freundii': 'cfr',
    'morganella morganii': 'mmo',
    'providencia stuartii': 'pst',
    'haemophilus influenzae': 'hin',
    'neisseria meningitidis': 'nme',
    'neisseria gonorrhoeae': 'ngo',
    'listeria monocytogenes': 'lmo',
    'clostridium difficile': 'cdi'
}

# Common antibiotic name to WHONET code mapping
WHONET_ANTIBIOTIC_CODES = {
    'ampicillin': 'AMP',
    'amoxicillin': 'AMX',
    'amoxicillin-clavulanic acid': 'AMC',
    'amoxicillin/clavulanic acid': 'AMC',
    'amox-clav': 'AMC',
    'piperacillin': 'PIP',
    'piperacillin-tazobactam': 'TZP',
    'pip-tazo': 'TZP',
    'cefazolin': 'CZO',
    'cefuroxime': 'CXM',
    'cefotaxime': 'CTX',
    'ceftriaxone': 'CRO',
    'ceftazidime': 'CAZ',
    'cefepime': 'FEP',
    'cefoxitin': 'FOX',
    'aztreonam': 'ATM',
    'imipenem': 'IPM',
    'meropenem': 'MEM',
    'ertapenem': 'ETP',
    'doripenem': 'DOR',
    'gentamicin': 'GEN',
    'amikacin': 'AMK',
    'tobramycin': 'TOB',
    'streptomycin': 'STR',
    'ciprofloxacin': 'CIP',
    'levofloxacin': 'LVX',
    'moxifloxacin': 'MXF',
    'norfloxacin': 'NOR',
    'ofloxacin': 'OFX',
    'nalidixic acid': 'NAL',
    'tetracycline': 'TCY',
    'doxycycline': 'DOX',
    'minocycline': 'MNO',
    'tigecycline': 'TGC',
    'trimethoprim-sulfamethoxazole': 'SXT',
    'cotrimoxazole': 'SXT',
    'tmp-smx': 'SXT',
    'chloramphenicol': 'CHL',
    'erythromycin': 'ERY',
    'azithromycin': 'AZM',
    'clarithromycin': 'CLR',
    'clindamycin': 'CLI',
    'vancomycin': 'VAN',
    'teicoplanin': 'TEC',
    'linezolid': 'LNZ',
    'daptomycin': 'DAP',
    'nitrofurantoin': 'NIT',
    'fosfomycin': 'FOS',
    'colistin': 'COL',
    'polymyxin b': 'PLB',
    'rifampicin': 'RIF',
    'rifampin': 'RIF',
    'penicillin': 'PEN',
    'oxacillin': 'OXA',
    'methicillin': 'MET',
    'cefoxitin screen': 'FOX'
}


def normalize_specimen_type(specimen: str) -> str:
    """Normalize specimen type to WHONET code."""
    if pd.isna(specimen):
        return 'unk'
    
    specimen_lower = specimen.lower().strip()
    for key, code in WHONET_SPECIMEN_CODES.items():
        if key in specimen_lower:
            return code
    return 'oth'


def normalize_organism(organism: str) -> Tuple[str, str]:
    """Normalize organism name to WHONET code and standardized name."""
    if pd.isna(organism):
        return ('unk', 'Unknown')
    
    organism_lower = organism.lower().strip()
    for key, code in WHONET_ORGANISM_CODES.items():
        if key in organism_lower or organism_lower in key:
            # Return proper case version
            proper_name = key.title().replace("'S", "'s")
            return (code, proper_name)
    
    # Generate code from first 3 letters if not found
    parts = organism.split()
    if len(parts) >= 2:
        code = parts[0][:1].lower() + parts[1][:2].lower()
    else:
        code = organism[:3].lower()
    
    return (code, organism)


def normalize_antibiotic(antibiotic: str) -> str:
    """Normalize antibiotic name to WHONET code."""
    if pd.isna(antibiotic):
        return ''
    
    antibiotic_lower = antibiotic.lower().strip()
    for key, code in WHONET_ANTIBIOTIC_CODES.items():
        if key in antibiotic_lower or antibiotic_lower in key:
            return code
    
    # Return first 3 letters uppercase if not found
    return antibiotic[:3].upper()


def convert_to_whonet_format(samples_df: pd.DataFrame, 
                              ast_df: pd.DataFrame,
                              lab_info: Optional[Dict] = None) -> pd.DataFrame:
    """
    Convert AMR data to WHONET format.
    
    WHONET format requires specific columns and data formatting.
    Reference: https://whonet.org/
    
    Args:
        samples_df: Sample data including demographics and specimen info
        ast_df: AST results with organism and antibiotic susceptibility
        lab_info: Optional laboratory information dict
    
    Returns:
        DataFrame in WHONET format
    """
    if samples_df.empty or ast_df.empty:
        return pd.DataFrame()
    
    # Merge data
    merged = ast_df.merge(samples_df, on='sample_id', how='left')
    
    if merged.empty:
        return pd.DataFrame()
    
    # Create WHONET formatted records
    whonet_records = []
    
    for _, row in merged.iterrows():
        org_code, org_name = normalize_organism(row.get('organism', ''))
        
        record = {
            # Laboratory info
            'LABORATORY': lab_info.get('code', 'GH001') if lab_info else 'GH001',
            
            # Patient info
            'PATIENT_ID': row.get('sample_id', ''),
            'LAST_NAME': '',  # Anonymized
            'FIRST_NAME': '',  # Anonymized
            'SEX': row.get('sex', row.get('gender', 'U'))[:1].upper() if pd.notna(row.get('sex', row.get('gender'))) else 'U',
            'DATE_BIRTH': '',  # Anonymized for privacy
            'AGE': row.get('age', '') if pd.notna(row.get('age')) else '',
            
            # Specimen info
            'SPEC_DATE': pd.to_datetime(row.get('collection_date', '')).strftime('%Y-%m-%d') if pd.notna(row.get('collection_date')) else '',
            'SPEC_NUM': row.get('sample_id', ''),
            'SPEC_TYPE': normalize_specimen_type(row.get('specimen_type', row.get('sample_type', ''))),
            
            # Clinical info
            'DEPARTMENT': row.get('department', row.get('ward', '')),
            'WARD': row.get('ward', ''),
            'PATIENT_TYPE': WHONET_PATIENT_TYPE.get(str(row.get('patient_type', 'unknown')).lower(), 'unk'),
            
            # Organism info
            'ORGANISM': org_name,
            'ORG_CODE': org_code,
            'ISOLATE_NUM': row.get('isolate_id', 1),
            
            # Country info
            'COUNTRY': 'GHA',  # Ghana ISO code
            'REGION': row.get('region', ''),
            
            # Data origin
            'DATA_SOURCE': 'AMR_Dashboard',
            'YEAR': pd.to_datetime(row.get('collection_date', datetime.now())).year if pd.notna(row.get('collection_date')) else datetime.now().year
        }
        
        # Add antibiotic result
        antibiotic = row.get('antibiotic', '')
        if pd.notna(antibiotic):
            ab_code = normalize_antibiotic(antibiotic)
            result = row.get('result', '')
            mic = row.get('mic', '')
            disk = row.get('disk_zone', row.get('zone_diameter', ''))
            
            record[ab_code] = result
            if pd.notna(mic) and str(mic).strip():
                record[f'{ab_code}_MIC'] = mic
            if pd.notna(disk) and str(disk).strip():
                record[f'{ab_code}_DISK'] = disk
        
        whonet_records.append(record)
    
    whonet_df = pd.DataFrame(whonet_records)
    
    # Consolidate records by sample/isolate (one row per isolate with all antibiotics)
    # Group by patient/specimen/organism and combine antibiotic columns
    key_cols = ['LABORATORY', 'PATIENT_ID', 'SPEC_NUM', 'ORG_CODE', 'ISOLATE_NUM']
    
    # Get antibiotic columns (3-letter uppercase codes)
    ab_cols = [col for col in whonet_df.columns if re.match(r'^[A-Z]{3}(_MIC|_DISK)?$', col)]
    
    consolidated = whonet_df.groupby(key_cols, as_index=False).first()
    
    # Aggregate antibiotic results
    for ab_col in ab_cols:
        if ab_col in whonet_df.columns:
            consolidated[ab_col] = whonet_df.groupby(key_cols)[ab_col].first().reset_index(drop=True)
    
    return consolidated


def export_to_whonet_txt(whonet_df: pd.DataFrame) -> str:
    """
    Export WHONET DataFrame to tab-delimited text format.
    Standard WHONET interchange format.
    """
    if whonet_df.empty:
        return ""
    
    output = StringIO()
    whonet_df.to_csv(output, sep='\t', index=False, na_rep='')
    return output.getvalue()


def export_to_whonet_excel(whonet_df: pd.DataFrame, 
                           include_metadata: bool = True) -> bytes:
    """
    Export WHONET data to Excel format with formatting.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    if whonet_df.empty:
        return bytes()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "WHONET_Data"
    
    # Style definitions
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(whonet_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for col_idx, column in enumerate(whonet_df.columns, 1):
        max_length = max(
            len(str(column)),
            whonet_df[column].astype(str).str.len().max() if not whonet_df.empty else 0
        )
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = min(max_length + 2, 30)
    
    # Add metadata sheet if requested
    if include_metadata:
        meta_ws = wb.create_sheet(title="Metadata")
        meta_ws['A1'] = "WHONET Export Metadata"
        meta_ws['A1'].font = Font(bold=True, size=14, color="0891B2")
        
        metadata = [
            ("Export Date", datetime.now().strftime('%Y-%m-%d %H:%M')),
            ("Format Version", "WHONET 5.6"),
            ("Country", "Ghana (GHA)"),
            ("Total Records", len(whonet_df)),
            ("Unique Organisms", whonet_df['ORGANISM'].nunique() if 'ORGANISM' in whonet_df.columns else 0),
            ("Date Range", f"{whonet_df['SPEC_DATE'].min()} to {whonet_df['SPEC_DATE'].max()}" if 'SPEC_DATE' in whonet_df.columns else 'N/A'),
            ("", ""),
            ("Reference", "https://whonet.org/"),
            ("GLASS Reporting", "https://www.who.int/initiatives/glass")
        ]
        
        for idx, (key, value) in enumerate(metadata, 3):
            meta_ws[f'A{idx}'] = key
            meta_ws[f'B{idx}'] = value
            meta_ws[f'A{idx}'].font = Font(bold=True)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_glass_report(whonet_df: pd.DataFrame) -> Dict:
    """
    Generate summary statistics for WHO GLASS reporting.
    
    GLASS (Global Antimicrobial Resistance Surveillance System) requires
    specific summary data for national reporting.
    """
    if whonet_df.empty:
        return {'error': 'No data available'}
    
    # Priority pathogens for GLASS
    glass_pathogens = {
        'eco': 'Escherichia coli',
        'kpn': 'Klebsiella pneumoniae', 
        'sau': 'Staphylococcus aureus',
        'spn': 'Streptococcus pneumoniae',
        'sal': 'Salmonella spp.',
        'sty': 'Salmonella typhi',
        'shi': 'Shigella spp.',
        'ngo': 'Neisseria gonorrhoeae',
        'aba': 'Acinetobacter baumannii'
    }
    
    # Key antibiotics for GLASS surveillance
    glass_antibiotics = {
        'eco': ['CRO', 'CAZ', 'FEP', 'MEM', 'CIP', 'GEN', 'SXT', 'COL'],
        'kpn': ['CRO', 'CAZ', 'FEP', 'MEM', 'CIP', 'GEN', 'SXT', 'COL'],
        'sau': ['OXA', 'FOX', 'VAN', 'LNZ', 'DAP', 'RIF', 'SXT'],
        'spn': ['PEN', 'CRO', 'ERY', 'LVX', 'SXT'],
        'sal': ['CIP', 'CRO', 'AZM', 'SXT', 'CHL'],
        'sty': ['CIP', 'CRO', 'AZM', 'SXT', 'CHL'],
        'shi': ['CIP', 'CRO', 'AZM', 'SXT'],
        'ngo': ['CRO', 'AZM', 'CIP', 'TET'],
        'aba': ['MEM', 'IPM', 'CIP', 'GEN', 'COL', 'TGC']
    }
    
    summary = {
        'report_date': datetime.now().isoformat(),
        'country': 'Ghana',
        'country_code': 'GHA',
        'total_isolates': len(whonet_df),
        'reporting_period': {
            'start': whonet_df['SPEC_DATE'].min() if 'SPEC_DATE' in whonet_df.columns else None,
            'end': whonet_df['SPEC_DATE'].max() if 'SPEC_DATE' in whonet_df.columns else None
        },
        'organisms': {},
        'specimen_distribution': {},
        'resistance_rates': {}
    }
    
    # Organism counts
    if 'ORG_CODE' in whonet_df.columns:
        org_counts = whonet_df['ORG_CODE'].value_counts().to_dict()
        summary['organisms'] = {
            glass_pathogens.get(code, code): count 
            for code, count in org_counts.items()
        }
    
    # Specimen distribution
    if 'SPEC_TYPE' in whonet_df.columns:
        summary['specimen_distribution'] = whonet_df['SPEC_TYPE'].value_counts().to_dict()
    
    # Calculate resistance rates for priority pathogen-antibiotic combinations
    resistance_data = {}
    
    for org_code, org_name in glass_pathogens.items():
        if 'ORG_CODE' not in whonet_df.columns:
            continue
            
        org_data = whonet_df[whonet_df['ORG_CODE'] == org_code]
        
        if org_data.empty:
            continue
        
        org_resistance = {}
        for ab_code in glass_antibiotics.get(org_code, []):
            if ab_code in org_data.columns:
                results = org_data[ab_code].dropna()
                if not results.empty:
                    total = len(results)
                    resistant = (results == 'R').sum()
                    intermediate = (results == 'I').sum()
                    susceptible = (results == 'S').sum()
                    
                    org_resistance[ab_code] = {
                        'tested': total,
                        'resistant': resistant,
                        'intermediate': intermediate,
                        'susceptible': susceptible,
                        'resistance_rate': (resistant / total * 100) if total > 0 else None,
                        'non_susceptible_rate': ((resistant + intermediate) / total * 100) if total > 0 else None
                    }
        
        if org_resistance:
            resistance_data[org_name] = {
                'isolate_count': len(org_data),
                'antibiotics': org_resistance
            }
    
    summary['resistance_rates'] = resistance_data
    
    return summary


def validate_whonet_data(whonet_df: pd.DataFrame) -> Dict:
    """
    Validate WHONET formatted data for completeness and quality.
    """
    validation = {
        'is_valid': True,
        'errors': [],
        'warnings': [],
        'statistics': {}
    }
    
    required_fields = ['LABORATORY', 'SPEC_NUM', 'SPEC_DATE', 'ORGANISM', 'ORG_CODE']
    
    # Check required fields
    for field in required_fields:
        if field not in whonet_df.columns:
            validation['errors'].append(f"Missing required field: {field}")
            validation['is_valid'] = False
        elif whonet_df[field].isna().all():
            validation['errors'].append(f"Required field {field} is empty")
            validation['is_valid'] = False
    
    # Check for valid organism codes
    if 'ORG_CODE' in whonet_df.columns:
        unknown_orgs = whonet_df[~whonet_df['ORG_CODE'].isin(WHONET_ORGANISM_CODES.values())]['ORG_CODE'].unique()
        if len(unknown_orgs) > 0:
            validation['warnings'].append(f"Non-standard organism codes: {', '.join(unknown_orgs[:5])}")
    
    # Check date format
    if 'SPEC_DATE' in whonet_df.columns:
        try:
            pd.to_datetime(whonet_df['SPEC_DATE'], errors='raise')
        except:
            validation['warnings'].append("Some specimen dates may not be in standard format")
    
    # Check for antibiotic results
    ab_cols = [col for col in whonet_df.columns if re.match(r'^[A-Z]{3}$', col)]
    if not ab_cols:
        validation['errors'].append("No antibiotic susceptibility data found")
        validation['is_valid'] = False
    else:
        validation['statistics']['antibiotic_count'] = len(ab_cols)
        validation['statistics']['antibiotics'] = ab_cols
    
    # Summary statistics
    validation['statistics']['total_records'] = len(whonet_df)
    validation['statistics']['unique_organisms'] = whonet_df['ORGANISM'].nunique() if 'ORGANISM' in whonet_df.columns else 0
    validation['statistics']['date_range'] = {
        'min': whonet_df['SPEC_DATE'].min() if 'SPEC_DATE' in whonet_df.columns else None,
        'max': whonet_df['SPEC_DATE'].max() if 'SPEC_DATE' in whonet_df.columns else None
    }
    
    return validation


def generate_glass_html_report(glass_report: Dict) -> str:
    """
    Generate a formatted HTML report for WHO GLASS submission.
    """
    if not glass_report or 'error' in glass_report:
        return "<p>No GLASS data available.</p>"
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>WHO GLASS Summary Report</title>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
            .container {{ max-width: 900px; margin: 0 auto; background: white; border-radius: 10px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #0891b2, #0d9488); color: white; padding: 25px; border-radius: 10px 10px 0 0; margin: -30px -30px 25px -30px; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .header p {{ margin: 5px 0 0 0; opacity: 0.9; }}
            .logo {{ font-size: 40px; margin-bottom: 10px; }}
            .section {{ margin: 25px 0; padding: 20px; background: #f8fafc; border-radius: 8px; }}
            .section h3 {{ color: #0f766e; margin-top: 0; border-bottom: 2px solid #0d9488; padding-bottom: 10px; }}
            .stat-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 20px; }}
            .stat-card {{ background: white; padding: 15px; border-radius: 8px; text-align: center; border: 1px solid #e2e8f0; }}
            .stat-value {{ font-size: 28px; font-weight: bold; color: #0d9488; }}
            .stat-label {{ color: #64748b; font-size: 12px; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #e2e8f0; }}
            th {{ background: #0f766e; color: white; }}
            tr:nth-child(even) {{ background: #f8fafc; }}
            .rate-high {{ color: #dc2626; font-weight: bold; }}
            .rate-medium {{ color: #f59e0b; font-weight: bold; }}
            .rate-low {{ color: #22c55e; font-weight: bold; }}
            .organism-card {{ background: white; padding: 15px; margin: 10px 0; border-radius: 8px; border-left: 4px solid #0d9488; }}
            .organism-name {{ font-weight: bold; color: #1f2937; font-size: 16px; }}
            .organism-count {{ color: #64748b; font-size: 14px; }}
            .antibiotic-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #e2e8f0; }}
            .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #e2e8f0; text-align: center; color: #64748b; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="logo">🌍</div>
                <h1>WHO GLASS Summary Report</h1>
                <p>Global Antimicrobial Resistance Surveillance System</p>
                <p>Country: {glass_report.get('country', 'Ghana')} ({glass_report.get('country_code', 'GHA')})</p>
            </div>
            
            <div class="stat-grid">
                <div class="stat-card">
                    <div class="stat-value">{glass_report.get('total_isolates', 0):,}</div>
                    <div class="stat-label">Total Isolates</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{len(glass_report.get('organisms', {}))}</div>
                    <div class="stat-label">Organisms</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{glass_report.get('reporting_period', {}).get('start', 'N/A')}</div>
                    <div class="stat-label">Period Start</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{glass_report.get('reporting_period', {}).get('end', 'N/A')}</div>
                    <div class="stat-label">Period End</div>
                </div>
            </div>
            
            <div class="section">
                <h3>Organism Distribution</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Organism</th>
                            <th>Isolate Count</th>
                            <th>Percentage</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    total_isolates = glass_report.get('total_isolates', 1)
    for org, count in glass_report.get('organisms', {}).items():
        pct = (count / total_isolates * 100) if total_isolates > 0 else 0
        html += f"""
                        <tr>
                            <td><strong>{org}</strong></td>
                            <td>{count}</td>
                            <td>{pct:.1f}%</td>
                        </tr>
        """
    
    html += """
                    </tbody>
                </table>
            </div>
            
            <div class="section">
                <h3>Specimen Type Distribution</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Specimen Type</th>
                            <th>Count</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    for spec, count in glass_report.get('specimen_distribution', {}).items():
        html += f"""
                        <tr>
                            <td>{spec}</td>
                            <td>{count}</td>
                        </tr>
        """
    
    html += """
                    </tbody>
                </table>
            </div>
            
            <div class="section">
                <h3>Priority Pathogen Resistance Rates</h3>
    """
    
    for org_name, data in glass_report.get('resistance_rates', {}).items():
        html += f"""
                <div class="organism-card">
                    <div class="organism-name">{org_name}</div>
                    <div class="organism-count">Isolates tested: {data.get('isolate_count', 0)}</div>
                    <table style="margin-top: 10px;">
                        <thead>
                            <tr>
                                <th>Antibiotic</th>
                                <th>Tested</th>
                                <th>Resistant</th>
                                <th>Resistance Rate</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        for ab_code, ab_data in data.get('antibiotics', {}).items():
            rate = ab_data.get('resistance_rate', 0) or 0
            rate_class = 'rate-high' if rate >= 50 else 'rate-medium' if rate >= 30 else 'rate-low'
            html += f"""
                            <tr>
                                <td>{ab_code}</td>
                                <td>{ab_data.get('tested', 0)}</td>
                                <td>{ab_data.get('resistant', 0)}</td>
                                <td class="{rate_class}">{rate:.1f}%</td>
                            </tr>
            """
        
        html += """
                        </tbody>
                    </table>
                </div>
        """
    
    html += f"""
            </div>
            
            <div class="footer">
                <p><strong>WHO GLASS Summary Report</strong></p>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
                <p>Ghana Food Safety Authority - AMR Surveillance Dashboard</p>
                <p><em>For official GLASS submission, please use the WHONET software with the exported data file.</em></p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html

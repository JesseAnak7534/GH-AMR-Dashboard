"""
Antibiogram Generator for AMR Surveillance Dashboard.
Generates CLSI-formatted cumulative antibiograms for clinical decision support.
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from io import BytesIO
import base64


# CLSI recommended minimum isolate count for reporting
CLSI_MIN_ISOLATES = 30

# Common organism groupings for antibiograms
ORGANISM_GROUPS = {
    "Enterobacterales": [
        "Escherichia coli", "Klebsiella pneumoniae", "Klebsiella oxytoca",
        "Enterobacter cloacae", "Enterobacter aerogenes", "Serratia marcescens",
        "Citrobacter freundii", "Citrobacter koseri", "Proteus mirabilis",
        "Proteus vulgaris", "Morganella morganii", "Providencia stuartii",
        "Salmonella", "Salmonella enterica", "Salmonella typhi", "Salmonella typhimurium",
        "Shigella", "Shigella flexneri", "Shigella sonnei"
    ],
    "Non-fermenting Gram-negatives": [
        "Pseudomonas aeruginosa", "Acinetobacter baumannii", "Acinetobacter",
        "Stenotrophomonas maltophilia", "Burkholderia cepacia"
    ],
    "Staphylococci": [
        "Staphylococcus aureus", "Staphylococcus epidermidis", 
        "Staphylococcus saprophyticus", "Staphylococcus haemolyticus",
        "Coagulase-negative staphylococci"
    ],
    "Enterococci": [
        "Enterococcus faecalis", "Enterococcus faecium", "Enterococcus"
    ],
    "Streptococci": [
        "Streptococcus pneumoniae", "Streptococcus pyogenes", 
        "Streptococcus agalactiae", "Viridans group streptococci"
    ],
    "Vibrio": [
        "Vibrio cholerae", "Vibrio parahaemolyticus", "Vibrio vulnificus"
    ],
    "Campylobacter": [
        "Campylobacter jejuni", "Campylobacter coli", "Campylobacter"
    ]
}


def get_organism_group(organism: str) -> Optional[str]:
    """Get the organism group for a given organism."""
    organism_lower = organism.lower()
    for group, organisms in ORGANISM_GROUPS.items():
        for org in organisms:
            if org.lower() in organism_lower or organism_lower in org.lower():
                return group
    return None


def calculate_susceptibility_rates(ast_df: pd.DataFrame, 
                                   organism_col: str = 'organism',
                                   antibiotic_col: str = 'antibiotic',
                                   result_col: str = 'result',
                                   min_isolates: int = CLSI_MIN_ISOLATES) -> pd.DataFrame:
    """
    Calculate susceptibility rates for each organism-antibiotic combination.
    Returns a pivoted DataFrame suitable for antibiogram display.
    """
    if ast_df.empty:
        return pd.DataFrame()
    
    # Get first isolate per patient/sample for deduplication (CLSI requirement)
    # Using isolate_id as proxy for unique isolates
    dedup_df = ast_df.drop_duplicates(subset=['isolate_id', antibiotic_col])
    
    # Calculate rates
    rates = []
    for (organism, antibiotic), group in dedup_df.groupby([organism_col, antibiotic_col]):
        total = len(group)
        susceptible = (group[result_col] == 'S').sum()
        intermediate = (group[result_col] == 'I').sum()
        resistant = (group[result_col] == 'R').sum()
        
        # Calculate percent susceptible (S only, not S+I)
        pct_susceptible = (susceptible / total) * 100 if total > 0 else None
        
        rates.append({
            'organism': organism,
            'antibiotic': antibiotic,
            'susceptible': susceptible,
            'intermediate': intermediate,
            'resistant': resistant,
            'total': total,
            'pct_susceptible': pct_susceptible,
            'meets_clsi': total >= min_isolates
        })
    
    return pd.DataFrame(rates)


def generate_antibiogram(ast_df: pd.DataFrame,
                         lab_name: Optional[str] = None,
                         time_period: Optional[str] = None,
                         min_isolates: int = CLSI_MIN_ISOLATES,
                         include_all: bool = False) -> Dict:
    """
    Generate a comprehensive antibiogram report.
    
    Args:
        ast_df: AST results DataFrame
        lab_name: Laboratory name for the report
        time_period: Time period description (e.g., "Q1 2026", "2025")
        min_isolates: Minimum isolates required for reporting (CLSI default: 30)
        include_all: Include combinations with fewer than min_isolates (marked with *)
    
    Returns:
        Dictionary containing antibiogram data and metadata
    """
    if ast_df.empty:
        return {'error': 'No data available', 'matrix': pd.DataFrame()}
    
    # Calculate rates
    rates_df = calculate_susceptibility_rates(ast_df, min_isolates=min_isolates)
    
    if rates_df.empty:
        return {'error': 'No valid combinations found', 'matrix': pd.DataFrame()}
    
    # Filter by minimum isolates if not including all
    if not include_all:
        display_df = rates_df[rates_df['meets_clsi']].copy()
    else:
        display_df = rates_df.copy()
    
    if display_df.empty:
        return {
            'error': f'No organism-antibiotic combinations with ≥{min_isolates} isolates',
            'matrix': pd.DataFrame(),
            'total_combinations': len(rates_df)
        }
    
    # Create pivot table for antibiogram matrix
    # Format: rows = organisms, columns = antibiotics, values = %S (n)
    def format_cell(row):
        if pd.isna(row['pct_susceptible']):
            return '-'
        
        marker = '' if row['meets_clsi'] else '*'
        return f"{row['pct_susceptible']:.0f}{marker} ({row['total']})"
    
    display_df['display_value'] = display_df.apply(format_cell, axis=1)
    
    matrix = display_df.pivot_table(
        index='organism',
        columns='antibiotic',
        values='display_value',
        aggfunc='first'
    ).fillna('-')
    
    # Create numeric matrix for heatmap
    numeric_matrix = display_df.pivot_table(
        index='organism',
        columns='antibiotic',
        values='pct_susceptible',
        aggfunc='first'
    )
    
    # Get counts matrix
    counts_matrix = display_df.pivot_table(
        index='organism',
        columns='antibiotic',
        values='total',
        aggfunc='first'
    ).fillna(0).astype(int)
    
    # Summary statistics
    summary = {
        'total_isolates': ast_df['isolate_id'].nunique(),
        'total_organisms': display_df['organism'].nunique(),
        'total_antibiotics': display_df['antibiotic'].nunique(),
        'total_combinations': len(display_df),
        'combinations_below_threshold': (~display_df['meets_clsi']).sum(),
        'min_isolates_threshold': min_isolates,
        'overall_susceptibility': display_df['pct_susceptible'].mean(),
        'highest_resistance': {
            'organism': None,
            'antibiotic': None,
            'rate': None
        },
        'lowest_susceptibility_combinations': []
    }
    
    # Find highest resistance (lowest susceptibility)
    if not display_df.empty:
        lowest_susc = display_df.loc[display_df['pct_susceptible'].idxmin()]
        summary['highest_resistance'] = {
            'organism': lowest_susc['organism'],
            'antibiotic': lowest_susc['antibiotic'],
            'rate': 100 - lowest_susc['pct_susceptible']
        }
        
        # Top 5 lowest susceptibility combinations
        low_susc = display_df.nsmallest(5, 'pct_susceptible')[
            ['organism', 'antibiotic', 'pct_susceptible', 'total']
        ].to_dict('records')
        summary['lowest_susceptibility_combinations'] = low_susc
    
    return {
        'matrix': matrix,
        'numeric_matrix': numeric_matrix,
        'counts_matrix': counts_matrix,
        'rates_data': display_df,
        'summary': summary,
        'lab_name': lab_name or 'All Laboratories',
        'time_period': time_period or 'All Time',
        'generated_at': datetime.now().isoformat(),
        'clsi_note': f"Values represent % Susceptible (number tested). * indicates <{min_isolates} isolates (interpret with caution)."
    }


def generate_antibiogram_by_facility(ast_df: pd.DataFrame, 
                                     samples_df: pd.DataFrame,
                                     min_isolates: int = CLSI_MIN_ISOLATES) -> Dict[str, Dict]:
    """
    Generate antibiograms for each laboratory/facility.
    """
    antibiograms = {}
    
    if ast_df.empty or samples_df.empty:
        return antibiograms
    
    # Merge to get lab names
    merged = ast_df.merge(samples_df[['sample_id', 'lab_name']], on='sample_id', how='left')
    
    for lab_name in merged['lab_name'].dropna().unique():
        lab_data = merged[merged['lab_name'] == lab_name]
        antibiograms[lab_name] = generate_antibiogram(
            lab_data, 
            lab_name=lab_name,
            min_isolates=min_isolates
        )
    
    return antibiograms


def generate_quarterly_antibiograms(ast_df: pd.DataFrame,
                                    samples_df: pd.DataFrame,
                                    year: int = None,
                                    min_isolates: int = 10) -> Dict[str, Dict]:
    """
    Generate antibiograms for each quarter.
    Uses lower threshold for quarterly data.
    """
    antibiograms = {}
    
    if ast_df.empty or samples_df.empty:
        return antibiograms
    
    # Merge to get collection dates
    merged = ast_df.merge(samples_df[['sample_id', 'collection_date']], on='sample_id', how='left')
    merged['collection_date'] = pd.to_datetime(merged['collection_date'], errors='coerce')
    merged = merged.dropna(subset=['collection_date'])
    
    if merged.empty:
        return antibiograms
    
    # Get year if not specified
    if year is None:
        year = merged['collection_date'].dt.year.max()
    
    # Filter by year
    merged = merged[merged['collection_date'].dt.year == year]
    merged['quarter'] = merged['collection_date'].dt.quarter
    
    for quarter in sorted(merged['quarter'].unique()):
        quarter_data = merged[merged['quarter'] == quarter]
        period_name = f"Q{quarter} {year}"
        
        antibiograms[period_name] = generate_antibiogram(
            quarter_data,
            time_period=period_name,
            min_isolates=min_isolates,
            include_all=True
        )
    
    return antibiograms


def compare_antibiograms(antibiogram1: Dict, antibiogram2: Dict,
                         period1_name: str = "Period 1",
                         period2_name: str = "Period 2") -> pd.DataFrame:
    """
    Compare two antibiograms and calculate differences.
    Useful for tracking changes over time.
    """
    if 'numeric_matrix' not in antibiogram1 or 'numeric_matrix' not in antibiogram2:
        return pd.DataFrame()
    
    matrix1 = antibiogram1['numeric_matrix']
    matrix2 = antibiogram2['numeric_matrix']
    
    # Align matrices
    common_organisms = set(matrix1.index) & set(matrix2.index)
    common_antibiotics = set(matrix1.columns) & set(matrix2.columns)
    
    if not common_organisms or not common_antibiotics:
        return pd.DataFrame()
    
    comparison_data = []
    for organism in common_organisms:
        for antibiotic in common_antibiotics:
            val1 = matrix1.loc[organism, antibiotic] if organism in matrix1.index and antibiotic in matrix1.columns else None
            val2 = matrix2.loc[organism, antibiotic] if organism in matrix2.index and antibiotic in matrix2.columns else None
            
            if pd.notna(val1) and pd.notna(val2):
                diff = val2 - val1
                comparison_data.append({
                    'Organism': organism,
                    'Antibiotic': antibiotic,
                    f'{period1_name} %S': val1,
                    f'{period2_name} %S': val2,
                    'Change': diff,
                    'Direction': '↑' if diff > 0 else ('↓' if diff < 0 else '→'),
                    'Significant': abs(diff) >= 10
                })
    
    return pd.DataFrame(comparison_data)


def antibiogram_to_html(antibiogram: Dict, include_legend: bool = True) -> str:
    """
    Convert antibiogram to styled HTML for reports.
    Uses color coding based on susceptibility rates.
    """
    if 'matrix' not in antibiogram or antibiogram['matrix'].empty:
        return "<p>No antibiogram data available.</p>"
    
    matrix = antibiogram['matrix']
    numeric = antibiogram.get('numeric_matrix', pd.DataFrame())
    
    # Color scale function
    def get_color(value):
        if pd.isna(value):
            return '#f0f0f0'
        if value >= 90:
            return '#10b981'  # Green - high susceptibility
        elif value >= 70:
            return '#84cc16'  # Lime
        elif value >= 50:
            return '#fbbf24'  # Yellow
        elif value >= 30:
            return '#f97316'  # Orange
        else:
            return '#ef4444'  # Red - high resistance
    
    html = f"""
    <div style="overflow-x: auto; max-width: 100%;">
        <h3 style="color: #0f766e; margin-bottom: 10px;">Cumulative Antibiogram</h3>
        <p style="color: #64748b; font-size: 0.9em;">
            <strong>Laboratory:</strong> {antibiogram.get('lab_name', 'All')}<br>
            <strong>Period:</strong> {antibiogram.get('time_period', 'All Time')}<br>
            <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M')}
        </p>
        
        <table style="border-collapse: collapse; font-size: 11px; min-width: 100%; table-layout: auto;">
            <thead>
                <tr style="background: #0f766e; color: white;">
                    <th style="padding: 8px; text-align: left; border: 1px solid #ddd; min-width: 150px; white-space: nowrap;">Organism</th>
    """
    
    # Header row with antibiotics - use abbreviations for cleaner display
    for antibiotic in matrix.columns:
        # Abbreviate long antibiotic names
        ab_display = antibiotic[:3].upper() if len(antibiotic) > 10 else antibiotic
        html += f'<th style="padding: 6px 4px; text-align: center; border: 1px solid #ddd; font-size: 10px; min-width: 40px;" title="{antibiotic}">{ab_display}</th>'
    
    html += "</tr></thead><tbody>"
    
    # Data rows
    for organism in matrix.index:
        html += f'<tr><td style="padding: 6px 8px; border: 1px solid #ddd; font-weight: 500; white-space: nowrap; background: #f8fafc;">{organism}</td>'
        
        for antibiotic in matrix.columns:
            value = matrix.loc[organism, antibiotic]
            numeric_val = numeric.loc[organism, antibiotic] if organism in numeric.index and antibiotic in numeric.columns else None
            
            bg_color = get_color(numeric_val) if pd.notna(numeric_val) else '#f0f0f0'
            text_color = 'white' if pd.notna(numeric_val) and numeric_val < 50 else '#1f2937'
            
            html += f'<td style="padding: 4px; text-align: center; border: 1px solid #ddd; background: {bg_color}; color: {text_color}; font-size: 10px;">{value}</td>'
        
        html += "</tr>"
    
    html += "</tbody></table>"
    
    # Legend
    if include_legend:
        html += """
        <div style="margin-top: 15px; padding: 10px; background: #f8fafc; border-radius: 8px;">
            <p style="font-weight: 600; margin-bottom: 5px;">Legend:</p>
            <div style="display: flex; gap: 15px; flex-wrap: wrap; font-size: 12px;">
                <span><span style="display: inline-block; width: 15px; height: 15px; background: #10b981; margin-right: 5px; vertical-align: middle;"></span>≥90% Susceptible</span>
                <span><span style="display: inline-block; width: 15px; height: 15px; background: #84cc16; margin-right: 5px; vertical-align: middle;"></span>70-89%</span>
                <span><span style="display: inline-block; width: 15px; height: 15px; background: #fbbf24; margin-right: 5px; vertical-align: middle;"></span>50-69%</span>
                <span><span style="display: inline-block; width: 15px; height: 15px; background: #f97316; margin-right: 5px; vertical-align: middle;"></span>30-49%</span>
                <span><span style="display: inline-block; width: 15px; height: 15px; background: #ef4444; margin-right: 5px; vertical-align: middle;"></span><30%</span>
            </div>
        """
        html += f"<p style='margin-top: 10px; font-size: 11px; color: #64748b;'>{antibiogram.get('clsi_note', '')}</p>"
        html += "</div></div>"
    
    return html


def antibiogram_to_excel(antibiogram: Dict) -> bytes:
    """
    Export antibiogram to Excel with formatting.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    if 'matrix' not in antibiogram or antibiogram['matrix'].empty:
        return bytes()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Antibiogram"
    
    # Title
    ws['A1'] = f"Cumulative Antibiogram - {antibiogram.get('lab_name', 'All Laboratories')}"
    ws['A1'].font = Font(bold=True, size=14, color="0F766E")
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Period: {antibiogram.get('time_period', 'All Time')} | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:H2')
    
    # Matrix data starting at row 4
    matrix = antibiogram['matrix']
    numeric = antibiogram.get('numeric_matrix', pd.DataFrame())
    
    # Colors
    green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    lime_fill = PatternFill(start_color="84CC16", end_color="84CC16", fill_type="solid")
    yellow_fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row
    ws.cell(row=4, column=1, value="Organism").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=4, column=1).fill = header_fill
    ws.cell(row=4, column=1).border = thin_border
    
    for col_idx, antibiotic in enumerate(matrix.columns, start=2):
        cell = ws.cell(row=4, column=col_idx, value=antibiotic)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(textRotation=90, horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, organism in enumerate(matrix.index, start=5):
        ws.cell(row=row_idx, column=1, value=organism).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).border = thin_border
        
        for col_idx, antibiotic in enumerate(matrix.columns, start=2):
            value = matrix.loc[organism, antibiotic]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
            # Color coding
            numeric_val = numeric.loc[organism, antibiotic] if organism in numeric.index and antibiotic in numeric.columns else None
            if pd.notna(numeric_val):
                if numeric_val >= 90:
                    cell.fill = green_fill
                elif numeric_val >= 70:
                    cell.fill = lime_fill
                elif numeric_val >= 50:
                    cell.fill = yellow_fill
                elif numeric_val >= 30:
                    cell.fill = orange_fill
                else:
                    cell.fill = red_fill
                    cell.font = Font(color="FFFFFF")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, len(matrix.columns) + 2):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = 8
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

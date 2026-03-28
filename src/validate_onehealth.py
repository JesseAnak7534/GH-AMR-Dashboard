"""
Validation and template generation for unified One Health data uploads.
Single Excel workbook with 6 sheets: samples, ast_results, pps_survey, prescriptions, amu_data, amc_data.
Users fill only the sheets they need.
"""
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Tuple, List, Dict, Optional
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from src.lab_management import APPROVED_LABS


# ════════════════════════════════════════════════════════════════════════════
# COLUMN SCHEMAS
# ════════════════════════════════════════════════════════════════════════════

REQUIRED_PPS_SURVEY_COLS = {
    'facility_name', 'survey_date', 'region', 'district',
    'total_patients', 'patients_on_antibiotics'
}

REQUIRED_PPS_RX_COLS = {
    'ward', 'patient_age_group', 'antibiotic_name', 'route',
    'indication', 'indication_documented', 'guideline_compliant', 'duration_days'
}

REQUIRED_AMU_COLS = {
    'facility_name', 'report_period', 'region', 'antibiotic_name',
    'quantity_dispensed'
}

REQUIRED_AMC_COLS = {
    'report_period', 'sector', 'antibiotic_class', 'quantity_kg'
}


# ════════════════════════════════════════════════════════════════════════════
# UNIFIED VALIDATOR  –  detects which sheets are present, validates each
# ════════════════════════════════════════════════════════════════════════════

def validate_unified_upload(file_obj) -> Tuple[bool, List[str], Dict]:
    """
    Validate a unified One Health Excel upload.

    Returns (ok, errors, result_dict) where result_dict may contain:
        'pps_survey'       -> dict   (single survey row)
        'pps_prescriptions'-> DataFrame
        'amu_data'         -> DataFrame
        'amc_data'         -> DataFrame
    Only keys for detected sheets are present.
    """
    errors: List[str] = []
    result: Dict = {}
    found_any = False

    try:
        wb = load_workbook(file_obj, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return False, [f"Cannot read Excel file: {e}"], {}

    # ── PPS (needs both sheets) ─────────────────────────────────────────
    if 'pps_survey' in sheets and 'prescriptions' in sheets:
        found_any = True
        file_obj.seek(0)
        survey_df = pd.read_excel(file_obj, sheet_name='pps_survey')
        file_obj.seek(0)
        rx_df = pd.read_excel(file_obj, sheet_name='prescriptions')

        if survey_df.empty:
            errors.append("[PPS] pps_survey sheet is empty")
        else:
            missing = REQUIRED_PPS_SURVEY_COLS - set(survey_df.columns)
            if missing:
                errors.append(f"[PPS] Missing pps_survey columns: {', '.join(sorted(missing))}")
            else:
                result['pps_survey'] = survey_df

        if rx_df.empty:
            errors.append("[PPS] prescriptions sheet is empty")
        else:
            missing = REQUIRED_PPS_RX_COLS - set(rx_df.columns)
            if missing:
                errors.append(f"[PPS] Missing prescriptions columns: {', '.join(sorted(missing))}")
            else:
                result['pps_prescriptions'] = rx_df

    elif 'pps_survey' in sheets or 'prescriptions' in sheets:
        errors.append("[PPS] Both 'pps_survey' and 'prescriptions' sheets are needed for PPS data")

    # ── AMU ─────────────────────────────────────────────────────────────
    if 'amu_data' in sheets:
        found_any = True
        file_obj.seek(0)
        amu_df = pd.read_excel(file_obj, sheet_name='amu_data')
        if amu_df.empty:
            errors.append("[AMU] amu_data sheet is empty")
        else:
            missing = REQUIRED_AMU_COLS - set(amu_df.columns)
            if missing:
                errors.append(f"[AMU] Missing columns: {', '.join(sorted(missing))}")
            else:
                for col in ['quantity_dispensed']:
                    if col in amu_df.columns:
                        bad = pd.to_numeric(amu_df[col], errors='coerce').isna() & amu_df[col].notna()
                        if bad.any():
                            errors.append(f"[AMU] Non-numeric values in {col}")
                if not errors or not any('[AMU]' in e for e in errors):
                    result['amu_data'] = amu_df

    # ── AMC ─────────────────────────────────────────────────────────────
    if 'amc_data' in sheets:
        found_any = True
        file_obj.seek(0)
        amc_df = pd.read_excel(file_obj, sheet_name='amc_data')
        if amc_df.empty:
            errors.append("[AMC] amc_data sheet is empty")
        else:
            missing = REQUIRED_AMC_COLS - set(amc_df.columns)
            if missing:
                errors.append(f"[AMC] Missing columns: {', '.join(sorted(missing))}")
            else:
                for col in ['quantity_kg', 'biomass_kg', 'mg_per_kg_biomass']:
                    if col in amc_df.columns:
                        bad = pd.to_numeric(amc_df[col], errors='coerce').isna() & amc_df[col].notna()
                        if bad.any():
                            errors.append(f"[AMC] Non-numeric values in {col}")
                valid_sectors = {'ANIMAL', 'AQUACULTURE'}
                if 'sector' in amc_df.columns:
                    inv = amc_df[~amc_df['sector'].isin(valid_sectors)]['sector'].dropna().unique()
                    if len(inv) > 0:
                        errors.append(f"[AMC] Invalid sector values: {', '.join(str(s) for s in inv)}")
                if not any('[AMC]' in e for e in errors):
                    result['amc_data'] = amc_df

    if not found_any:
        errors.append(
            "No recognised One Health sheets found. "
            "Expected some of: pps_survey, prescriptions, amu_data, amc_data"
        )

    return len(errors) == 0, errors, result


# ════════════════════════════════════════════════════════════════════════════
# UNIFIED TEMPLATE  –  one Excel file, 6 sheets
# ════════════════════════════════════════════════════════════════════════════

def create_unified_template() -> bytes:
    """
    Generate ONE Excel template with all sheets:
      samples, ast_results, pps_survey, prescriptions, amu_data, amc_data.
    Each sheet has example rows and colour-coded headers.
    """
    lab_names = sorted(APPROVED_LABS.keys())

    # ── Samples ─────────────────────────────────────────────────────────
    samples = pd.DataFrame({
        'sample_id': ['SAMPLE_001', 'SAMPLE_002', 'SAMPLE_003'],
        'lab_name': [lab_names[0] if lab_names else 'Lab 1',
                     lab_names[1] if len(lab_names) > 1 else 'Lab 2',
                     lab_names[2] if len(lab_names) > 2 else 'Lab 3'],
        'collection_date': ['2024-01-15', '2024-01-20', '2024-01-25'],
        'region': ['Ashanti', 'Greater Accra', 'Eastern'],
        'district': ['Kumasi', 'Accra', 'Koforidua'],
        'site_type': ['Water Treatment Plant', 'Retail Market', 'Hospital Lab'],
        'source_category': ['ENVIRONMENT', 'FOOD', 'HUMAN'],
        'source_type': ['treated_water', 'raw_chicken', 'clinical_specimen'],
        'food_matrix': ['', 'chicken', ''],
        'environment_matrix': ['treated_water', '', ''],
        'latitude': [6.6326, 5.6037, 6.1256],
        'longitude': [-1.6243, -0.1870, -0.3597],
    })

    ast = pd.DataFrame({
        'sample_id': ['SAMPLE_001', 'SAMPLE_001', 'SAMPLE_002'],
        'isolate_id': ['ISO_001', 'ISO_002', 'ISO_003'],
        'organism': ['E. coli', 'E. coli', 'Salmonella'],
        'antibiotic': ['Ampicillin', 'Ciprofloxacin', 'Ampicillin'],
        'result': ['R', 'S', 'I'],
        'method': ['DD', 'DD', 'MIC'],
        'guideline': ['CLSI', 'EUCAST', 'CLSI'],
        'test_date': ['2024-01-20', '2024-01-20', '2024-01-22'],
        'mic_value': [np.nan, np.nan, 0.5],
        'zone_diameter': [15.0, 28.0, np.nan],
    })

    # ── PPS ─────────────────────────────────────────────────────────────
    pps_survey = pd.DataFrame({
        'facility_name': ['Korle-Bu Teaching Hospital'],
        'survey_date': ['2026-03-15'],
        'region': ['Greater Accra'],
        'district': ['Accra Metropolis'],
        'total_patients': [120],
        'patients_on_antibiotics': [45],
    })

    prescriptions = pd.DataFrame({
        'ward': ['Medical', 'Surgical', 'Paediatric'],
        'patient_age_group': ['Adult (25-44)', 'Geriatric (65+)', 'Child (5-14)'],
        'antibiotic_name': ['Amoxicillin', 'Ceftriaxone', 'Metronidazole'],
        'route': ['Oral', 'IV', 'Oral'],
        'indication': ['Community-acquired pneumonia', 'Surgical prophylaxis', 'Intra-abdominal infection'],
        'indication_documented': [1, 1, 0],
        'guideline_compliant': [1, 0, 1],
        'duration_days': [7, 1, 5],
    })

    # ── AMU ─────────────────────────────────────────────────────────────
    amu = pd.DataFrame({
        'facility_name': ['Korle-Bu Teaching Hospital', 'Korle-Bu Teaching Hospital', 'Tamale Teaching Hospital'],
        'report_period': ['2026-Q1', '2026-Q1', '2026-Q1'],
        'region': ['Greater Accra', 'Greater Accra', 'Northern'],
        'district': ['Accra Metropolis', 'Accra Metropolis', 'Tamale Metropolis'],
        'sector': ['HUMAN', 'HUMAN', 'HUMAN'],
        'antibiotic_name': ['Amoxicillin', 'Ciprofloxacin', 'Ceftriaxone'],
        'atc_code': ['J01CA04', 'J01MA02', 'J01DD04'],
        'formulation': ['500mg capsule', '500mg tablet', '1g injection'],
        'unit_of_measure': ['DDD', 'DDD', 'DDD'],
        'quantity_dispensed': [1500, 800, 350],
        'ddd_per_1000': [45.2, 22.1, 12.8],
        'patient_days': [33200, 33200, 15600],
    })

    # ── AMC ─────────────────────────────────────────────────────────────
    amc = pd.DataFrame({
        'report_period': ['2026-Q1', '2026-Q1', '2026-Q1'],
        'region': ['Greater Accra', 'Ashanti', 'Northern'],
        'sector': ['ANIMAL', 'ANIMAL', 'AQUACULTURE'],
        'species': ['Poultry', 'Bovine', 'Tilapia'],
        'production_type': ['Broiler', 'Dairy', 'Farm'],
        'antibiotic_class': ['Tetracyclines', 'Penicillins', 'Fluoroquinolones'],
        'antibiotic_name': ['Oxytetracycline', 'Amoxicillin', 'Enrofloxacin'],
        'atc_vet_code': ['QJ01AA06', 'QJ01CA04', 'QJ01MA90'],
        'quantity_kg': [120.5, 45.0, 8.2],
        'biomass_kg': [500000, 300000, 100000],
        'mg_per_kg_biomass': [241.0, 150.0, 82.0],
        'route': ['Oral (water)', 'Injection', 'Oral (feed)'],
        'purpose': ['Therapeutic', 'Therapeutic', 'Prophylactic'],
    })

    # ── Write to Excel with coloured headers ────────────────────────────
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        samples.to_excel(w, sheet_name='samples', index=False)
        ast.to_excel(w, sheet_name='ast_results', index=False)
        pps_survey.to_excel(w, sheet_name='pps_survey', index=False)
        prescriptions.to_excel(w, sheet_name='prescriptions', index=False)
        amu.to_excel(w, sheet_name='amu_data', index=False)
        amc.to_excel(w, sheet_name='amc_data', index=False)

    # Style headers + add lab dropdown validation
    from openpyxl.styles import PatternFill, Font, Alignment

    HEADER_STYLES = {
        'samples':       PatternFill('solid', fgColor='1F4E79'),
        'ast_results':   PatternFill('solid', fgColor='2E75B6'),
        'pps_survey':    PatternFill('solid', fgColor='548235'),
        'prescriptions': PatternFill('solid', fgColor='70AD47'),
        'amu_data':      PatternFill('solid', fgColor='BF8F00'),
        'amc_data':      PatternFill('solid', fgColor='C55A11'),
    }
    white_font = Font(color='FFFFFF', bold=True)

    buf.seek(0)
    wb = load_workbook(buf)

    for sheet_name, fill in HEADER_STYLES.items():
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')

    # Lab name dropdown on samples sheet
    if 'lab_lists' not in wb.sheetnames:
        list_ws = wb.create_sheet('lab_lists')
    else:
        list_ws = wb['lab_lists']
    list_ws['A1'] = 'lab_name_list'
    for idx, lab in enumerate(lab_names, start=2):
        list_ws[f"A{idx}"] = lab
    list_range = f"lab_lists!$A$2:$A${len(lab_names) + 1}"
    dv = DataValidation(type="list", formula1=list_range, allow_blank=False)
    wb['samples'].add_data_validation(dv)
    dv.add("B2:B1000")
    list_ws.sheet_state = 'hidden'

    # Instructions sheet
    instr = wb.create_sheet('_INSTRUCTIONS', 0)
    instructions = [
        "AMR ONE HEALTH SURVEILLANCE – UNIFIED DATA TEMPLATE",
        "",
        "Fill ONLY the sheets relevant to your data submission:",
        "",
        "  SHEET                 PURPOSE",
        "  ─────────────────     ──────────────────────────────────────",
        "  samples               Sample metadata (ID, location, source)",
        "  ast_results           Antimicrobial susceptibility test results",
        "  pps_survey            Point Prevalence Survey facility summary",
        "  prescriptions         Individual prescription records (PPS)",
        "  amu_data              Antimicrobial Use data (human, DDD)",
        "  amc_data              Antimicrobial Consumption (animal/aqua, kg)",
        "",
        "RULES:",
        "  • samples + ast_results are the core AMR sheets (always needed for resistance data)",
        "  • pps_survey + prescriptions must both be filled for PPS",
        "  • amu_data and amc_data are independent",
        "  • Leave unused sheets empty (do not delete them)",
        "  • Yellow/orange columns = required fields; others are optional",
        "",
        "Delete the example rows before entering your real data.",
    ]
    for i, line in enumerate(instructions, 1):
        instr.cell(row=i, column=1, value=line)
    instr.column_dimensions['A'].width = 80
    title_fill = PatternFill('solid', fgColor='0D1117')
    instr.cell(row=1, column=1).fill = title_fill
    instr.cell(row=1, column=1).font = Font(color='FFFFFF', bold=True, size=14)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# Legacy single-module templates (kept for backward compat in dashboard tabs)
# ════════════════════════════════════════════════════════════════════════════

def create_pps_template() -> bytes:
    pps_survey = pd.DataFrame({
        'facility_name': ['Korle-Bu Teaching Hospital'],
        'survey_date': ['2026-03-15'],
        'region': ['Greater Accra'],
        'district': ['Accra Metropolis'],
        'total_patients': [120],
        'patients_on_antibiotics': [45],
    })
    rx = pd.DataFrame({
        'ward': ['Medical', 'Surgical', 'Paediatric'],
        'patient_age_group': ['Adult (25-44)', 'Geriatric (65+)', 'Child (5-14)'],
        'antibiotic_name': ['Amoxicillin', 'Ceftriaxone', 'Metronidazole'],
        'route': ['Oral', 'IV', 'Oral'],
        'indication': ['Community-acquired pneumonia', 'Surgical prophylaxis', 'Intra-abdominal infection'],
        'indication_documented': [1, 1, 0],
        'guideline_compliant': [1, 0, 1],
        'duration_days': [7, 1, 5],
    })
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        pps_survey.to_excel(w, sheet_name='pps_survey', index=False)
        rx.to_excel(w, sheet_name='prescriptions', index=False)
    return buf.getvalue()


def create_amu_template() -> bytes:
    data = pd.DataFrame({
        'facility_name': ['Korle-Bu Teaching Hospital', 'Tamale Teaching Hospital'],
        'report_period': ['2026-Q1', '2026-Q1'],
        'region': ['Greater Accra', 'Northern'],
        'district': ['Accra Metropolis', 'Tamale Metropolis'],
        'sector': ['HUMAN', 'HUMAN'],
        'antibiotic_name': ['Amoxicillin', 'Ceftriaxone'],
        'atc_code': ['J01CA04', 'J01DD04'],
        'formulation': ['500mg capsule', '1g injection'],
        'unit_of_measure': ['DDD', 'DDD'],
        'quantity_dispensed': [1500, 350],
        'ddd_per_1000': [45.2, 12.8],
        'patient_days': [33200, 15600],
    })
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        data.to_excel(w, sheet_name='amu_data', index=False)
    return buf.getvalue()


def create_amc_template() -> bytes:
    data = pd.DataFrame({
        'report_period': ['2026-Q1', '2026-Q1'],
        'region': ['Greater Accra', 'Ashanti'],
        'sector': ['ANIMAL', 'AQUACULTURE'],
        'species': ['Poultry', 'Tilapia'],
        'production_type': ['Broiler', 'Farm'],
        'antibiotic_class': ['Tetracyclines', 'Fluoroquinolones'],
        'antibiotic_name': ['Oxytetracycline', 'Enrofloxacin'],
        'atc_vet_code': ['QJ01AA06', 'QJ01MA90'],
        'quantity_kg': [120.5, 8.2],
        'biomass_kg': [500000, 100000],
        'mg_per_kg_biomass': [241.0, 82.0],
        'route': ['Oral (water)', 'Oral (feed)'],
        'purpose': ['Therapeutic', 'Prophylactic'],
    })
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        data.to_excel(w, sheet_name='amc_data', index=False)
    return buf.getvalue()


# Legacy aliases
def validate_pps_upload(file_obj):
    """Legacy: validate PPS-only upload via unified validator."""
    ok, errs, result = validate_unified_upload(file_obj)
    survey = result.get('pps_survey', {})
    rx = result.get('pps_prescriptions', pd.DataFrame())
    return ok, errs, survey, rx

def validate_amu_upload(file_obj):
    ok, errs, result = validate_unified_upload(file_obj)
    return ok, errs, result.get('amu_data', pd.DataFrame())

def validate_amc_upload(file_obj):
    ok, errs, result = validate_unified_upload(file_obj)
    return ok, errs, result.get('amc_data', pd.DataFrame())

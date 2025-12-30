"""
Validation module for AMR data uploads.
Handles Excel parsing, schema validation, and automated interpretation.
"""
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Tuple, List, Dict
from openpyxl import load_workbook
import io

# Import interpretation engine
from src.interpretation import batch_interpret_results, interpret_ast_result


# Required columns per sheet
REQUIRED_SAMPLES_COLUMNS = {
    'sample_id', 'collection_date', 'region', 'district', 'site_type',
    'source_category', 'source_type', 'food_matrix', 'environment_matrix',
    'latitude', 'longitude'
}

REQUIRED_AST_COLUMNS = {
    'sample_id', 'isolate_id', 'organism', 'antibiotic', 'result',
    'method', 'guideline', 'test_date', 'mic_value', 'zone_diameter'
}

VALID_SOURCE_CATEGORIES = {'ENVIRONMENT', 'FOOD', 'HUMAN', 'ANIMAL', 'AQUACULTURE'}
VALID_RESULTS = {'S', 'I', 'R'}
VALID_METHODS = {'DD', 'MIC'}
VALID_GUIDELINES = {'CLSI', 'EUCAST'}
# Note: source_type and site_type accept all values (no restrictions)


def validate_excel_structure(file_obj) -> Tuple[bool, str, Dict]:
    """Validate that Excel file has required sheets."""
    try:
        wb = load_workbook(file_obj)
        sheets = wb.sheetnames
        
        if 'samples' not in sheets or 'ast_results' not in sheets:
            return False, "Excel must have 'samples' and 'ast_results' sheets", {}
        
        return True, "Structure valid", {"sheets": sheets}
    except Exception as e:
        return False, f"Error reading Excel: {str(e)}", {}


def load_excel_sheets(file_obj) -> Tuple[pd.DataFrame, pd.DataFrame, str]:
    """Load samples and ast_results sheets from Excel."""
    try:
        samples_df = pd.read_excel(file_obj, sheet_name='samples')
        ast_df = pd.read_excel(file_obj, sheet_name='ast_results')
        return samples_df, ast_df, ""
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame(), f"Error loading sheets: {str(e)}"


def validate_samples(df: pd.DataFrame) -> Tuple[bool, List[str]]:
    """Validate samples dataframe."""
    errors = []

    # Check required columns
    missing_cols = REQUIRED_SAMPLES_COLUMNS - set(df.columns)
    if missing_cols:
        errors.append(f"Missing columns in samples: {', '.join(sorted(missing_cols))}")

    # Strip whitespace from all columns
    df.columns = df.columns.str.strip()
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].astype(str).str.strip()

    # Check for required column presence after strip
    existing_cols = set(df.columns)
    if not REQUIRED_SAMPLES_COLUMNS.issubset(existing_cols):
        errors.append(f"Missing required columns")
        return False, errors

    # Check for duplicates
    if df['sample_id'].duplicated().any():
        dup_ids = df[df['sample_id'].duplicated()]['sample_id'].unique()
        errors.append(f"Duplicate sample_id: {', '.join(dup_ids.astype(str))}")

    # Validate source_category
    invalid_source = df[~df['source_category'].isin(VALID_SOURCE_CATEGORIES)]['source_category'].unique()
    if len(invalid_source) > 0:
        errors.append(f"Invalid source_category (must be one of ENVIRONMENT, FOOD, HUMAN, ANIMAL, AQUACULTURE): {', '.join(invalid_source)}")

    # Validate dates
    for idx, row in df.iterrows():
        try:
            if pd.notna(row['collection_date']):
                pd.to_datetime(row['collection_date'], format='%Y-%m-%d')
        except:
            errors.append(f"Invalid date format in row {idx + 2}: {row['collection_date']} (use YYYY-MM-DD)")
            break

    # Check coordinates if provided
    for idx, row in df.iterrows():
        if pd.notna(row['latitude']) and pd.notna(row['longitude']):
            try:
                lat = float(row['latitude'])
                lon = float(row['longitude'])
                if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                    errors.append(f"Invalid coordinates in row {idx + 2}")
            except:
                errors.append(f"Coordinates must be numeric in row {idx + 2}")
                break

    # Check for empty required fields
    for col in ['sample_id', 'region', 'district']:
        if df[col].isna().any() or (df[col].astype(str).str.strip() == '').any():
            errors.append(f"Missing values in required column: {col}")
            break

    return len(errors) == 0, errors


def validate_ast_results(df: pd.DataFrame, sample_ids: set) -> Tuple[bool, List[str]]:
    """Validate AST results dataframe."""
    errors = []

    # Check required columns
    missing_cols = REQUIRED_AST_COLUMNS - set(df.columns)
    if missing_cols:
        errors.append(f"Missing columns in ast_results: {', '.join(sorted(missing_cols))}")

    # Strip whitespace
    df.columns = df.columns.str.strip()
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].astype(str).str.strip()

    # Check for required column presence after strip
    existing_cols = set(df.columns)
    if not REQUIRED_AST_COLUMNS.issubset(existing_cols):
        errors.append(f"Missing required columns in ast_results")
        return False, errors

    # Check for duplicates in isolate_id + antibiotic
    dup_check = df.groupby(['isolate_id', 'antibiotic']).size()
    if (dup_check > 1).any():
        dups = dup_check[dup_check > 1].index.tolist()
        errors.append(f"Duplicate isolate_id + antibiotic combination: {dups[:5]}")

    # Validate result values
    invalid_results = df[~df['result'].isin(VALID_RESULTS)]['result'].unique()
    if len(invalid_results) > 0:
        errors.append(f"Invalid result values (must be S, I, or R): {', '.join(invalid_results)}")

    # Validate method values
    invalid_methods = df[~df['method'].isin(VALID_METHODS)]['method'].unique()
    if len(invalid_methods) > 0:
        errors.append(f"Invalid method values (must be DD or MIC): {', '.join(invalid_methods)}")

    # Validate guideline values
    invalid_guidelines = df[~df['guideline'].isin(VALID_GUIDELINES)]['guideline'].unique()
    if len(invalid_guidelines) > 0:
        errors.append(f"Invalid guideline values (must be CLSI or EUCAST): {', '.join(invalid_guidelines)}")

    # Validate dates
    for idx, row in df.iterrows():
        try:
            if pd.notna(row['test_date']):
                pd.to_datetime(row['test_date'], format='%Y-%m-%d')
        except:
            errors.append(f"Invalid date format in ast_results row {idx + 2}: {row['test_date']}")
            break

    # Check sample_id references
    orphan_samples = set(df['sample_id'].unique()) - sample_ids
    if orphan_samples:
        errors.append(f"AST results reference non-existent sample_id: {list(orphan_samples)[:5]}")

    # Check for empty required fields
    for col in ['sample_id', 'isolate_id', 'organism', 'antibiotic']:
        if df[col].isna().any() or (df[col].astype(str).str.strip() == '').any():
            errors.append(f"Missing values in required column: {col}")
            break

    # Validate mic_value is numeric
    for idx, row in df.iterrows():
        if pd.notna(row['mic_value']):
            try:
                float(row['mic_value'])
            except:
                errors.append(f"mic_value must be numeric in row {idx + 2}")
                break

    # Validate zone_diameter is numeric
    for idx, row in df.iterrows():
        if pd.notna(row['zone_diameter']):
            try:
                float(row['zone_diameter'])
            except:
                errors.append(f"zone_diameter must be numeric in row {idx + 2}")
                break

    return len(errors) == 0, errors


def perform_automated_interpretation(ast_df: pd.DataFrame) -> pd.DataFrame:
    """Perform automated interpretation of AST results using breakpoint database."""
    if ast_df.empty:
        return ast_df

    # Create a copy to avoid modifying original
    interpreted_df = ast_df.copy()

    # Add interpretation columns
    interpreted_df['auto_interpreted'] = False
    interpreted_df['interpreted_result'] = None
    interpreted_df['interpretation_guideline'] = None
    interpreted_df['interpretation_confidence'] = None
    interpreted_df['suspected_mechanism'] = None
    interpreted_df['interpretation_notes'] = None

    # Process each row
    for idx, row in interpreted_df.iterrows():
        # Skip if result is already provided and valid
        if pd.notna(row['result']) and row['result'] in VALID_RESULTS:
            continue

        # Skip if no test values available
        if row['method'] == 'MIC' and pd.isna(row['mic_value']):
            continue
        elif row['method'] == 'DD' and pd.isna(row['zone_diameter']):
            continue

        # Determine guideline to use (default to CLSI if not specified)
        guideline = row['guideline'] if pd.notna(row['guideline']) else 'CLSI'

        try:
            # Perform interpretation
            interpretation = interpret_ast_result(
                organism=row['organism'],
                antibiotic=row['antibiotic'],
                method=row['method'],
                mic_value=row['mic_value'],
                zone_diameter=row['zone_diameter'],
                guideline=guideline
            )

            # Update the row
            interpreted_df.at[idx, 'auto_interpreted'] = True
            interpreted_df.at[idx, 'result'] = interpretation['interpretation']
            interpreted_df.at[idx, 'interpreted_result'] = interpretation['interpretation']
            interpreted_df.at[idx, 'interpretation_guideline'] = interpretation['guideline']
            interpreted_df.at[idx, 'interpretation_confidence'] = interpretation['confidence']
            interpreted_df.at[idx, 'suspected_mechanism'] = interpretation['suspected_mechanism']
            interpreted_df.at[idx, 'interpretation_notes'] = interpretation['notes']

        except Exception as e:
            # Log interpretation errors but don't fail validation
            interpreted_df.at[idx, 'interpretation_notes'] = f"Interpretation failed: {str(e)}"

    return interpreted_df


def validate_upload(file_obj) -> Tuple[bool, List[str], pd.DataFrame, pd.DataFrame]:
    """
    Complete validation pipeline.
    Returns: (is_valid, error_list, samples_df, ast_df)
    """
    errors = []

    # Validate structure
    valid_struct, msg, _ = validate_excel_structure(file_obj)
    if not valid_struct:
        return False, [msg], pd.DataFrame(), pd.DataFrame()

    # Load sheets
    file_obj.seek(0)  # Reset file pointer
    samples_df, ast_df, load_error = load_excel_sheets(file_obj)
    if load_error:
        return False, [load_error], pd.DataFrame(), pd.DataFrame()

    if samples_df.empty or ast_df.empty:
        return False, ["One or both sheets are empty"], pd.DataFrame(), pd.DataFrame()

    # Validate samples
    samples_valid, sample_errors = validate_samples(samples_df)
    errors.extend(sample_errors)

    # Validate AST results (need sample_ids from samples)
    sample_ids = set(samples_df['sample_id'].dropna().astype(str).str.strip())
    ast_valid, ast_errors = validate_ast_results(ast_df, sample_ids)
    errors.extend(ast_errors)

    # Perform automated interpretation if validation passed
    if len(errors) == 0:
        ast_df = perform_automated_interpretation(ast_df)

    return len(errors) == 0, errors, samples_df, ast_df


def create_template_excel() -> bytes:
    """Create a template Excel file."""
    samples_data = {
        'sample_id': ['SAMPLE_001', 'SAMPLE_002', 'SAMPLE_003'],
        'collection_date': ['2024-01-15', '2024-01-20', '2024-01-25'],
        'region': ['Ashanti', 'Greater Accra', 'Eastern'],
        'district': ['Kumasi', 'Accra', 'Koforidua'],
        'site_type': ['Water Treatment Plant', 'Retail Market', 'Hospital Lab'],
        'source_category': ['ENVIRONMENT', 'FOOD', 'HUMAN'],
        'source_type': ['treated_water', 'raw_chicken', 'clinical_specimen'],
        'food_matrix': ['', 'chicken', ''],
        'environment_matrix': ['treated_water', '', ''],
        'latitude': [6.6326, 5.6037, 6.1256],
        'longitude': [-1.6243, -0.1870, -0.3597]
    }

    ast_data = {
        'sample_id': ['SAMPLE_001', 'SAMPLE_001', 'SAMPLE_002'],
        'isolate_id': ['ISO_001', 'ISO_002', 'ISO_003'],
        'organism': ['E. coli', 'E. coli', 'Salmonella'],
        'antibiotic': ['Ampicillin', 'Ciprofloxacin', 'Ampicillin'],
        'result': ['R', 'S', 'I'],
        'method': ['DD', 'DD', 'MIC'],
        'guideline': ['CLSI', 'EUCAST', 'CLSI'],
        'test_date': ['2024-01-20', '2024-01-20', '2024-01-22'],
        'mic_value': [np.nan, np.nan, 0.5],
        'zone_diameter': [15.0, 28.0, np.nan]
    }

    with pd.ExcelWriter('templates/AMR_ENV_FOOD_template_v1.xlsx', engine='openpyxl') as writer:
        pd.DataFrame(samples_data).to_excel(writer, sheet_name='samples', index=False)
        pd.DataFrame(ast_data).to_excel(writer, sheet_name='ast_results', index=False)

    # Read back as bytes
    with open('templates/AMR_ENV_FOOD_template_v1.xlsx', 'rb') as f:
        return f.read()
